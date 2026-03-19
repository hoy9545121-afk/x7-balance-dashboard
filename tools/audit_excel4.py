"""
성장_경제 / 몬스터_카드 / 장비_제작 시트 감사
"""
import openpyxl

BASE = r'C:\AI_simulator\기획서'

def dump_rows(path, sheet, r_from, r_to):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"  ⚠ 시트 없음: {sheet}")
        wb.close()
        return
    ws = wb[sheet]
    print(f"\n--- [{path.split(chr(92))[-1]}] {sheet} R{r_from}~R{r_to} ---")
    for r in range(r_from, r_to+1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 14)]
        if any(v is not None for v in vals):
            print(f"  R{r}: {[v for v in vals if v is not None]}")
    wb.close()

GROWTH = fr'{BASE}\X7_성장_경제_밸런스_통합기획서.xlsx'
CARD   = fr'{BASE}\X7_몬스터_카드_아이템_목록.xlsx'

# 성장 경제: 수치 요약 시트 (사냥 속도 관련 내용 확인)
dump_rows(GROWTH, '📊 캐릭터 레벨', 1, 30)
dump_rows(GROWTH, '🏆 숙련도', 1, 30)
dump_rows(GROWTH, '📈 수치 요약', 1, 50)

# 몬스터 카드: 몬스터 등급 관련
dump_rows(CARD, '📋 전체 카드 목록', 1, 30)
