"""
X7 스킬 밸런스 Excel 생성기 v2
모든 48개 스킬 슬롯 (4무기 × 3티어 × 4커맨드) 포함
무기별 시트 (한손검/양손검/지팡이/단검) + 스킬 리스트 시트
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, RadarChart, Reference

# ── 색상 팔레트 ──────────────────────────────────────────
CLR = {
    "title_bg":   "1F2D3D",
    "title_fg":   "E8B84B",
    "header_bg":  "2C3E50",
    "header_fg":  "FFFFFF",
    "tier_bg":    "34495E",
    "tier_fg":    "ECF0F1",
    "q_bg":       "D5F5E3",
    "w_bg":       "D6EAF8",
    "e_bg":       "F9EBEA",
    "r_bg":       "FEF9E7",
    "q_partial":  "EAFAF1",  # 편집 필요 (연한 버전)
    "w_partial":  "EBF5FB",
    "e_partial":  "FDEDEC",
    "r_partial":  "FEFDE7",
    "no_data":    "F2F3F4",  # 데이터 없음
    "section_bg": "EBF5FB",
    "formula_bg": "FDFEFE",
    "warn_bg":    "FADBD8",
    "heal_bg":    "E8F8F5",
}

CMD_BG       = {"Q": "q_bg",      "W": "w_bg",      "E": "e_bg",      "R": "r_bg"}
CMD_PARTIAL  = {"Q": "q_partial", "W": "w_partial", "E": "e_partial", "R": "r_partial"}

# ── 스킬 데이터 ──────────────────────────────────────────
# dmg 단위: 1.0 = 100% / 레벨 미확보는 lv별 동일값 or 0 사용
# data_full: Lv1~5 전 레벨 확보 여부 (True면 분석 섹션 생성)

WEAPON_DATA = {

    # ────────────────────────── 한손검 ──────────────────────────
    "한손검": {
        "concept": "방패 + 반격 스타일의 근접 전사 — 방어와 CC를 활용해 적을 제압",
        "tiers": [
            {
                "tier": "1티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "철벽의 반격", "sk_type": "방향 공격",
                        "desc": "방패를 들어 방어하다가 공격에 반격합니다.",
                        "levels": [
                            {"lv": 1, "cd": 5,  "mp": 30, "dmg": 2.25, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5,  "mp": 30, "dmg": 2.25, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 5,  "mp": 30, "dmg": 2.25, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 5,  "mp": 30, "dmg": 2.25, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 5,  "mp": 30, "dmg": 2.25, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "방패 돌진", "sk_type": "단일 공격",
                        "desc": "방패로 적에게 돌진하여 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 0,  "mp": 0,  "dmg": 2.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 0,  "mp": 0,  "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 0,  "mp": 0,  "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 0,  "mp": 0,  "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 0,  "mp": 0,  "dmg": 2.0,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "방패 투척", "sk_type": "방향 공격",
                        "desc": "방패를 투척하여 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 30, "dmg": 2.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 30, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 15, "mp": 30, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 15, "mp": 30, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 15, "mp": 30, "dmg": 2.0,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "파멸의 징벌", "sk_type": "범위 공격 (위치 지정, max5)",
                        "desc": "지정 위치 반경 300에 충격을 가해 최대 5대상에게 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 45, "mp": 30, "dmg": 4.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 45, "mp": 30, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 45, "mp": 30, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 45, "mp": 30, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 45, "mp": 30, "dmg": 4.0,  "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "2티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "제압", "sk_type": "단일 대상",
                        "desc": "적을 제압하여 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 5,  "mp": 0,  "dmg": 1.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5,  "mp": 0,  "dmg": 1.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 5,  "mp": 0,  "dmg": 1.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 5,  "mp": 0,  "dmg": 1.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 5,  "mp": 0,  "dmg": 1.0, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "복수의 갑옷", "sk_type": "버프",
                        "desc": "복수의 갑옷을 두릅니다. (방어력 강화)",
                        "levels": [
                            {"lv": 1, "cd": 10, "mp": 0,  "dmg": 0,   "buff": "방어력↑", "confirmed": True},
                            {"lv": 2, "cd": 10, "mp": 0,  "dmg": 0,   "buff": "방어력↑", "confirmed": False},
                            {"lv": 3, "cd": 10, "mp": 0,  "dmg": 0,   "buff": "방어력↑", "confirmed": False},
                            {"lv": 4, "cd": 10, "mp": 0,  "dmg": 0,   "buff": "방어력↑", "confirmed": False},
                            {"lv": 5, "cd": 10, "mp": 0,  "dmg": 0,   "buff": "방어력↑", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "모두 쉿", "sk_type": "자기중심 광역 (CC)",
                        "desc": "주변 적에게 침묵을 걸어 스킬 사용을 막습니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 50, "dmg": 0,   "buff": "침묵", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 50, "dmg": 0,   "buff": "침묵", "confirmed": False},
                            {"lv": 3, "cd": 15, "mp": 50, "dmg": 0,   "buff": "침묵", "confirmed": False},
                            {"lv": 4, "cd": 15, "mp": 50, "dmg": 0,   "buff": "침묵", "confirmed": False},
                            {"lv": 5, "cd": 15, "mp": 50, "dmg": 0,   "buff": "침묵", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "진공 폭발", "sk_type": "자기중심 광역",
                        "desc": "강력한 진공 폭발로 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 30, "mp": 0,  "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 30, "mp": 0,  "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 30, "mp": 0,  "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 30, "mp": 0,  "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 30, "mp": 0,  "dmg": 3.0, "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "3티어",
                "data_full": False,
                "note": "⚖️ 균형 추정값 (3티어 양손검 기준, Lv.1)",
                "skills": [
                    {
                        "cmd": "Q", "name": "성스러운 일격", "sk_type": "단일 공격",
                        "desc": "성스러운 빛으로 적을 강타합니다.",
                        "levels": [
                            {"lv": 1, "cd": 5,  "mp": 10, "dmg": 2.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5,  "mp": 10, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 5,  "mp": 10, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 5,  "mp": 10, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 5,  "mp": 10, "dmg": 2.0, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "가호의 빛", "sk_type": "버프",
                        "desc": "빛의 가호로 방어력을 강화합니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 30, "dmg": 0,   "buff": "방호 강화", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 30, "dmg": 0,   "buff": "방호 강화", "confirmed": False},
                            {"lv": 3, "cd": 12, "mp": 30, "dmg": 0,   "buff": "방호 강화", "confirmed": False},
                            {"lv": 4, "cd": 12, "mp": 30, "dmg": 0,   "buff": "방호 강화", "confirmed": False},
                            {"lv": 5, "cd": 12, "mp": 30, "dmg": 0,   "buff": "방호 강화", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "폭주하는 광휘", "sk_type": "범위 공격",
                        "desc": "광휘가 폭주하며 주변 적에게 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "광휘의 심판", "sk_type": "단일 공격",
                        "desc": "강력한 광휘로 심판의 일격을 가합니다.",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
        ],
    },

    # ────────────────────────── 양손검 ──────────────────────────
    "양손검": {
        "concept": "강력한 범위 공격과 버스트를 가진 근접 딜러",
        "tiers": [
            {
                "tier": "1티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "폭주", "sk_type": "공격 / 버프",
                        "desc": "폭주 상태로 강력한 공격을 가합니다.",
                        "levels": [
                            {"lv": 1, "cd": 0,  "mp": 15, "dmg": 3.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 0,  "mp": 15, "dmg": 3.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 0,  "mp": 15, "dmg": 3.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 0,  "mp": 15, "dmg": 3.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 0,  "mp": 15, "dmg": 3.0,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "분노의 일격", "sk_type": "버프",
                        "desc": "분노를 담아 공격력을 강화합니다.",
                        "levels": [
                            {"lv": 1, "cd": 10, "mp": 15, "dmg": 0,    "buff": "공격력 강화 / 5s (수치 미확인)", "buff_dur": 5, "confirmed": True},
                            {"lv": 2, "cd": 10, "mp": 15, "dmg": 0,    "buff": "공격력 강화 / 5s (수치 미확인)", "buff_dur": 5, "confirmed": False},
                            {"lv": 3, "cd": 10, "mp": 15, "dmg": 0,    "buff": "공격력 강화 / 5s (수치 미확인)", "buff_dur": 5, "confirmed": False},
                            {"lv": 4, "cd": 10, "mp": 15, "dmg": 0,    "buff": "공격력 강화 / 5s (수치 미확인)", "buff_dur": 5, "confirmed": False},
                            {"lv": 5, "cd": 10, "mp": 15, "dmg": 0,    "buff": "공격력 강화 / 5s (수치 미확인)", "buff_dur": 5, "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "대지 분쇄", "sk_type": "범위 공격",
                        "desc": "대지를 분쇄하여 주변 적에게 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 0,  "mp": 1,  "dmg": 0.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 0,  "mp": 1,  "dmg": 0.5,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 0,  "mp": 1,  "dmg": 0.5,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 0,  "mp": 1,  "dmg": 0.5,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 0,  "mp": 1,  "dmg": 0.5,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "공간 가르기", "sk_type": "단일 공격",
                        "desc": "공간을 가르는 강력한 일격을 가합니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 30, "dmg": 6.25, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 30, "dmg": 6.25, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 12, "mp": 30, "dmg": 6.25, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 12, "mp": 30, "dmg": 6.25, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 12, "mp": 30, "dmg": 6.25, "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "2티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "몰아치기", "sk_type": "버프",
                        "desc": "연속 공격으로 몰아칩니다.",
                        "levels": [
                            {"lv": 1, "cd": 6,  "mp": 15, "dmg": 0,    "buff": "공속+330%(3300천분률) / 5s", "buff_dur": 5, "buff_atk_speed": 3300, "confirmed": True},
                            {"lv": 2, "cd": 6,  "mp": 15, "dmg": 0,    "buff": "공속+330%(3300천분률) / 5s", "buff_dur": 5, "buff_atk_speed": 3300, "confirmed": False},
                            {"lv": 3, "cd": 6,  "mp": 15, "dmg": 0,    "buff": "공속+330%(3300천분률) / 5s", "buff_dur": 5, "buff_atk_speed": 3300, "confirmed": False},
                            {"lv": 4, "cd": 6,  "mp": 15, "dmg": 0,    "buff": "공속+330%(3300천분률) / 5s", "buff_dur": 5, "buff_atk_speed": 3300, "confirmed": False},
                            {"lv": 5, "cd": 6,  "mp": 15, "dmg": 0,    "buff": "공속+330%(3300천분률) / 5s", "buff_dur": 5, "buff_atk_speed": 3300, "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "찌르기", "sk_type": "단일 공격",
                        "desc": "검으로 적을 찌릅니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "결투 돌입", "sk_type": "단일 공격",
                        "desc": "결투에 돌입하며 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 0,  "mp": 15, "dmg": 1.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 0,  "mp": 15, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 0,  "mp": 15, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 0,  "mp": 15, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 0,  "mp": 15, "dmg": 1.5,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "섬광의 길", "sk_type": "방향 범위 공격 (Box300×800, max5)",
                        "desc": "전방 방향으로 섬광 가르기를 가해 직선 범위(300×800) 최대 5대상에게 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 8,  "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 8,  "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 8,  "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 8,  "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 8,  "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "3티어",
                "data_full": True,
                "note": "✅ 전 레벨 데이터 확보 (시뮬레이터 구현)",
                "skills": [
                    {
                        "cmd": "Q", "name": "가벼운 손놀림", "sk_type": "자기중심 광역 (평타 강화)",
                        "desc": "손놀림을 가볍게 하여 다음 기본 공격을 강화합니다.",
                        "levels": [
                            {"lv": 1, "cd": 5, "mp": 10, "dmg": 2.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5, "mp": 10, "dmg": 2.2, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 5, "mp": 10, "dmg": 2.4, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 5, "mp": 10, "dmg": 2.6, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 5, "mp": 10, "dmg": 2.8, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "W", "name": "대지 가르기", "sk_type": "방향 공격",
                        "desc": "검으로 땅을 내리쳐 충격파로 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 30, "dmg": 2.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 35, "dmg": 2.75, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 12, "mp": 40, "dmg": 3.0,  "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 12, "mp": 45, "dmg": 3.25, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 12, "mp": 50, "dmg": 3.5,  "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "E", "name": "휘몰이", "sk_type": "범위 공격 (6-hit 합산)",
                        "desc": "주변을 검으로 6회 연속 베어 피해를 입힙니다. (50%×6=300% Lv1)",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 20, "dmg": 3.3, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 15, "mp": 25, "dmg": 3.6, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 15, "mp": 30, "dmg": 3.9, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 15, "mp": 35, "dmg": 4.2, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "R", "name": "적진으로", "sk_type": "돌진 + 착지 범위 공격 (max100)",
                        "desc": "지정 위치로 돌진 후 착지 반경 300 내 최대 100대상에게 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 50, "dmg": 4.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 38, "mp": 50, "dmg": 4.4, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 36, "mp": 50, "dmg": 4.8, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 34, "mp": 50, "dmg": 5.2, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 32, "mp": 50, "dmg": 5.6, "buff": "", "confirmed": True},
                        ],
                    },
                ],
            },
        ],
    },

    # ────────────────────────── 지팡이 ──────────────────────────
    "지팡이": {
        "concept": "DoT·다중 타격·힐·CC를 갖춘 다재다능한 마법사",
        "tiers": [
            {
                "tier": "1티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "화염 장막", "sk_type": "DoT 장판 (6틱 합산)",
                        "desc": "화염 장막 소환. life_time=2.9s, Interval=0.5s, 6틱×36%=216% (Lv1)",
                        "levels": [
                            {"lv": 1, "cd": 8,  "mp": 30, "dmg": 2.16, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 8,  "mp": 30, "dmg": 2.16, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 8,  "mp": 30, "dmg": 2.16, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 8,  "mp": 30, "dmg": 2.16, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 8,  "mp": 30, "dmg": 2.16, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "불의 비", "sk_type": "범위 공격",
                        "desc": "화염 비를 내려 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 50, "dmg": 4.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 50, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 15, "mp": 50, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 15, "mp": 50, "dmg": 4.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 15, "mp": 50, "dmg": 4.0,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "화염 폭발", "sk_type": "단일 공격",
                        "desc": "화염 폭발로 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 20, "mp": 20, "dmg": 2.0,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 20, "mp": 20, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 20, "mp": 20, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 20, "mp": 20, "dmg": 2.0,  "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 20, "mp": 20, "dmg": 2.0,  "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "별똥별", "sk_type": "범위 공격 (위치 지정, max8)",
                        "desc": "지정 위치에 별똥별을 소환하여 반경 300 내 최대 8대상에게 강력한 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "2티어",
                "data_full": True,
                "note": "✅ 전 레벨 데이터 확보",
                "skills": [
                    {
                        "cmd": "Q", "name": "신비한 가시 나무", "sk_type": "DoT 장판 (6틱 합산)",
                        "desc": "가시 나무 DoT. life_time=2.9s, Interval=0.5s, 6틱 합산. 36/42/48/54/60%×6",
                        "levels": [
                            {"lv": 1, "cd": 8, "mp": 20, "dmg": 2.16, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 8, "mp": 22, "dmg": 2.52, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 8, "mp": 24, "dmg": 2.88, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 8, "mp": 26, "dmg": 3.24, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 8, "mp": 28, "dmg": 3.60, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "W", "name": "가시 나무 덩쿨", "sk_type": "범위 공격 + CC",
                        "desc": "덩쿨로 적을 잡아 피해를 입히고 부유 상태로 만듭니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 40,  "dmg": 2.5, "buff": "부유", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 44,  "dmg": 2.7, "buff": "부유", "confirmed": True},
                            {"lv": 3, "cd": 12, "mp": 48,  "dmg": 2.9, "buff": "부유", "confirmed": True},
                            {"lv": 4, "cd": 12, "mp": 52,  "dmg": 3.1, "buff": "부유", "confirmed": True},
                            {"lv": 5, "cd": 12, "mp": 56,  "dmg": 3.3, "buff": "부유", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "E", "name": "치유의 꽃잎", "sk_type": "힐",
                        "desc": "꽃잎으로 자신의 HP를 회복합니다. (피해량=0)",
                        "levels": [
                            {"lv": 1, "cd": 1,  "mp": 10,  "dmg": 0, "buff": "힐", "confirmed": True},
                            {"lv": 2, "cd": 1,  "mp": 11,  "dmg": 0, "buff": "힐", "confirmed": True},
                            {"lv": 3, "cd": 1,  "mp": 12,  "dmg": 0, "buff": "힐", "confirmed": True},
                            {"lv": 4, "cd": 1,  "mp": 13,  "dmg": 0, "buff": "힐", "confirmed": True},
                            {"lv": 5, "cd": 1,  "mp": 14,  "dmg": 0, "buff": "힐", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "R", "name": "치유의 바람", "sk_type": "힐 DoT",
                        "desc": "치유의 바람이 지속적으로 HP를 회복합니다. (피해량=0)",
                        "levels": [
                            {"lv": 1, "cd": 30, "mp": 100, "dmg": 0, "buff": "힐DoT", "confirmed": True},
                            {"lv": 2, "cd": 30, "mp": 110, "dmg": 0, "buff": "힐DoT", "confirmed": True},
                            {"lv": 3, "cd": 30, "mp": 120, "dmg": 0, "buff": "힐DoT", "confirmed": True},
                            {"lv": 4, "cd": 30, "mp": 130, "dmg": 0, "buff": "힐DoT", "confirmed": True},
                            {"lv": 5, "cd": 30, "mp": 140, "dmg": 0, "buff": "힐DoT", "confirmed": True},
                        ],
                    },
                ],
            },
            {
                "tier": "3티어",
                "data_full": True,
                "note": "✅ 전 레벨 데이터 확보 (시뮬레이터 구현)",
                "skills": [
                    {
                        "cmd": "Q", "name": "얼음 화살", "sk_type": "방향 투사체",
                        "desc": "차가운 얼음 화살을 발사하여 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 3,  "mp": 30,  "dmg": 2.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 3,  "mp": 33,  "dmg": 2.2, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 3,  "mp": 36,  "dmg": 2.4, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 3,  "mp": 39,  "dmg": 2.6, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 3,  "mp": 42,  "dmg": 2.8, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "W", "name": "다후타의 손짓", "sk_type": "DoT 장판 (5틱 합산)",
                        "desc": "얼음 폭풍 장판. life_time=4.9s, Interval=1s, 5틱×80%=400% (Lv1)",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 50,  "dmg": 4.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 55,  "dmg": 4.4, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 15, "mp": 60,  "dmg": 4.8, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 15, "mp": 65,  "dmg": 5.2, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 15, "mp": 70,  "dmg": 5.6, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "E", "name": "심판의 낙뢰", "sk_type": "위치 지정 (2-hit)",
                        "desc": "강력한 낙뢰 2회 피해. (400%+200%=600% Lv1)",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 100, "dmg": 6.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 38, "mp": 110, "dmg": 6.6, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 36, "mp": 120, "dmg": 7.2, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 34, "mp": 130, "dmg": 7.8, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 32, "mp": 140, "dmg": 8.4, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "R", "name": "방울방울", "sk_type": "CC (부유)",
                        "desc": "적을 부유 상태로 만드는 CC 스킬. 피해량 없음.",
                        "levels": [
                            {"lv": i, "cd": 20, "mp": 50, "dmg": 0.0, "buff": "부유 8s", "confirmed": True}
                            for i in range(1, 6)
                        ],
                    },
                ],
            },
        ],
    },

    # ────────────────────────── 단검 ──────────────────────────
    "단검": {
        "concept": "빠른 속도와 높은 치명타로 적을 순식간에 처치하는 민첩한 딜러",
        "tiers": [
            {
                "tier": "1티어",
                "data_full": False,
                "note": "Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
                "skills": [
                    {
                        "cmd": "Q", "name": "질풍 가르기", "sk_type": "방향 공격",
                        "desc": "빠르게 베어 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 5,  "mp": 30,  "dmg": 1.5, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5,  "mp": 30,  "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 5,  "mp": 30,  "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 5,  "mp": 30,  "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 5,  "mp": 30,  "dmg": 1.5, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "무기 연마", "sk_type": "버프",
                        "desc": "무기를 연마하여 공격 속도를 높입니다.",
                        "levels": [
                            {"lv": 1, "cd": 10, "mp": 15, "dmg": 0,   "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                            {"lv": 2, "cd": 10, "mp": 15, "dmg": 0,   "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": False},
                            {"lv": 3, "cd": 10, "mp": 15, "dmg": 0,   "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": False},
                            {"lv": 4, "cd": 10, "mp": 15, "dmg": 0,   "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": False},
                            {"lv": 5, "cd": 10, "mp": 15, "dmg": 0,   "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "어둠의 일격", "sk_type": "단일 공격",
                        "desc": "어둠의 힘으로 강력한 일격을 가합니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "암살자의 분노", "sk_type": "버프 (평타→광역)",
                        "desc": "버프 활성 시 평타가 범위200 최대5타 광역(420%)으로 전환. CriVaryper/CriDmg/공속/이속 대폭 강화. 지속15s",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 100, "dmg": 0,  "buff": "치명률+300|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 300, "buff_cri_dmg": 500, "confirmed": True},
                            {"lv": 2, "cd": 40, "mp": 100, "dmg": 0,  "buff": "치명률+330|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 330, "buff_cri_dmg": 500, "confirmed": False},
                            {"lv": 3, "cd": 40, "mp": 100, "dmg": 0,  "buff": "치명률+360|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 360, "buff_cri_dmg": 500, "confirmed": False},
                            {"lv": 4, "cd": 40, "mp": 100, "dmg": 0,  "buff": "치명률+390|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 390, "buff_cri_dmg": 500, "confirmed": False},
                            {"lv": 5, "cd": 40, "mp": 100, "dmg": 0,  "buff": "치명률+420|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 420, "buff_cri_dmg": 500, "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "2티어",
                "data_full": False,
                "note": "⚖️ 일부 균형 추정값 적용 (그림자 매·진격·전율하는 매, Lv.1)",
                "skills": [
                    {
                        "cmd": "Q", "name": "그림자 매", "sk_type": "단일 공격",
                        "desc": "그림자처럼 빠르게 적을 베어 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 5,  "mp": 20, "dmg": 1.5, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5,  "mp": 20, "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 5,  "mp": 20, "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 5,  "mp": 20, "dmg": 1.5, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 5,  "mp": 20, "dmg": 1.5, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "W", "name": "쌍수 달인", "sk_type": "버프",
                        "desc": "쌍수 기술로 공격 속도를 강화합니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 20, "dmg": 0, "buff": "공속+35%(350천분률) / 15s", "buff_dur": 15, "buff_atk_speed": 350, "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 25, "dmg": 0, "buff": "공속+35%(350천분률) / 15s", "buff_dur": 15, "buff_atk_speed": 350, "confirmed": False},
                            {"lv": 3, "cd": 12, "mp": 30, "dmg": 0, "buff": "공속+35%(350천분률) / 15s", "buff_dur": 15, "buff_atk_speed": 350, "confirmed": False},
                            {"lv": 4, "cd": 12, "mp": 35, "dmg": 0, "buff": "공속+35%(350천분률) / 15s", "buff_dur": 15, "buff_atk_speed": 350, "confirmed": False},
                            {"lv": 5, "cd": 12, "mp": 40, "dmg": 0, "buff": "공속+35%(350천분률) / 15s", "buff_dur": 15, "buff_atk_speed": 350, "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "E", "name": "진격", "sk_type": "단일 공격",
                        "desc": "적에게 빠르게 돌진하며 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 12, "mp": 25, "dmg": 2.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 12, "mp": 25, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 12, "mp": 25, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 12, "mp": 25, "dmg": 2.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 12, "mp": 25, "dmg": 2.0, "buff": "", "confirmed": False},
                        ],
                    },
                    {
                        "cmd": "R", "name": "전율하는 매", "sk_type": "단일 공격",
                        "desc": "강력한 일격으로 적을 전율하게 만듭니다.",
                        "levels": [
                            {"lv": 1, "cd": 30, "mp": 60, "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 30, "mp": 60, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 3, "cd": 30, "mp": 60, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 4, "cd": 30, "mp": 60, "dmg": 3.0, "buff": "", "confirmed": False},
                            {"lv": 5, "cd": 30, "mp": 60, "dmg": 3.0, "buff": "", "confirmed": False},
                        ],
                    },
                ],
            },
            {
                "tier": "3티어",
                "data_full": True,
                "note": "✅ 전 레벨 데이터 확보",
                "skills": [
                    {
                        "cmd": "Q", "name": "질풍 가르기", "sk_type": "방향 공격",
                        "desc": "빠르게 베어 피해를 입힙니다.",
                        "levels": [
                            {"lv": 1, "cd": 5, "mp": 30, "dmg": 1.5,  "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 5, "mp": 30, "dmg": 1.65, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 5, "mp": 30, "dmg": 1.8,  "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 5, "mp": 30, "dmg": 1.95, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 5, "mp": 30, "dmg": 2.1,  "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "W", "name": "무기 연마", "sk_type": "버프",
                        "desc": "무기를 연마하여 공격 속도를 높입니다.",
                        "levels": [
                            {"lv": 1, "cd": 10, "mp": 15, "dmg": 0, "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                            {"lv": 2, "cd": 10, "mp": 17, "dmg": 0, "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                            {"lv": 3, "cd": 10, "mp": 19, "dmg": 0, "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                            {"lv": 4, "cd": 10, "mp": 21, "dmg": 0, "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                            {"lv": 5, "cd": 10, "mp": 23, "dmg": 0, "buff": "다음 평타 강화(3s)", "buff_dur": 3, "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "E", "name": "어둠의 일격", "sk_type": "단일 공격",
                        "desc": "어둠의 힘으로 강력한 일격을 가합니다.",
                        "levels": [
                            {"lv": 1, "cd": 15, "mp": 15, "dmg": 3.0, "buff": "", "confirmed": True},
                            {"lv": 2, "cd": 15, "mp": 17, "dmg": 3.3, "buff": "", "confirmed": True},
                            {"lv": 3, "cd": 15, "mp": 19, "dmg": 3.6, "buff": "", "confirmed": True},
                            {"lv": 4, "cd": 15, "mp": 21, "dmg": 3.9, "buff": "", "confirmed": True},
                            {"lv": 5, "cd": 15, "mp": 23, "dmg": 4.2, "buff": "", "confirmed": True},
                        ],
                    },
                    {
                        "cmd": "R", "name": "암살자의 분노", "sk_type": "버프 (평타→광역)",
                        "desc": "버프 활성 시 평타가 범위200 최대5타 광역(420%)으로 전환. CriVaryper/CriDmg/공속/이속 대폭 강화. 지속15s",
                        "levels": [
                            {"lv": 1, "cd": 40, "mp": 100, "dmg": 0, "buff": "치명률+300|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 300, "buff_cri_dmg": 500, "confirmed": True},
                            {"lv": 2, "cd": 38, "mp": 100, "dmg": 0, "buff": "치명률+330|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 330, "buff_cri_dmg": 500, "confirmed": True},
                            {"lv": 3, "cd": 36, "mp": 100, "dmg": 0, "buff": "치명률+360|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 360, "buff_cri_dmg": 500, "confirmed": True},
                            {"lv": 4, "cd": 34, "mp": 100, "dmg": 0, "buff": "치명률+390|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 390, "buff_cri_dmg": 500, "confirmed": True},
                            {"lv": 5, "cd": 32, "mp": 100, "dmg": 0, "buff": "치명률+420|치명피해+500|공속+500|이속+500 / 평타→광역420%×5 / 지속15s", "buff_dur": 15, "buff_atk_speed": 500, "buff_cri": 420, "buff_cri_dmg": 500, "confirmed": True},
                        ],
                    },
                ],
            },
        ],
    },
}


# ── 스타일 헬퍼 ─────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN  = Side(style="thin",   color="AAAAAA")
THICK = Side(style="medium", color="666666")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_cell(cell, bg=None, bold=False, size=10, fg="000000",
               h_align="center", v_align="center", wrap=False, italic=False):
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = align(h=h_align, v=v_align, wrap=wrap)
    cell.border = thin_border()


# ── 티어 섹션 쓰기 ───────────────────────────────────────
def write_tier_section(ws, tier_name, tier_data, start_row):
    """
    하나의 티어 섹션을 작성하고 다음 빈 행 번호를 반환.
    returns (next_row, skill_rows_dict)
      skill_rows_dict: {cmd -> (first_row, last_row)}  — 분석용
    """
    r = start_row
    skills    = tier_data["skills"]
    data_full = tier_data["data_full"]
    note_text = tier_data.get("note", "")

    # ── 티어 구분선 ──
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:M{r}")
    c = ws[f"B{r}"]
    c.value = f"▶ {tier_name}  |  {note_text}"
    style_cell(c, bg=CLR["tier_bg"], bold=True, size=12, fg=CLR["tier_fg"], h_align="left")
    r += 1

    # ── 스킬 기본 설명 ──
    hdr_desc = ["커맨드", "스킬명", "유형", "설명"]
    hdr_cols = ["B", "C", "D", "E"]
    for col, h in zip(hdr_cols, hdr_desc):
        c = ws[f"{col}{r}"]
        c.value = h
        style_cell(c, bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"])
    ws.merge_cells(f"E{r}:M{r}")
    r += 1

    for sk in skills:
        cmd = sk["cmd"]
        bg  = CLR[CMD_BG[cmd]]
        ws[f"B{r}"].value = cmd
        ws[f"C{r}"].value = sk["name"]
        ws[f"D{r}"].value = sk["sk_type"]
        ws.merge_cells(f"E{r}:M{r}")
        ws[f"E{r}"].value = sk["desc"]
        for col in ["B", "C", "D"]:
            style_cell(ws[f"{col}{r}"], bg=bg, h_align="center" if col=="B" else "left")
        style_cell(ws[f"E{r}"], bg=bg, h_align="left", wrap=True, size=9)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1  # 빈 행

    # ── 스킬 강화 정보 테이블 ──
    ws.merge_cells(f"B{r}:M{r}")
    ws[f"B{r}"].value = f"⚔️ 스킬 강화 정보 — {tier_name}"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, size=11, fg=CLR["title_fg"], h_align="left")
    r += 1

    headers = ["커맨드", "레벨", "쿨타임(s)", "마나소모", "공격유형",
               "피해량(배율)", "버프/디버프", "DPS", "분당횟수", "1분누적피해", "1분마나소모"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        c = ws[f"{col}{r}"]
        c.value = h
        style_cell(c, bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"], wrap=True)
        ws.row_dimensions[r].height = 30
    r += 1

    skill_rows = {}  # cmd -> (first_row, last_row)

    for sk in skills:
        cmd   = sk["cmd"]
        first = r
        for lv_data in sk["levels"]:
            lv        = lv_data["lv"]
            confirmed = lv_data.get("confirmed", True)
            bg_key    = CMD_BG[cmd] if confirmed else CMD_PARTIAL[cmd]
            bg        = CLR[bg_key]

            cd  = lv_data["cd"]
            mp  = lv_data["mp"]
            dmg = lv_data["dmg"]

            row_vals = [
                cmd if lv == 1 else "",
                lv,
                cd if cd > 0 else "-",
                mp,
                sk["sk_type"],
                dmg,
                lv_data.get("buff", ""),
            ]
            for i, val in enumerate(row_vals):
                col = get_column_letter(2 + i)
                c = ws[f"{col}{r}"]
                c.value = val
                h = "left" if i == 0 else "center"
                style_cell(c, bg=bg, h_align=h)
                if not confirmed:
                    c.font = Font(bold=False, size=10, color="999999", italic=True)

            # DPS = 피해량 / 쿨타임  (버프는 업타임 표시)
            c_dps = ws[f"I{r}"]
            buff_dur = lv_data.get("buff_dur", 0)
            if cd > 0 and dmg > 0:
                c_dps.value = f"=G{r}/D{r}"
                c_dps.number_format = "0.00"
            elif buff_dur > 0 and cd > 0:
                uptime = min(buff_dur / cd, 1.0)
                if uptime >= 1.0:
                    c_dps.value = "상시 유지"
                else:
                    c_dps.value = f"업타임 {uptime:.0%}"
            else:
                c_dps.value = "-" if dmg == 0 else "0"
            style_cell(c_dps, bg=bg)

            # 분당 횟수
            c_cnt = ws[f"J{r}"]
            if cd > 0:
                c_cnt.value = f"=60/D{r}"
                c_cnt.number_format = "0.0"
            else:
                c_cnt.value = "-"
            style_cell(c_cnt, bg=bg)

            # 1분 누적 피해
            c_acc = ws[f"K{r}"]
            if cd > 0 and dmg > 0:
                c_acc.value = f"=J{r}*G{r}"
                c_acc.number_format = "0.00"
            else:
                c_acc.value = "0"
            style_cell(c_acc, bg=bg)

            # 1분 마나
            c_mp = ws[f"L{r}"]
            if cd > 0 and mp > 0:
                c_mp.value = f"=J{r}*E{r}"
                c_mp.number_format = "0"
            else:
                c_mp.value = "0"
            style_cell(c_mp, bg=bg)

            r += 1

        skill_rows[cmd] = (first, r - 1)

    r += 1  # 빈 행

    # ── 분석 섹션 (data_full=True 일 때만) ──
    if data_full:
        r = _write_analysis(ws, tier_name, skills, skill_rows, r)

    # 주석
    ws.merge_cells(f"B{r}:M{r}")
    ws[f"B{r}"].value = (
        "※ 피해량 배율 1.0 = ATK×100%  |  노란 셀 = Lv.1 확인 데이터, 연한 셀 = 편집 필요  "
        "|  D(쿨타임)·E(마나)·G(피해량) 직접 수정 → 오른쪽 분석 자동 갱신"
    )
    style_cell(ws[f"B{r}"], bg="FEF9E7", fg="7D6608", h_align="left", size=9, italic=True)
    r += 2

    return r, skill_rows


def _add_charts(ws, tier_name, cmds, dps_header_row, dps_data_start, chart_row):
    """
    차트 1: 스킬별 레벨 성장 꺾은선 (Q/W/E/R 각각 선 + 마커)
    차트 2: 레벨별 스킬 피해 클러스터 막대 (나란히 비교)
    """
    n = len(cmds)
    colors   = {"Q": "43A047", "W": "1565C0", "E": "C62828", "R": "E65100"}
    markers  = {"Q": "circle", "W": "diamond", "E": "triangle", "R": "square"}
    cats     = Reference(ws, min_col=2,
                         min_row=dps_data_start, max_row=dps_data_start + 4)

    # ── 차트 1: 스킬별 레벨 성장 꺾은선 ──────────────────────
    line = LineChart()
    line.title           = f"{tier_name}  스킬별 레벨 성장 (1분 누적 피해)"
    line.y_axis.title    = "1분 누적 피해 (배율)"
    line.x_axis.title    = "스킬 레벨"
    line.style           = 10
    line.smooth          = True
    line.width           = 17
    line.height          = 13

    for j, cmd in enumerate(cmds):
        ref = Reference(ws, min_col=3 + j, max_col=3 + j,
                        min_row=dps_header_row, max_row=dps_data_start + 4)
        line.add_data(ref, titles_from_data=True)

    line.set_categories(cats)

    for i, cmd in enumerate(cmds):
        if i >= len(line.series):
            break
        s   = line.series[i]
        clr = colors.get(cmd, "555555")
        s.graphicalProperties.line.solidFill  = clr
        s.graphicalProperties.line.width      = 22000   # 2.2pt
        s.marker.symbol                        = markers.get(cmd, "circle")
        s.marker.size                          = 6
        s.marker.graphicalProperties.solidFill            = clr
        s.marker.graphicalProperties.line.solidFill       = clr

    ws.add_chart(line, f"B{chart_row + 1}")

    # ── 차트 2: 레벨별 스킬 피해 클러스터 막대 ──────────────────
    bar = BarChart()
    bar.type             = "col"
    bar.grouping         = "clustered"
    bar.overlap          = 0
    bar.title            = f"{tier_name}  레벨별 스킬 피해 비교"
    bar.y_axis.title     = "1분 누적 피해 (배율)"
    bar.x_axis.title     = "스킬 레벨"
    bar.style            = 10
    bar.width            = 17
    bar.height           = 13

    data = Reference(ws, min_col=3, max_col=2 + n,
                     min_row=dps_header_row, max_row=dps_data_start + 4)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    for i, cmd in enumerate(cmds):
        if i < len(bar.series):
            bar.series[i].graphicalProperties.solidFill = colors.get(cmd, "555555")

    ws.add_chart(bar, f"K{chart_row + 1}")


def _write_buff_analysis(ws, tier_name, skills, r):
    """
    버프 스킬이 있으면 업타임·DPS 기여 환산 테이블을 작성하고 다음 행을 반환.
    버프 스킬이 하나도 없으면 그대로 r 반환.
    """
    # 버프 스킬만 추림 (buff_dur > 0)
    buff_skills = [sk for sk in skills if any(lv.get("buff_dur", 0) > 0 for lv in sk["levels"])]
    if not buff_skills:
        return r

    # 섹션 헤더
    ws.merge_cells(f"B{r}:M{r}")
    ws[f"B{r}"].value = f"⚡ 버프 효과 환산 — {tier_name}  (공속 단위: 천분률, 1000=100%)"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, size=11, fg=CLR["title_fg"], h_align="left")
    r += 1

    # 테이블 헤더
    hdrs = ["스킬", "레벨", "쿨타임", "지속", "업타임", "공속 기여(평타DPS×)", "치명 기여(critMult+)", "종합 DPS 기여 추정"]
    hdr_cols = ["B","C","D","E","F","G","H","I"]
    for col, h in zip(hdr_cols, hdrs):
        ws[f"{col}{r}"].value = h
        style_cell(ws[f"{col}{r}"], bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"], wrap=True)
    ws.merge_cells(f"I{r}:M{r}")
    ws.row_dimensions[r].height = 30
    r += 1

    for sk in buff_skills:
        cmd = sk["cmd"]
        bg_base = CLR[CMD_BG[cmd]]
        bg_light = CLR[CMD_PARTIAL[cmd]]
        first_in_skill = True

        for lv_data in sk["levels"]:
            lv       = lv_data["lv"]
            cd       = lv_data["cd"]
            buff_dur = lv_data.get("buff_dur", 0)
            if buff_dur == 0:
                continue

            confirmed = lv_data.get("confirmed", True)
            bg = bg_base if confirmed else bg_light

            uptime = min(buff_dur / cd, 1.0) if cd > 0 else 1.0
            uptime_str = "상시 유지" if uptime >= 1.0 else f"{uptime:.0%}"

            # 공속 기여
            atk_spd = lv_data.get("buff_atk_speed", 0)
            if atk_spd > 0:
                # 평타 DPS 증폭 = 1 + uptime * (atk_spd/1000)
                atk_contrib = uptime * (atk_spd / 1000)
                atk_str = f"×{1 + atk_contrib:.2f}  (+{atk_contrib:.0%})"
            else:
                atk_str = "-"

            # 치명 기여 (base cri=0% 가정)
            buff_cri = lv_data.get("buff_cri", 0)
            buff_cri_dmg = lv_data.get("buff_cri_dmg", 0)
            if buff_cri > 0 and buff_cri_dmg > 0:
                # critMult_during_buff = 1 + (cri/100)*(criDmg/100)
                # 천분률: cri/1000 * 100 = cri/10 (%), criDmg/1000 * 100 = criDmg/10 (%)
                crit_mult_buff = (buff_cri / 1000) * (buff_cri_dmg / 1000)
                crit_eff = uptime * crit_mult_buff
                crit_str = f"+{crit_eff:.1%}  (버프시 +{crit_mult_buff:.1%})"
            else:
                crit_str = "-"

            # 종합 (공속+치명 합산, 스킬DPS는 제외)
            total_parts = []
            if atk_spd > 0:
                total_parts.append(f"평타+{uptime*(atk_spd/1000):.0%}")
            if buff_cri > 0:
                total_parts.append(f"치명+{uptime*(buff_cri/1000)*(buff_cri_dmg/1000):.1%}")
            total_str = " / ".join(total_parts) if total_parts else "스킬 외부 수치 미확인"

            # 셀 쓰기
            sk_label = f"{cmd}  {sk['name']}" if first_in_skill else ""
            vals = [sk_label, lv, f"{cd}s" if cd > 0 else "-", f"{buff_dur}s", uptime_str, atk_str, crit_str]
            for i, (col, val) in enumerate(zip(hdr_cols, vals)):
                c = ws[f"{col}{r}"]
                c.value = val
                align_h = "left" if col in ("B",) else "center"
                style_cell(c, bg=bg, h_align=align_h, size=9)
                if not confirmed:
                    c.font = Font(bold=False, size=9, color="999999", italic=True)

            ws.merge_cells(f"I{r}:M{r}")
            c_total = ws[f"I{r}"]
            c_total.value = total_str
            style_cell(c_total, bg=bg, h_align="left", size=9)
            if not confirmed:
                c_total.font = Font(bold=False, size=9, color="999999", italic=True)

            ws.row_dimensions[r].height = 16
            first_in_skill = False
            r += 1

    # 주석
    ws.merge_cells(f"B{r}:M{r}")
    ws[f"B{r}"].value = (
        "※ 공속 기여: 평타 DPS × (1 + 업타임 × 공속/1000)  |  "
        "치명 기여: 업타임 × (치명률/1000) × (치명피해/1000)  |  "
        "base 치명률=0 가정  |  스킬 데미지는 별도 (위 DPS 표 참조)"
    )
    style_cell(ws[f"B{r}"], bg="FEF9E7", fg="7D6608", h_align="left", size=8, italic=True)
    ws.row_dimensions[r].height = 14
    r += 2

    return r


def _write_analysis(ws, tier_name, skills, skill_rows, start_row):
    """분석 섹션 (지속 딜링 + 비중표 + 마나소모 + 차트). start_row 반환."""
    r = start_row
    cmds = [sk["cmd"] for sk in skills]
    n    = len(cmds)

    # ── 지속 딜링 능력 ──
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"].value = f"📊 지속 딜링 능력 (1분 기준) — {tier_name}"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, fg=CLR["title_fg"], h_align="left")
    r += 1

    dps_hdrs = ["레벨"] + cmds + ["1분 합계", "지속DPS", "레벨업 증가율"]
    for i, h in enumerate(dps_hdrs):
        col = get_column_letter(2 + i)
        ws[f"{col}{r}"].value = h
        style_cell(ws[f"{col}{r}"], bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"])
    dps_header_row = r       # 차트용
    dps_data_start = r + 1   # 첫 데이터 행
    r += 1

    for i in range(5):
        ws[f"B{r}"].value = i + 1
        total_col_idx = 3 + n   # 합계 컬럼 (cmds 바로 다음)
        for j, cmd in enumerate(cmds):
            col   = get_column_letter(3 + j)
            sr, _ = skill_rows[cmd]
            ws[f"{col}{r}"].value = f"=K{sr + i}"
            ws[f"{col}{r}"].number_format = "0.00"
        total_col = get_column_letter(total_col_idx)
        first_c   = get_column_letter(3)
        last_c    = get_column_letter(2 + n)
        ws[f"{total_col}{r}"].value = f"=SUM({first_c}{r}:{last_c}{r})"
        ws[f"{total_col}{r}"].number_format = "0.00"
        dps_col  = get_column_letter(total_col_idx + 1)
        ws[f"{dps_col}{r}"].value = f"={total_col}{r}/60"
        ws[f"{dps_col}{r}"].number_format = "0.00"
        rate_col = get_column_letter(total_col_idx + 2)
        if i == 0:
            ws[f"{rate_col}{r}"].value = "-"
        else:
            ws[f"{rate_col}{r}"].value = f"={dps_col}{r}/{dps_col}{r-1}"
            ws[f"{rate_col}{r}"].number_format = "0.0%"
        for col_idx in range(2, total_col_idx + 3):
            style_cell(ws[f"{get_column_letter(col_idx)}{r}"], bg=CLR["section_bg"])
        r += 1

    r += 1

    # ── 비중표 ──
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"].value = f"📊 딜 비중표 (%) — {tier_name}"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, fg=CLR["title_fg"], h_align="left")
    r += 1

    ratio_hdrs = ["레벨"] + cmds
    for i, h in enumerate(ratio_hdrs):
        col = get_column_letter(2 + i)
        ws[f"{col}{r}"].value = h
        style_cell(ws[f"{col}{r}"], bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"])
    r += 1

    total_col_l = get_column_letter(3 + n)   # 합계 컬럼 참조
    for i in range(5):
        dr = dps_data_start + i
        ws[f"B{r}"].value = i + 1
        for j in range(n):
            col_dst = get_column_letter(3 + j)
            col_src = get_column_letter(3 + j)
            ws[f"{col_dst}{r}"].value = f"={col_src}{dr}/{total_col_l}{dr}"
            ws[f"{col_dst}{r}"].number_format = "0.0%"
        for col_idx in range(2, 3 + n):
            style_cell(ws[f"{get_column_letter(col_idx)}{r}"], bg=CLR["formula_bg"])
        r += 1

    r += 1

    # ── 마나 소모량 ──
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"].value = f"💧 마나 소모량 (1분 기준) — {tier_name}"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, fg=CLR["title_fg"], h_align="left")
    r += 1

    mp_hdrs = ["레벨"] + cmds + ["합계", "증가율"]
    for i, h in enumerate(mp_hdrs):
        col = get_column_letter(2 + i)
        ws[f"{col}{r}"].value = h
        style_cell(ws[f"{col}{r}"], bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"])
    r += 1

    for i in range(5):
        ws[f"B{r}"].value = i + 1
        for j, cmd in enumerate(cmds):
            col   = get_column_letter(3 + j)
            sr, _ = skill_rows[cmd]
            ws[f"{col}{r}"].value = f"=L{sr + i}"
            ws[f"{col}{r}"].number_format = "0"
        total_col_mp = get_column_letter(3 + n)
        first_c2     = get_column_letter(3)
        last_c2      = get_column_letter(2 + n)
        ws[f"{total_col_mp}{r}"].value = f"=SUM({first_c2}{r}:{last_c2}{r})"
        ws[f"{total_col_mp}{r}"].number_format = "0"
        rate_col2 = get_column_letter(4 + n)
        if i == 0:
            ws[f"{rate_col2}{r}"].value = "-"
        else:
            ws[f"{rate_col2}{r}"].value = f"={total_col_mp}{r}/{total_col_mp}{r-1}"
            ws[f"{rate_col2}{r}"].number_format = "0.0%"
        for col_idx in range(2, 5 + n):
            style_cell(ws[f"{get_column_letter(col_idx)}{r}"], bg=CLR["formula_bg"])
        r += 1

    r += 1

    # ── 차트 ──
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"].value = f"📈 스킬 분석 차트 — {tier_name}"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, fg=CLR["title_fg"], h_align="left")
    _add_charts(ws, tier_name, cmds, dps_header_row, dps_data_start, r)
    r += 23   # 차트 높이 공간 확보

    r += 1
    r = _write_buff_analysis(ws, tier_name, skills, r)
    return r


# ── 무기 시트 쓰기 ────────────────────────────────────────
def write_weapon_sheet(ws, weapon_name, weapon_data):
    # 열 너비
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 9
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12

    r = 1

    # 제목
    ws.row_dimensions[r].height = 30
    ws.merge_cells(f"B{r}:M{r}")
    c = ws[f"B{r}"]
    c.value = f"X7  {weapon_name}  스킬 밸런스"
    style_cell(c, bg=CLR["title_bg"], bold=True, size=16, fg=CLR["title_fg"], h_align="left")
    r += 1

    ws.merge_cells(f"B{r}:M{r}")
    c = ws[f"B{r}"]
    c.value = weapon_data["concept"]
    style_cell(c, bg="2C3E50", fg="BDC3C7", h_align="left", size=11, italic=True)
    r += 2

    # 각 티어
    for tier_data in weapon_data["tiers"]:
        r, _ = write_tier_section(ws, tier_data["tier"], tier_data, r)


# ── 스킬 리스트 시트 ─────────────────────────────────────
def write_skill_list_sheet(ws):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30

    r = 1
    ws.merge_cells(f"B{r}:G{r}")
    ws[f"B{r}"].value = "X7 스킬 전체 리스트 ([X7] 스킬 밸런스 - 스킬 리스트.csv 기준)"
    style_cell(ws[f"B{r}"], bg=CLR["title_bg"], bold=True, size=14, fg=CLR["title_fg"], h_align="left")
    r += 1

    headers = ["장비 구분", "커맨드", "GroupId", "스킬명 (Lv1)", "데이터 상태", "비고"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws[f"{col}{r}"].value = h
        style_cell(ws[f"{col}{r}"], bg=CLR["header_bg"], bold=True, fg=CLR["header_fg"])
    r += 1

    skill_list = [
        # 한손검
        ("1티어 한손검", "Q", "100201", "철벽의 반격",      "🟡 Lv.1만 확보", "cd=5s/MP=30/dmg=225%"),
        ("1티어 한손검", "W", "100251", "방패 돌진",        "🟡 Lv.1만 확보", "cd=0s(서브)/dmg=200%"),
        ("1티어 한손검", "E", "100301", "방패 투척",        "🟡 Lv.1만 확보", "cd=15s/MP=30/dmg=200%"),
        ("1티어 한손검", "R", "100351", "파멸의 징벌",      "🟡 Lv.1만 확보", "cd=45s/MP=30/dmg=400%"),
        ("2티어 한손검", "Q", "100401", "제압",             "🟡 Lv.1만 확보", "cd=5s/dmg=100%"),
        ("2티어 한손검", "W", "100451", "복수의 갑옷",      "🟡 Lv.1만 확보", "cd=10s/버프"),
        ("2티어 한손검", "E", "100501", "모두 쉿",          "🟡 Lv.1만 확보", "cd=15s/MP=50/CC(침묵)"),
        ("2티어 한손검", "R", "100551", "진공 폭발",        "🟡 Lv.1만 확보", "cd=30s/dmg=300%"),
        ("3티어 한손검", "Q", "100601", "성스러운 일격",    "⚖️ 균형 추정값", "cd=5s/MP=10/dmg=200% (양손검 3티어 Q 기준)"),
        ("3티어 한손검", "W", "100651", "가호의 빛",        "⚖️ 균형 추정값", "cd=12s/MP=30/버프"),
        ("3티어 한손검", "E", "100701", "폭주하는 광휘",    "⚖️ 균형 추정값", "cd=15s/MP=15/dmg=300% (양손검 3티어 E 기준)"),
        ("3티어 한손검", "R", "100751", "광휘의 심판",      "⚖️ 균형 추정값", "cd=40s/MP=50/dmg=400% (양손검 3티어 R 기준)"),
        # 양손검
        ("1티어 양손검", "Q", "200401", "폭주",             "🟡 Lv.1만 확보", "cd=0s/MP=15/dmg=300%"),
        ("1티어 양손검", "W", "200451", "분노의 일격",      "🟡 Lv.1만 확보", "cd=10s/버프"),
        ("1티어 양손검", "E", "200501", "대지 분쇄",        "🟡 Lv.1만 확보", "cd=0s/dmg=50%"),
        ("1티어 양손검", "R", "200551", "공간 가르기",      "🟡 Lv.1만 확보", "cd=12s/MP=30/dmg=625%"),
        ("2티어 양손검", "Q", "200601", "몰아치기",         "🟡 Lv.1만 확보", "cd=6s/버프"),
        ("2티어 양손검", "W", "200651", "찌르기",           "🟡 Lv.1만 확보", "cd=12s/MP=30/dmg=250%"),
        ("2티어 양손검", "E", "200701", "결투 돌입",        "🟡 Lv.1만 확보", "cd=0s/dmg=150%"),
        ("2티어 양손검", "R", "200751", "섬광의 길",        "🟡 Lv.1만 확보", "cd=8s/MP=30/dmg=150%"),
        ("3티어 양손검", "Q", "200801", "가벼운 손놀림",    "✅ 전 레벨 확보", "cd=5s/MP=10/dmg=200~280%"),
        ("3티어 양손검", "W", "200851", "대지 가르기",      "✅ 전 레벨 확보", "cd=12s/MP=30~50/dmg=250~350%"),
        ("3티어 양손검", "E", "200901", "휘몰이",           "✅ 전 레벨 확보", "cd=15s/6-hit/dmg=300~420%"),
        ("3티어 양손검", "R", "200951", "적진으로",         "✅ 전 레벨 확보", "cd=40~32s/MP=50/dmg=400~560%"),
        # 지팡이
        ("1티어 지팡이", "Q", "400201", "화염 장막",        "🟡 Lv.1만 확보", "cd=8s/DoT 6틱 합산=216%"),
        ("1티어 지팡이", "W", "400251", "불의 비",          "🟡 Lv.1만 확보", "cd=15s/MP=50/dmg=400%"),
        ("1티어 지팡이", "E", "400301", "화염 폭발",        "🟡 Lv.1만 확보", "cd=20s/MP=20/dmg=200%"),
        ("1티어 지팡이", "R", "400351", "별똥별",           "🟡 Lv.1만 확보", "cd=40s/MP=100/dmg=600%"),
        ("2티어 지팡이", "Q", "100001", "신비한 가시 나무", "✅ 전 레벨 확보", "cd=8s/DoT 6틱 합산=216~360%"),
        ("2티어 지팡이", "W", "100051", "가시 나무 덩쿨",   "✅ 전 레벨 확보", "cd=12s/CC+dmg=250~330%"),
        ("2티어 지팡이", "E", "100101", "치유의 꽃잎",      "✅ 전 레벨 확보", "cd=1s/힐"),
        ("2티어 지팡이", "R", "100150", "치유의 바람",      "✅ 전 레벨 확보", "cd=30s/힐DoT"),
        ("3티어 지팡이", "Q", "400601", "얼음 화살",        "✅ 전 레벨 확보", "cd=3s/MP=30~42/dmg=200~280%"),
        ("3티어 지팡이", "W", "400651", "다후타의 손짓",    "✅ 전 레벨 확보", "cd=15s/DoT 5틱=400~560%"),
        ("3티어 지팡이", "E", "400701", "심판의 낙뢰",      "✅ 전 레벨 확보", "cd=40~32s/2-hit=600~840%"),
        ("3티어 지팡이", "R", "400751", "방울방울",         "✅ 전 레벨 확보", "cd=20s/CC(부유)"),
        # 단검
        ("1티어 단검",   "Q", "500401", "질풍 가르기",      "🟡 Lv.1만 확보", "cd=5s/MP=30/dmg=150%"),
        ("1티어 단검",   "W", "500451", "무기 연마",        "🟡 Lv.1만 확보", "cd=10s/버프(공격속도)"),
        ("1티어 단검",   "E", "500501", "어둠의 일격",      "🟡 Lv.1만 확보", "cd=15s/MP=15/dmg=300%"),
        ("1티어 단검",   "R", "500551", "암살자의 분노",    "🟡 Lv.1만 확보", "cd=40s/MP=100/버프"),
        ("2티어 단검",   "Q", "500601", "그림자 매",        "⚖️ 균형 추정값", "cd=5s/MP=20/dmg=150%"),
        ("2티어 단검",   "W", "500651", "쌍수 달인",        "✅ 전 레벨 확보", "cd=12s/버프(공격속도)"),
        ("2티어 단검",   "E", "500701", "진격",             "⚖️ 균형 추정값", "cd=12s/MP=25/dmg=200%"),
        ("2티어 단검",   "R", "500751", "전율하는 매",      "⚖️ 균형 추정값", "cd=30s/MP=60/dmg=300%"),
        ("3티어 단검",   "Q", "500801", "질풍 가르기",      "✅ 전 레벨 확보", "cd=5s/dmg=150~210%"),
        ("3티어 단검",   "W", "500851", "무기 연마",        "✅ 전 레벨 확보", "cd=10s/버프(공격속도)"),
        ("3티어 단검",   "E", "500901", "어둠의 일격",      "✅ 전 레벨 확보", "cd=15s/dmg=300~420%"),
        ("3티어 단검",   "R", "500951", "암살자의 분노",    "✅ 전 레벨 확보", "cd=40~32s/버프"),
    ]

    weapon_colors = {"한손검": "EBF5FB", "양손검": "FEF9E7", "지팡이": "F3E5F5", "단검": "E8F8F5"}
    cmd_colors    = {"Q": CLR["q_bg"], "W": CLR["w_bg"], "E": CLR["e_bg"], "R": CLR["r_bg"]}
    status_colors = {"✅": "DFFFDF", "🟡": "FFFDE7", "❌": "FFE0E0"}

    for row_data in skill_list:
        equip, cmd, gid, name, status, note = row_data
        wpn_key = next((k for k in weapon_colors if k in equip), None)
        bg_base = weapon_colors[wpn_key] if wpn_key else "FFFFFF"
        bg_cmd  = cmd_colors[cmd]
        s_key   = next((k for k in status_colors if k in status), None)
        bg_st   = status_colors[s_key] if s_key else "FFFFFF"

        vals = [equip, cmd, gid, name, status, note]
        bgs  = [bg_base, bg_cmd, bg_base, bg_st, bg_st, bg_base]
        for i, (val, bg) in enumerate(zip(vals, bgs)):
            col = get_column_letter(2 + i)
            ws[f"{col}{r}"].value = val
            style_cell(ws[f"{col}{r}"], bg=bg, h_align="left" if i >= 2 else "center")
        r += 1


# ── 무기 비교 개요 시트 ──────────────────────────────────
def write_overview_sheet(ws):
    """첫 번째 시트 — 무기 컨셉 + 레이더 차트 + 클러스터 막대 차트."""

    WPN_ORDER  = ["한손검", "양손검", "지팡이", "단검"]
    WPN_COLORS = {"한손검": "5D6D7E", "양손검": "7B241C",
                  "지팡이": "1A4F72", "단검":   "1E6B3E"}
    CMD_ORDER  = ["Q", "W", "E", "R"]
    CMD_BG_OV  = {"Q": "D5F5E3", "W": "D6EAF8", "E": "F9EBEA", "R": "FEF9E7"}

    WPN_CONCEPT = {
        "한손검": "방어 + 반격 스타일 — 버프·CC로 생존하며 안전한 타이밍에 강타",
        "양손검": "파워 딜링 — 높은 계수 스킬로 짧은 시간에 대량 피해 집중",
        "지팡이": "마법 컨트롤 — DoT·다중히트·CC로 딜링과 견제를 동시에 수행",
        "단검":   "속도 우선 — 낮은 쿨다운과 높은 타격 빈도로 지속 DPS 극대화",
    }

    # 컬럼 너비: A(indent) B(label) C~F(한손검~단검) G(합산)
    for col, w in zip("ABCDEFG", [3, 12, 15, 15, 15, 15, 14]):
        ws.column_dimensions[col].width = w

    r = 1

    # ── 타이틀 ──────────────────────────────────────────
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "  ⚔️  X7  무기별 스킬 밸런스  —  비교 개요"
    style_cell(c, bg="1F2D3D", fg="E8B84B", bold=True, size=14, h_align="left")
    ws.row_dimensions[r].height = 36
    r += 2

    # ── 무기 컨셉 ─────────────────────────────────────────
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value = "  📋  무기별 설계 컨셉"
    style_cell(ws[f"A{r}"], bg="2C3E50", fg="E8B84B", bold=True, size=11, h_align="left")
    ws.row_dimensions[r].height = 24
    r += 1

    for idx, wpn in enumerate(WPN_ORDER):
        bg   = WPN_COLORS[wpn]
        rbg  = "F2F3F4" if idx % 2 == 0 else "EBF5FB"
        ws[f"B{r}"].value = wpn
        style_cell(ws[f"B{r}"], bg=bg, fg="FFFFFF", bold=True, size=10)
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"].value = WPN_CONCEPT[wpn]
        style_cell(ws[f"C{r}"], bg=rbg, fg="1A2A3A", h_align="left", size=10, wrap=True)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1

    # ── 비교 데이터 테이블 ────────────────────────────────
    # 레이아웃: B=커맨드, C=한손검, D=양손검, E=지팡이, F=단검, G=최고값
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value = "  📊  3티어 Lv.1  커맨드별 1분 누적 피해 (배율×) — 차트 데이터"
    style_cell(ws[f"A{r}"], bg="2C3E50", fg="E8B84B", bold=True, size=11, h_align="left")
    ws.row_dimensions[r].height = 24
    r += 1

    data_header_row = r
    # 헤더 행
    ws[f"B{r}"].value = "커맨드"
    style_cell(ws[f"B{r}"], bg="34495E", fg="FFFFFF", bold=True, size=9)
    for j, wpn in enumerate(WPN_ORDER):
        col = get_column_letter(3 + j)
        ws[f"{col}{r}"].value = wpn
        style_cell(ws[f"{col}{r}"], bg=WPN_COLORS[wpn], fg="FFFFFF", bold=True, size=9)
    ws[f"G{r}"].value = "최고"
    style_cell(ws[f"G{r}"], bg="E8B84B", fg="1F2D3D", bold=True, size=9)
    ws.row_dimensions[r].height = 20
    r += 1

    data_start_row = r

    # 값 계산: 1분 누적 피해 = (60/cd) × dmg
    for cmd in CMD_ORDER:
        ws[f"B{r}"].value = cmd
        cbg = CMD_BG_OV[cmd]
        style_cell(ws[f"B{r}"], bg=cbg, fg="1A2A3A", bold=True, size=10)
        row_vals = []
        for j, wpn in enumerate(WPN_ORDER):
            tier3  = next(t for t in WEAPON_DATA[wpn]["tiers"] if t["tier"] == "3티어")
            skills = {sk["cmd"]: sk for sk in tier3["skills"]}
            lv1    = skills[cmd]["levels"][0] if cmd in skills else None
            cd     = (lv1["cd"]  or 0) if lv1 else 0
            dmg    = (lv1["dmg"] or 0.0) if lv1 else 0.0
            val    = round((60 / cd) * dmg, 1) if cd > 0 and dmg > 0 else 0
            row_vals.append(val)
            col = get_column_letter(3 + j)
            ws[f"{col}{r}"].value = val
            style_cell(ws[f"{col}{r}"], bg=cbg, fg="1A2A3A", size=10)
        # 최고값 강조
        best = max(row_vals)
        ws[f"G{r}"].value = best
        style_cell(ws[f"G{r}"], bg="F9E79F", fg="7D6608", bold=True, size=10)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1

    # ── 차트 영역 ─────────────────────────────────────────
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value = "  📈  무기별 스킬 밸런스 비교 차트"
    style_cell(ws[f"A{r}"], bg="2C3E50", fg="E8B84B", bold=True, size=11, h_align="left")
    ws.row_dimensions[r].height = 24
    chart_row = r

    # ── 클러스터 막대: 커맨드별 무기 비교 ────────────────
    bar = BarChart()
    bar.type     = "col"
    bar.grouping = "clustered"
    bar.overlap  = 0
    bar.title    = "커맨드별 무기 피해 비교  (3티어 Lv.1 기준  /  1분 누적 피해)"
    bar.y_axis.title = "1분 누적 피해 (배율)"
    bar.x_axis.title = "스킬 커맨드"
    bar.style    = 10
    bar.width    = 24
    bar.height   = 16

    wpn_color_list = [WPN_COLORS[w] for w in WPN_ORDER]

    b_data = Reference(ws, min_col=3, max_col=6,
                       min_row=data_header_row, max_row=data_start_row + 3)
    b_cats = Reference(ws, min_col=2,
                       min_row=data_start_row, max_row=data_start_row + 3)
    bar.add_data(b_data, titles_from_data=True)
    bar.set_categories(b_cats)

    for i, clr in enumerate(wpn_color_list):
        if i < len(bar.series):
            bar.series[i].graphicalProperties.solidFill = clr

    ws.add_chart(bar, f"B{chart_row + 1}")


# ── 메인 ─────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    # 첫 번째 시트: 무기 비교 개요
    wb.active.title = "무기 비교 개요"
    write_overview_sheet(wb.active)

    # 무기별 시트
    for name in ["한손검", "양손검", "지팡이", "단검"]:
        ws = wb.create_sheet(name)
        write_weapon_sheet(ws, name, WEAPON_DATA[name])

    ws_list = wb.create_sheet("스킬 리스트")
    write_skill_list_sheet(ws_list)

    out_path = r"C:\AI_simulator\스킬 밸런스 기획\X7_무기_스킬_밸런스.xlsx"
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")
    print("  시트: 무기 비교 개요 / 한손검 / 양손검 / 지팡이 / 단검 / 스킬 리스트")


if __name__ == "__main__":
    main()
