"""
X7_장비_제작_및_드랍.xlsx — 📋 장비 아이템 리스트 시트 수정
[X7] 성장_경제 밸런스 - 시트29.csv 기준으로 교정

수정 내용:
1. "마나 소모량 감소" → "마력 자연 회복" (RegenMpVary 오역 전체 수정)
2. 2T 전리품제작 신발 3개 추가 (판금/가죽/천 신발 누락)
3. 2T 보스드랍 중복 3개 삭제 (CSV에 없는 항목 제거)
"""

import openpyxl
from openpyxl import load_workbook
from copy import copy

# ── 스타일 복사 헬퍼 ──────────────────────────────────────────────
def copy_cell_fmt(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

def insert_row_styled(ws, row_idx, data, style_src_row):
    """row_idx 위치에 새 행 삽입 후 style_src_row 서식 복사"""
    ws.insert_rows(row_idx)
    for col, val in enumerate(data, 1):
        new_cell = ws.cell(row=row_idx, column=col, value=val)
        src_cell = ws.cell(row=style_src_row, column=col)
        copy_cell_fmt(src_cell, new_cell)

# ── 파일 열기 ─────────────────────────────────────────────────────
wb = load_workbook("기획서/X7_장비_제작_및_드랍.xlsx")
ws = wb["📋 장비 아이템 리스트"]

# ════════════════════════════════════════════════════════════════
# STEP 1: "마나 소모량 감소" → "마력 자연 회복" 전체 치환
#   RegenMpVary 의 올바른 한국어 명칭
# ════════════════════════════════════════════════════════════════
fixed_cells = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "마나 소모량 감소" in cell.value:
            old = cell.value
            cell.value = cell.value.replace("마나 소모량 감소", "마력 자연 회복")
            fixed_cells.append(f"  {cell.coordinate}: '{old}' → '{cell.value}'")

print(f"[STEP 1] '마나 소모량 감소' 수정 {len(fixed_cells)}곳:")
for s in fixed_cells:
    print(s)

# ════════════════════════════════════════════════════════════════
# STEP 2: 2T 전리품제작 신발 3개 삽입
#   CSV 기준:
#     판금 신발: 방어력 / 생명력 자연 회복  (after 판금 장갑 row18)
#     가죽 신발: 공격력 / 최대 생명력      (after 가죽 장갑 row21 → row22 after 1st insert)
#     천  신발: 마력 자연 회복 / 최대 생명력 (after 천 장갑  row24 → row26 after 2nd insert)
# ════════════════════════════════════════════════════════════════
BIGO_T2_LOOT = "T2 전리품 필요"

판금신발 = ("2T", "전리품 제작", "판금", "신발", "방어력",        "생명력 자연 회복", "-", BIGO_T2_LOOT)
가죽신발 = ("2T", "전리품 제작", "가죽", "신발", "공격력",        "최대 생명력",     "-", BIGO_T2_LOOT)
천신발   = ("2T", "전리품 제작", "천",  "신발", "마력 자연 회복", "최대 생명력",     "-", BIGO_T2_LOOT)

# 1) 판금 신발 → row18(판금 장갑) 다음에 삽입
insert_row_styled(ws, 19, 판금신발, style_src_row=18)
print("\n[STEP 2a] 판금 신발 row19 삽입")

# 2) 가죽 신발 → 1차 삽입으로 가죽 장갑이 row22로 이동 → row23에 삽입
insert_row_styled(ws, 23, 가죽신발, style_src_row=22)
print("[STEP 2b] 가죽 신발 row23 삽입")

# 3) 천 신발 → 2차 삽입으로 천 장갑이 row26으로 이동 → row27에 삽입
insert_row_styled(ws, 27, 천신발,   style_src_row=26)
print("[STEP 2c] 천 신발 row27 삽입")

# ════════════════════════════════════════════════════════════════
# STEP 3: 2T 보스드랍 과잉 항목 3개 삭제
#   3회 삽입 후 원래 row34~39 → row37~42 로 이동
#
#   row37: 판금신발 방어력/스킬가속       ✓ KEEP
#   row38: 판금신발 방어력/생명력자연회복  ✗ DELETE  (CSV에 없음)
#   row39: 가죽신발 공격력/최대생명력      ✗ DELETE  (CSV에 없음)
#   row40: 가죽신발 공격력/치명타확률      ✓ KEEP
#   row41: 천신발  마력자연회복/최대생명력  ✗ DELETE  (CSV에 없음)
#   row42: 천신발  마력자연회복/공격력     ✓ KEEP
#
#   삭제는 아래→위 순서로 (row 번호 밀림 방지)
# ════════════════════════════════════════════════════════════════

# 삭제 전 확인 출력
print("\n[STEP 3] 삭제 전 보스드랍 구간 (row37~42):")
for r in range(37, 43):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
    print(f"  row{r}: {vals}")

ws.delete_rows(41)  # 천신발 과잉 (아래부터)
ws.delete_rows(39)  # 가죽신발 과잉
ws.delete_rows(38)  # 판금신발 과잉
print("\n[STEP 3] 행 38, 39, 41 삭제 완료")

# ════════════════════════════════════════════════════════════════
# STEP 4: 요약 수치 업데이트
# ════════════════════════════════════════════════════════════════
t1 = t2 = t3 = 0
basic = loot = dungeon = boss = 0
for row_vals in ws.iter_rows(min_row=4, values_only=True):
    tier = row_vals[0]
    acq  = row_vals[1]
    if tier == "1T": t1 += 1
    elif tier == "2T": t2 += 1
    elif tier == "3T": t3 += 1
    if acq == "기본 제작": basic += 1
    elif acq == "전리품 제작": loot += 1
    elif acq == "던전 코어 제작": dungeon += 1
    elif acq == "보스 드랍": boss += 1

# 요약 행 찾아서 업데이트
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "티어별 아이템 수" in cell.value:
            r = cell.row
            ws.cell(row=r,   column=3).value = f"1T  :  {t1}개"
            ws.cell(row=r,   column=4).value = f"2T  :  {t2}개"
            ws.cell(row=r,   column=5).value = f"3T  :  {t3}개"
            ws.cell(row=r+1, column=3).value = f"기본 제작  :  {basic}개"
            ws.cell(row=r+1, column=4).value = f"전리품 제작  :  {loot}개"
            ws.cell(row=r+1, column=5).value = f"던전 코어 제작  :  {dungeon}개"
            ws.cell(row=r+1, column=6).value = f"보스 드랍  :  {boss}개"
            print(f"\n[STEP 4] 요약 갱신 (row{r}~{r+1})")
            break

# ── 저장 ──────────────────────────────────────────────────────
wb.save("기획서/X7_장비_제작_및_드랍.xlsx")
print("\n✅ 저장 완료: 기획서/X7_장비_제작_및_드랍.xlsx")

# ── 최종 검증 ─────────────────────────────────────────────────
print("\n=== 최종 검증 ===")
wb2 = load_workbook("기획서/X7_장비_제작_및_드랍.xlsx")
ws2 = wb2["📋 장비 아이템 리스트"]
t1=t2=t3=basic=loot=dungeon=boss=0
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if rv[0] == "1T": t1 += 1
    elif rv[0] == "2T": t2 += 1
    elif rv[0] == "3T": t3 += 1
    if rv[1] == "기본 제작": basic += 1
    elif rv[1] == "전리품 제작": loot += 1
    elif rv[1] == "던전 코어 제작": dungeon += 1
    elif rv[1] == "보스 드랍": boss += 1

print(f"  티어별  : 1T={t1} / 2T={t2} / 3T={t3}")
print(f"  획득별  : 기본={basic} / 전리품={loot} / 던전코어={dungeon} / 보스={boss}")
print()
print("  [2T 전리품제작 12개 확인]")
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if rv[0] == "2T" and rv[1] == "전리품 제작":
        print(f"    {rv[2]} {rv[3]} | {rv[4]} / {rv[5]}")
print()
print("  [2T 보스드랍 3개 확인]")
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if rv[0] == "2T" and rv[1] == "보스 드랍":
        print(f"    {rv[2]} {rv[3]} | {rv[4]} / {rv[5]}")
print()
# 마나소모량감소 잔존 여부
remain = [c.coordinate for row in ws2.iter_rows()
          for c in row if isinstance(c.value, str) and "마나 소모량 감소" in c.value]
if remain:
    print(f"  ⚠️ '마나 소모량 감소' 잔존: {remain}")
else:
    print("  ✅ '마나 소모량 감소' 잔존 없음")
