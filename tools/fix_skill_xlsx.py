"""
X7_무기_스킬.xlsx 수정:
1. 열 너비 교정 (모든 시트)
2. 3티어 기본셋 내용 검증 → 불일치 시 셀 수정
"""
import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"

# rebuild_bow_sheet.py 기준 참조 너비
SKILL_SHEET_WIDTHS = {
    "A": 2, "B": 6, "C": 5, "D": 9, "E": 9,
    "F": 14, "G": 12, "H": 36, "I": 9, "J": 9, "K": 12, "L": 12
}

# constants.py SKILLS 3티어 기본셋 Lv.1
SKILLS_3T = {
    "양손검": [
        {"cmd": "Q", "name": "가벼운 손놀림", "cd": 5.0,  "mp": 15,  "dmg": 200},
        {"cmd": "W", "name": "대지 가르기",   "cd": 12.0, "mp": 30,  "dmg": 250},
        {"cmd": "E", "name": "휘몰이",        "cd": 15.0, "mp": 25,  "dmg": 300},
        {"cmd": "R", "name": "적진으로",      "cd": 40.0, "mp": 100, "dmg": 400},
    ],
    "한손검": [
        {"cmd": "Q", "name": "성스러운 일격", "cd": 6.0,  "mp": 10,  "dmg": 150},
        {"cmd": "W", "name": "가호의 빛",     "cd": 12.0, "mp": 20,  "dmg": 200},
        {"cmd": "E", "name": "폭주하는 광휘", "cd": 20.0, "mp": 30,  "dmg": 250},
        {"cmd": "R", "name": "광휘의 심판",   "cd": 45.0, "mp": 100, "dmg": 500},
    ],
    "활": [
        {"cmd": "Q", "name": "충격 사격",   "cd": 5.0,  "mp": 10,  "dmg": 220},
        {"cmd": "W", "name": "연속 쏘기",   "cd": 12.0, "mp": 20,  "dmg": 280},
        {"cmd": "E", "name": "급습",        "cd": 20.0, "mp": 30,  "dmg": 350},
        {"cmd": "R", "name": "발묶음",      "cd": 45.0, "mp": 100, "dmg": 500},
    ],
    "지팡이": [
        {"cmd": "Q", "name": "얼음 화살",      "cd": 4.0,  "mp": 20,  "dmg": 200},
        {"cmd": "W", "name": "다후타의 손짓",  "cd": 15.0, "mp": 40,  "dmg": 400},
        {"cmd": "E", "name": "심판의 낙뢰",    "cd": 40.0, "mp": 50,  "dmg": 600},
        {"cmd": "R", "name": "방울방울",       "cd": 20.0, "mp": 100, "dmg": 300},
    ],
    "단검": [
        {"cmd": "Q", "name": "추격자의 발톱", "cd": 6.0,  "mp": 15,  "dmg": 220},
        {"cmd": "W", "name": "침묵의 일격",   "cd": 13.0, "mp": 30,  "dmg": 120},
        {"cmd": "E", "name": "사방의 비수",   "cd": 15.0, "mp": 40,  "dmg": 250},
        {"cmd": "R", "name": "유혈의 장막",   "cd": 40.0, "mp": 100, "dmg": 500},
    ],
}

WEAPON_SHEETS = ["한손검", "양손검", "활", "지팡이", "단검"]

wb = openpyxl.load_workbook(XLSX_PATH)
print("시트 목록:", wb.sheetnames)

# ── 1. 열 너비 교정 ────────────────────────────────────────────
print("\n=== 열 너비 교정 ===")
for sheet_name in WEAPON_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"  {sheet_name}: 시트 없음")
        continue
    ws = wb[sheet_name]
    old_b = ws.column_dimensions["B"].width
    old_e = ws.column_dimensions["E"].width
    for col_letter, width in SKILL_SHEET_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    print(f"  {sheet_name}: B {old_b:.1f}→6, E {old_e:.1f}→9, H→36 교정")

# 무기 비교 개요
if "무기 비교 개요" in wb.sheetnames:
    ws = wb["무기 비교 개요"]
    old = {c: ws.column_dimensions[c].width for c in "ABCDEFGH"}
    # 구조 파악: A가 60이면 row-label column → 20으로; C가 60이면 description → 35로
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    # D-H 유지 (15/15/15/14/13 등 이미 적절)
    print(f"  무기 비교 개요: A {old['A']:.1f}→20, C {old['C']:.1f}→35")

# 스킬 리스트
if "스킬 리스트" in wb.sheetnames:
    ws = wb["스킬 리스트"]
    old_b = ws.column_dimensions["B"].width
    old_g = ws.column_dimensions["G"].width
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["G"].width = 30
    print(f"  스킬 리스트: B {old_b:.1f}→20, G {old_g:.1f}→30")

# ── 2. 내용 검증 + 수정 ────────────────────────────────────────
print("\n=== 3티어 기본셋 내용 검증 (Lv.1 기준) ===")
total_fixed = 0

for weapon in WEAPON_SHEETS:
    if weapon not in wb.sheetnames:
        continue
    ws = wb[weapon]
    expected_list = SKILLS_3T[weapon]

    # 3티어 섹션 헤더 행 찾기
    t3_row = None
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and "3티어" in str(val):
                t3_row = row_idx
                break
        if t3_row:
            break

    if not t3_row:
        print(f"  {weapon}: 3티어 섹션 미발견 — 건너뜀")
        continue

    # 컬럼 헤더 행 찾기 (3티어 이후 "쿨타임" 포함 행)
    header_row = None
    for r in range(t3_row + 1, min(t3_row + 20, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and "쿨타임" in str(val):
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        print(f"  {weapon}: 컬럼 헤더 미발견 — 건너뜀")
        continue

    # 컬럼 인덱스 매핑 (쿨타임, 마나소모, 피해량)
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "")
        if "쿨타임" in val:   col_map["cd"] = c
        if "마나" in val:     col_map["mp"] = c
        if "피해량" in val:   col_map["dmg"] = c
        if "레벨" in val:     col_map["lv"] = c
        if "커맨드" in val:   col_map["cmd"] = c

    print(f"  {weapon}: 3T행={t3_row}, 헤더행={header_row}, 컬럼={col_map}")

    if len(col_map) < 4:
        print(f"    ⚠ 컬럼 매핑 불완전, 건너뜀")
        continue

    # Lv.1 행 수집 (헤더 이후)
    lv1_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        lv_val = ws.cell(row=r, column=col_map["lv"]).value
        if lv_val and str(lv_val).strip() in ("Lv.1", "1"):
            # cmd 열도 확인 (Q/W/E/R 포함 행)
            cmd_val = str(ws.cell(row=r, column=col_map["cmd"]).value or "")
            for sk in ["Q", "W", "E", "R"]:
                if sk in cmd_val:
                    lv1_rows.append((r, sk))
                    break

    print(f"    Lv.1 행: {[(r, sk) for r, sk in lv1_rows]}")

    # 각 스킬 비교 및 수정
    for exp in expected_list:
        cmd = exp["cmd"]
        # 해당 cmd의 Lv.1 행 찾기
        match_row = next((r for r, sk in lv1_rows if sk == cmd), None)
        if not match_row:
            print(f"    ⚠ {weapon} {cmd} 행 미발견")
            continue

        issues = []
        # CD
        if "cd" in col_map:
            cell_cd = ws.cell(row=match_row, column=col_map["cd"])
            val = cell_cd.value
            try:
                if val is not None and float(val) != exp["cd"]:
                    issues.append(f"CD {val}→{exp['cd']}")
                    cell_cd.value = exp["cd"]
            except (ValueError, TypeError):
                pass

        # MP
        if "mp" in col_map:
            cell_mp = ws.cell(row=match_row, column=col_map["mp"])
            val = cell_mp.value
            try:
                if val is not None and int(float(val)) != exp["mp"]:
                    issues.append(f"MP {val}→{exp['mp']}")
                    cell_mp.value = exp["mp"]
            except (ValueError, TypeError):
                pass

        # DMG (피해량은 배율이므로 소수점 비교)
        if "dmg" in col_map:
            cell_dmg = ws.cell(row=match_row, column=col_map["dmg"])
            val = cell_dmg.value
            try:
                if val is not None and abs(float(val) - exp["dmg"]) > 0.01:
                    issues.append(f"DMG {val}→{exp['dmg']}")
                    cell_dmg.value = exp["dmg"]
            except (ValueError, TypeError):
                pass

        if issues:
            print(f"    ✏ {weapon} {cmd} {exp['name']}: {', '.join(issues)}")
            total_fixed += len(issues)
        else:
            print(f"    ✓ {weapon} {cmd} {exp['name']}: OK")

print(f"\n수정된 셀: {total_fixed}개")

# ── 3. 저장 ───────────────────────────────────────────────────
try:
    wb.save(XLSX_PATH)
    print(f"저장 완료: {XLSX_PATH}")
except PermissionError:
    alt = XLSX_PATH.replace(".xlsx", "_fixed.xlsx")
    wb.save(alt)
    print(f"원본 잠김 — 대체 저장: {alt}")
