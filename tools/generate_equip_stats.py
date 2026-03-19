"""
tools/generate_equip_stats.py
장비 능력치 기획서 Excel 생성
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\AI_simulator\기획서\장비_능력치_기획서.xlsx"

# ── 데이터 ─────────────────────────────────────────────────────
TIERS       = ["Tier1","Tier2","Tier3","Tier4","Tier5","Tier6","Tier7"]
TIER_LABELS = ["T1","T2","T3","T4","T5","T6","T7"]
GRADES      = ["일반","고급","희귀","고대","영웅","유일","유물"]
GRADE_ORDER = {g: i for i, g in enumerate(GRADES)}

TIER_MAX_GRADE = {
    "Tier1":"고급","Tier2":"희귀","Tier3":"고대",
    "Tier4":"영웅","Tier5":"유일","Tier6":"유물","Tier7":"유물",
}

WPN_BASE  = {"Tier1":60,"Tier2":80,"Tier3":120,"Tier4":180,"Tier5":260,"Tier6":360,"Tier7":480}
WPN_GRADE = {"일반":0,"고급":20,"희귀":40,"고대":60,"영웅":80,"유일":100,"유물":120}
WPN_ENH   = {"Tier1":8,"Tier2":10,"Tier3":12,"Tier4":14,"Tier5":16,"Tier6":18,"Tier7":20}

ARM_BASE  = {"Tier1":30,"Tier2":40,"Tier3":60,"Tier4":90,"Tier5":130,"Tier6":180,"Tier7":240}
ARM_GRADE = {"일반":0,"고급":10,"희귀":20,"고대":30,"영웅":40,"유일":50,"유물":55}
ARM_ENH   = {"Tier1":4,"Tier2":5,"Tier3":6,"Tier4":7,"Tier5":8,"Tier6":9,"Tier7":10}

RING_BASE  = {"Tier1":0,"Tier2":0,"Tier3":30,"Tier4":45,"Tier5":65,"Tier6":90,"Tier7":120}
RING_GRADE = {"일반":0,"고급":5,"희귀":10,"고대":15,"영웅":20,"유일":25,"유물":30}
RING_ENH   = {"Tier1":4,"Tier2":5,"Tier3":6,"Tier4":7,"Tier5":8,"Tier6":9,"Tier7":10}

NECK_BASE  = {"Tier1":0,"Tier2":0,"Tier3":60,"Tier4":90,"Tier5":130,"Tier6":180,"Tier7":240}
NECK_GRADE = {"일반":0,"고급":10,"희귀":20,"고대":30,"영웅":40,"유일":50,"유물":60}
NECK_ENH   = {"Tier1":4,"Tier2":5,"Tier3":6,"Tier4":7,"Tier5":8,"Tier6":9,"Tier7":10}

EAR_BASE  = NECK_BASE.copy()
EAR_GRADE = NECK_GRADE.copy()
EAR_ENH   = NECK_ENH.copy()

MAX_ENH   = 10
ENH_RANGE = list(range(MAX_ENH + 1))

CAPS = [
    ("공격력 증가",    "AttackVaryper",        50,  "%",  "무기/반지 기본 공격력에 비율로 가산"),
    ("방어력 증가",    "DefenseVaryper",       50,  "%",  "방어구/악세 기본 방어력에 비율로 가산"),
    ("추가 공격력",    "AddAttackVary",       300,  "고정", "최종 ATK 계산 후 고정값 추가"),
    ("추가 방어력",    "AddDefenseVary",      600,  "고정", "최종 DEF 계산 후 고정값 추가"),
    ("스킬 가속",      "SkillCooldownAccVary",100,  "n",  "유효CD = 기본CD × 100 / (100+n)"),
    ("주는 피해 증가", "DamageUpVaryper",     100,  "%",  "PVE·PVP 공통 적용"),
    ("받는 피해 감소", "DamageDownVaryper",    50,  "%",  "PVE·PVP 공통 적용"),
    ("PVE 피해 증가",  "PVEDamageUpVaryper",  100,  "%",  "PVE 전용, 주는 피해 증가와 곱연산"),
    ("PVE 피해 감소",  "PVEDamageDownVaryper", 50,  "%",  "PVE 전용"),
    ("PVP 피해 증가",  "PVPDamageUpVaryper",  100,  "%",  "PVP 전용"),
    ("PVP 피해 감소",  "PVPDamageDownVaryper", 50,  "%",  "PVP 전용"),
    ("치명타 확률",    "CriVaryper",           80,  "%",  "치명타 가능 스킬에만 적용"),
    ("치명타 피해",    "CriDamageVaryper",    300,  "%",  "치명타 배율 = 1 + (확률/100)×(피해/100)"),
    ("공격 속도",      "AtkSpeedVaryper",     300,  "%",  "실질 공속 = 0.91 × (1 + 공격속도%/100)"),
]

# ── 컬러 ──────────────────────────────────────────────────────
BG_DARK  = "0D0D1A"
BG_PANEL = "141428"
BG_HEAD  = "1A1F3A"
BG_SUBH  = "0F2D5A"
BG_SUBH2 = "1A3A6A"
BG_ROW_A = "181830"
BG_ROW_B = "1E2040"
BG_NA    = "111118"
BG_CODE  = "0A1020"

GOLD     = "E8C050"
GOLD2    = "F5D878"
GOLD_DIM = "9A7A28"
CREAM    = "EEE8D8"
DIM      = "7A6A5A"
NA_FG    = "2E2E3A"

TIER_C = {
    "Tier1":"50C870","Tier2":"7DCC50","Tier3":"C8C846",
    "Tier4":"E8C050","Tier5":"E68C32","Tier6":"D44040","Tier7":"A060E8",
}
GRADE_C = {
    "일반":"8A9BB0","고급":"50C870","희귀":"4090F0",
    "고대":"C89040","영웅":"A04BE8","유일":"E84B4B","유물":"E8C050",
}

# ── 헬퍼 ─────────────────────────────────────────────────────
def _s(style, color): return Side(style=style, color=color)
def _thin():  return _s("thin",   "252535")
def _mid():   return _s("medium", "383858")
def _gold():  return _s("medium", GOLD_DIM)
def _none():  return Side(style=None)

def brd(t=True, b=True, l=True, r=True, accent=False):
    s = _gold() if accent else _mid()
    return Border(
        top    = s if t else _none(),
        bottom = s if b else _none(),
        left   = s if l else _none(),
        right  = s if r else _none(),
    )

def brd_thin():
    return Border(top=_thin(), bottom=_thin(), left=_thin(), right=_thin())

def fill(c): return PatternFill("solid", fgColor=c)

def cx(ws, row, col, value="", bg=BG_PANEL, fg=CREAM,
       bold=False, size=10, h="center", v="center",
       wrap=False, border=None, italic=False):
    c = ws.cell(row=row, column=col)
    # ⚠ '=' 으로 시작하는 문자열은 엑셀이 수식으로 인식 → 앞에 공백 삽입
    if isinstance(value, str) and value.startswith("="):
        value = "\u200b" + value   # 폭 없는 공백(ZWS) 선행 → 텍스트로 강제
    c.value = value
    c.fill  = fill(bg)
    c.font  = Font(color=fg, bold=bold, size=size,
                   name="맑은 고딕", italic=italic)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border is not None:
        c.border = border
    return c

def row_h(ws, r, h): ws.row_dimensions[r].height = h
def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def no_grid(ws): ws.sheet_view.showGridLines = False

def title_row(ws, row, text, span_end_col, size=13):
    """전체 너비 타이틀 행"""
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 32)
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = fill(BG_HEAD)
    c.font  = Font(color=GOLD2, bold=True, size=size, name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1, wrap_text=False)
    c.border = Border(bottom=_gold())

def section_bar(ws, row, text, span_end_col, sub=False):
    """섹션 구분 바"""
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 22)
    bg = BG_SUBH2 if sub else BG_SUBH
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = fill(bg)
    c.font  = Font(color=GOLD, bold=True, size=11, name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=_s("medium", GOLD_DIM),
                      top=_s("thin","303050"))

def blank_row(ws, row, span_end_col):
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 8)
    ws.cell(row=row, column=1).fill = fill(BG_DARK)

def is_avail(tier, grade):
    return GRADE_ORDER[grade] <= GRADE_ORDER[TIER_MAX_GRADE[tier]]

# ════════════════════════════════════════════════════════════════
# SHEET 1  설계 의도
# ════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)
ws1 = wb.create_sheet("📋 설계 의도")
no_grid(ws1)

col_w(ws1, 1, 24)
col_w(ws1, 2, 70)
for c in range(3, 9): col_w(ws1, c, 4)

SPAN = 2
title_row(ws1, 1, "⚔  장비 능력치 — 설계 의도", SPAN, size=14)

INTENT = [
    # (섹션헤더, None) | (라벨, 내용)
    ("◆ 핵심 설계 원칙", None),
    ("티어 차등 강조",
     "티어 간 기본 수치 격차를 크게 설계\n"
     "→ 상위 티어 장비 확보 자체가 핵심 성장 목표가 되도록 함"),
    ("티어별 강화 보상 차등",
     "강화 1단계 증가량을 티어가 높을수록 크게 설정  (T1: +8 / T7: +20)\n"
     "→ 고티어 장비일수록 강화 시 더 큰 수치 상승 경험"),
    ("강화 리스크 보장",
     "강화 증가폭이 클수록 실패 시 손실도 커짐\n"
     "→ 고티어 강화는 자원 투자 대비 리스크가 명확히 존재"),
    ("등급 수치 동일, 상한으로 가치 차별화",
     "등급 보너스 자체는 티어 무관 동일  (예: 희귀 +40 ATK — 모든 티어 공통)\n"
     "단, 티어별 최대 등급 제한으로 실질 등급 가치 차별화\n"
     "  T1 최대 고급  /  T2 최대 희귀  /  T3 최대 고대  /  T4 최대 영웅\n"
     "  T5 최대 유일  /  T6 · T7 최대 유물"),
    ("캐릭터 기본 능력치 최소화",
     "캐릭터는 HP · MP 기본값만 보유 (레벨업 성장)\n"
     "→ 공격력 · 방어력은 100% 장비에서 획득  →  장비 수집·강화가 성장의 핵심 동력"),
    ("◆ 수치 구조 요약", None),
    ("무기 기본 ATK (0강 · 일반)",
     "T1: 60  /  T2: 80  /  T3: 120  /  T4: 180  /  T5: 260  /  T6: 360  /  T7: 480"),
    ("방어구 기본 DEF (0강 · 일반)",
     "T1: 30  /  T2: 40  /  T3: 60   /  T4: 90   /  T5: 130  /  T6: 180  /  T7: 240"),
    ("무기 강화 증가 / 단계",
     "T1: +8  /  T2: +10  /  T3: +12  /  T4: +14  /  T5: +16  /  T6: +18  /  T7: +20"),
    ("방어구 강화 증가 / 단계",
     "T1: +4  /  T2: +5   /  T3: +6   /  T4: +7   /  T5: +8   /  T6: +9   /  T7: +10"),
    ("최대 강화 단계", "10강"),
    ("악세사리 (T1·T2)", "반지 · 목걸이 · 귀걸이는 T3부터 장착 가능  (T1·T2 슬롯 없음)"),
]

r = 3
for item in INTENT:
    label, desc = item
    if desc is None:
        # 섹션 헤더
        blank_row(ws1, r, SPAN); r += 1
        section_bar(ws1, r, label, SPAN); r += 1
        continue

    lines = desc.count("\n") + 1
    row_h(ws1, r, max(18, lines * 15 + 4))

    cx(ws1, r, 1, label, bg=BG_HEAD, fg=CREAM, bold=True,
       h="left", v="top", size=10, border=brd())

    c = ws1.cell(row=r, column=2, value=desc)
    bg = BG_ROW_A if r % 2 == 0 else BG_ROW_B
    c.fill      = fill(bg)
    c.font      = Font(color=CREAM, size=10, name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.border    = brd_thin()
    r += 1

blank_row(ws1, r, SPAN)


# ════════════════════════════════════════════════════════════════
# 공통: 장비 스탯 시트 생성
# ════════════════════════════════════════════════════════════════
# 열 레이아웃
# A=티어(5)  B=최고등급(8)  C~I=등급별 0강(9×7=63)  +빈열  J~T=0강~10강(8×11=88)  U=+/단계(10)
COL_TIER   = 1   # A
COL_MAXG   = 2   # B
COL_G0     = 3   # C  (등급 7개: C~I)
COL_ENH0   = 11  # K  (강화 0강 시작)
COL_STEP   = COL_ENH0 + MAX_ENH + 1   # 강화증가/단계
COL_GAP    = 10  # J — 빈 구분열
COL_MAX3   = 6   # 격차 표: 5열

TOTAL_COLS = COL_STEP  # 마지막 열

def setup_equip_cols(ws):
    no_grid(ws)
    col_w(ws, COL_TIER,  5)
    col_w(ws, COL_MAXG,  9)
    for gi in range(7):
        col_w(ws, COL_G0 + gi, 8)
    col_w(ws, COL_GAP, 2)          # 빈 구분열
    for e in range(MAX_ENH + 1):
        col_w(ws, COL_ENH0 + e, 7)
    col_w(ws, COL_STEP, 10)

def write_equip_sheet(ws, title, icon, base_d, grade_d, enh_d, stat_nm):
    setup_equip_cols(ws)
    title_row(ws, 1, f"{icon}  {title} — {stat_nm} 수치표", TOTAL_COLS, size=13)

    # ─── 등급 색상 범례 ───────────────────────────────
    r = 3
    section_bar(ws, r, "등급 색상 범례", TOTAL_COLS, sub=True); r += 1
    row_h(ws, r, 16)
    ws.merge_cells(f"A{r}:B{r}")
    cx(ws, r, 1, "등급 →", bg=BG_HEAD, fg=DIM, size=9, h="right")
    for gi, g in enumerate(GRADES):
        cx(ws, r, COL_G0 + gi, g, bg=BG_HEAD, fg=GRADE_C[g],
           bold=True, size=9, border=brd_thin())
    # 강화 열 범례
    cx(ws, r, COL_ENH0, "0강", bg=BG_HEAD, fg=DIM, size=9, border=brd_thin())
    ws.merge_cells(f"{get_column_letter(COL_ENH0+1)}{r}:{get_column_letter(COL_ENH0+MAX_ENH-1)}{r}")
    cx(ws, r, COL_ENH0+1, "← 강화 단계 →", bg=BG_HEAD, fg=DIM, size=9, border=brd_thin())
    cx(ws, r, COL_ENH0+MAX_ENH, f"+{MAX_ENH}강", bg=BG_HEAD, fg=GOLD,
       bold=True, size=9, border=brd())
    cx(ws, r, COL_STEP, "+/단계", bg=BG_HEAD, fg=DIM, size=9, border=brd_thin())
    r += 1

    # ─── 섹션 A: 0강 티어×등급 표 ──────────────────────
    blank_row(ws, r, TOTAL_COLS); r += 1
    section_bar(ws, r, f"▶ [A]  0강 기준  티어 × 등급  {stat_nm} 표"
                f"  (회색 셀 = 해당 티어에서 획득 불가 등급)", TOTAL_COLS); r += 1

    row_h(ws, r, 18)
    cx(ws, r, COL_TIER, "티어", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    cx(ws, r, COL_MAXG, "최대등급", bg=BG_HEAD, fg=GOLD_DIM, size=9, border=brd())
    for gi, g in enumerate(GRADES):
        cx(ws, r, COL_G0+gi, g, bg=BG_HEAD, fg=GRADE_C[g], bold=True, border=brd())
    # 강화 열 헤더도 표시
    for e in ENH_RANGE:
        fg_ = GOLD if e == MAX_ENH else CREAM
        cx(ws, r, COL_ENH0+e, f"+{e}강", bg=BG_HEAD, fg=fg_,
           bold=(e==MAX_ENH), size=9, border=brd(accent=(e==MAX_ENH)))
    cx(ws, r, COL_STEP, "+/단계", bg=BG_HEAD, fg=DIM, size=9, border=brd())
    r += 1

    for ti, tier in enumerate(TIERS):
        row_h(ws, r, 16)
        max_g  = TIER_MAX_GRADE[tier]
        tc     = TIER_C[tier]
        row_bg = BG_ROW_A if ti % 2 == 0 else BG_ROW_B

        cx(ws, r, COL_TIER, TIER_LABELS[ti], bg=BG_HEAD, fg=tc,
           bold=True, border=brd(accent=True))
        cx(ws, r, COL_MAXG, max_g, bg=BG_HEAD, fg=GRADE_C[max_g],
           size=9, border=brd())

        for gi, grade in enumerate(GRADES):
            if is_avail(tier, grade):
                val = base_d[tier] + grade_d[grade]
                cx(ws, r, COL_G0+gi, val, bg=row_bg, fg=CREAM, border=brd_thin())
            else:
                c = ws.cell(row=r, column=COL_G0+gi, value="—")
                c.fill      = fill(BG_NA)
                c.font      = Font(color=NA_FG, size=10, name="맑은 고딕")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = brd_thin()

        # 강화 진행 (최고등급 기준) — 같은 행에 이어서
        base_val = base_d[tier] + grade_d[max_g]
        for e in ENH_RANGE:
            val    = base_val + enh_d[tier] * e
            is_max = (e == MAX_ENH)
            cx(ws, r, COL_ENH0+e, val, bg=row_bg,
               fg=GOLD if is_max else CREAM,
               bold=is_max, border=brd(accent=is_max))

        cx(ws, r, COL_STEP, f"+{enh_d[tier]}", bg=BG_HEAD, fg=DIM,
           size=9, border=brd())
        r += 1

    # ─── 섹션 B: 티어 간 격차 요약 ──────────────────────
    blank_row(ws, r, TOTAL_COLS); r += 1
    section_bar(ws, r, f"▶ [B]  티어 간 격차 요약  (최고등급 {MAX_ENH}강 기준)", TOTAL_COLS); r += 1

    row_h(ws, r, 18)
    for ci, h in enumerate(["티어","최고등급",f"최대 {stat_nm}","전 티어 대비 +","배율","비고"]):
        cx(ws, r, 1+ci, h, bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    r += 1

    prev = None
    for ti, tier in enumerate(TIERS):
        row_h(ws, r, 16)
        max_g   = TIER_MAX_GRADE[tier]
        max_val = base_d[tier] + grade_d[max_g] + enh_d[tier] * MAX_ENH
        diff    = max_val - prev if prev else "—"
        ratio   = f"×{max_val/prev:.2f}" if prev else "기준"
        note    = f"최고등급: {max_g}"
        row_bg  = BG_ROW_A if ti % 2 == 0 else BG_ROW_B

        cx(ws, r, 1, TIER_LABELS[ti], bg=BG_HEAD, fg=TIER_C[tier], bold=True, border=brd())
        cx(ws, r, 2, max_g, bg=BG_HEAD, fg=GRADE_C[max_g], size=9, border=brd())
        cx(ws, r, 3, max_val, bg=row_bg, fg=CREAM, bold=True, border=brd())
        cx(ws, r, 4, diff, bg=row_bg,
           fg=GOLD_DIM if diff != "—" else DIM, border=brd())
        cx(ws, r, 5, ratio, bg=row_bg,
           fg=GOLD if ratio != "기준" else DIM, border=brd())
        cx(ws, r, 6, note, bg=row_bg, fg=DIM, size=9, h="left", border=brd_thin())
        prev = max_val
        r += 1


ws2 = wb.create_sheet("⚔ 무기 공격력")
write_equip_sheet(ws2, "무기", "⚔", WPN_BASE, WPN_GRADE, WPN_ENH, "ATK")

ws3 = wb.create_sheet("🛡 방어구 방어력")
write_equip_sheet(ws3, "방어구", "🛡", ARM_BASE, ARM_GRADE, ARM_ENH, "DEF")


# ════════════════════════════════════════════════════════════════
# SHEET 4  악세사리
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("💍 악세사리")
no_grid(ws4)

col_w(ws4, 1,  10)   # 슬롯
col_w(ws4, 2,   5)   # 티어
col_w(ws4, 3,   9)   # 최고등급
for e in range(MAX_ENH + 1):
    col_w(ws4, 4 + e, 7)
col_w(ws4, 4 + MAX_ENH + 1, 9)  # 최대값
col_w(ws4, 4 + MAX_ENH + 2, 8)  # +/단계
SPAN4 = 4 + MAX_ENH + 2

title_row(ws4, 1, "💍  악세사리 — 능력치 수치표", SPAN4, size=13)

ACC_SPECS = [
    ("💍 반지", "ATK", RING_BASE, RING_GRADE, RING_ENH,
     "×2 장착 (ATK 합산) · T1·T2 장착 불가"),
    ("📿 목걸이", "DEF", NECK_BASE, NECK_GRADE, NECK_ENH,
     "DEF 제공 · T1·T2 장착 불가"),
    ("👂 귀걸이", "DEF", EAR_BASE, EAR_GRADE, EAR_ENH,
     "DEF 제공 · T1·T2 장착 불가"),
]

r = 3
for slot_name, stat_nm, base_d, grade_d, enh_d, note in ACC_SPECS:
    blank_row(ws4, r, SPAN4); r += 1
    section_bar(ws4, r, f"{slot_name}  —  {note}", SPAN4); r += 1

    # 컬럼 헤더
    row_h(ws4, r, 18)
    cx(ws4, r, 1, "슬롯", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    cx(ws4, r, 2, "티어", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    cx(ws4, r, 3, "최고등급", bg=BG_HEAD, fg=GOLD_DIM, size=9, border=brd())
    for e in ENH_RANGE:
        fg_ = GOLD if e == MAX_ENH else CREAM
        cx(ws4, r, 4+e, f"+{e}강", bg=BG_HEAD, fg=fg_,
           bold=(e==MAX_ENH), size=9, border=brd(accent=(e==MAX_ENH)))
    cx(ws4, r, 4+MAX_ENH+1, f"최대{stat_nm}", bg=BG_HEAD, fg=GOLD,
       bold=True, border=brd(accent=True))
    cx(ws4, r, 4+MAX_ENH+2, "+/단계", bg=BG_HEAD, fg=DIM, size=9, border=brd())
    r += 1

    first_row = r
    for ti, tier in enumerate(TIERS):
        row_h(ws4, r, 16)
        max_g   = TIER_MAX_GRADE[tier]
        is_na   = (base_d[tier] == 0)
        row_bg  = BG_ROW_A if ti % 2 == 0 else BG_ROW_B

        # 슬롯열 — 첫 티어만 슬롯명 표시, 나머지 병합
        if ti == 0:
            ws4.merge_cells(f"A{first_row}:A{first_row+6}")
            cx(ws4, first_row, 1, slot_name, bg=BG_HEAD, fg=GOLD,
               bold=True, v="center", border=brd())

        cx(ws4, r, 2, TIER_LABELS[ti], bg=BG_HEAD, fg=TIER_C[tier],
           bold=True, border=brd())

        if is_na:
            cx(ws4, r, 3, "없음", bg=BG_NA, fg=NA_FG, size=9, border=brd_thin())
            for e in ENH_RANGE:
                c = ws4.cell(row=r, column=4+e, value="N/A")
                c.fill = fill(BG_NA)
                c.font = Font(color=NA_FG, size=9, name="맑은 고딕")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = brd_thin()
            cx(ws4, r, 4+MAX_ENH+1, "—", bg=BG_NA, fg=NA_FG, border=brd_thin())
            cx(ws4, r, 4+MAX_ENH+2, "—", bg=BG_NA, fg=NA_FG, size=9, border=brd_thin())
        else:
            cx(ws4, r, 3, max_g, bg=BG_HEAD, fg=GRADE_C[max_g], size=9, border=brd())
            base_val = base_d[tier] + grade_d[max_g]
            for e in ENH_RANGE:
                val    = base_val + enh_d[tier] * e
                is_max = (e == MAX_ENH)
                cx(ws4, r, 4+e, val, bg=row_bg,
                   fg=GOLD if is_max else CREAM,
                   bold=is_max, border=brd(accent=is_max))
            cx(ws4, r, 4+MAX_ENH+1, base_val + enh_d[tier]*MAX_ENH,
               bg=row_bg, fg=GOLD, bold=True, border=brd(accent=True))
            cx(ws4, r, 4+MAX_ENH+2, f"+{enh_d[tier]}", bg=BG_HEAD,
               fg=DIM, size=9, border=brd())
        r += 1


# ════════════════════════════════════════════════════════════════
# SHEET 5  스탯 캡 & 공식
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📌 스탯 캡 & 공식")
no_grid(ws5)

col_w(ws5, 1, 22)   # 능력치 이름
col_w(ws5, 2,  8)   # 캡 값
col_w(ws5, 3,  8)   # 단위
col_w(ws5, 4, 28)   # 내부 키
col_w(ws5, 5, 52)   # 비고
SPAN5 = 5

title_row(ws5, 1, "📌  스탯 캡 & 데미지 공식", SPAN5, size=13)

# ─── 스탯 캡 표 ─────────────────────────────────────────────
r = 3
section_bar(ws5, r, "◆ 능력치 상한값 (CAP)", SPAN5); r += 1

row_h(ws5, r, 18)
for ci, h in enumerate(["능력치", "상한", "단위", "내부 키", "설명 / 공식"]):
    cx(ws5, r, 1+ci, h, bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
r += 1

for idx, (kr, key, cap, unit, note) in enumerate(CAPS):
    row_h(ws5, r, 16)
    bg = BG_ROW_A if idx % 2 == 0 else BG_ROW_B
    cx(ws5, r, 1, kr,   bg=bg,     fg=CREAM,    bold=True, h="left",  border=brd())
    cx(ws5, r, 2, cap,  bg=BG_HEAD, fg=GOLD2,   bold=True,            border=brd(accent=True))
    cx(ws5, r, 3, unit, bg=BG_HEAD, fg=DIM,     size=9,               border=brd())
    cx(ws5, r, 4, key,  bg=bg,     fg=DIM,      size=9,  h="left",    border=brd_thin())
    cx(ws5, r, 5, note, bg=bg,     fg=CREAM,    size=9,  h="left",    border=brd_thin())
    r += 1

# ─── 공식 구역 ──────────────────────────────────────────────
blank_row(ws5, r, SPAN5); r += 1
section_bar(ws5, r, "◆ 공식 정의", SPAN5); r += 1

# 공식 헤더
row_h(ws5, r, 18)
cx(ws5, r, 1, "항목", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
ws5.merge_cells(f"B{r}:E{r}")
cx(ws5, r, 2, "공식 / 설명", bg=BG_HEAD, fg=GOLD, bold=True,
   h="left", border=brd(accent=True))
r += 1

# ⚠ 중요: '='로 시작하는 문자열은 cx() 내부에서 ZWS 삽입으로 수식 오인 방지
FORMULAS = [
    ("데미지 (기본식)",
     "ATK  ×  [ 500 / (DEF + 500) ]\n"
     "     ×  (1 + 피해증가% / 100)  ×  (1 − 피해감소% / 100)\n"
     "     ×  (1 + 모드피해증가% / 100)  ×  (1 − 모드피해감소% / 100)\n"
     "     ×  (스킬계수 / 100)\n"
     "     ×  치명타배율",
     5),

    ("방어 계수  [K=500]",
     "DEF =    0  →  500/(0+500)   = 1.00  (피해 100%)\n"
     "DEF =  500  →  500/(500+500) = 0.50  (피해  50%)\n"
     "DEF = 1000  →  500/(1000+500)= 0.33  (피해  33%)",
     3),

    ("치명타 배율  (치명타 가능 스킬만)",
     "배율  =  1  +  (치명타확률 / 100)  ×  (치명타피해 / 100)\n"
     "예)  확률 20%,  피해 50%  →  배율 = 1 + 0.20 × 0.50 = 1.10",
     2),

    ("스킬 가속 적용 CD",
     "유효CD  =  기본CD  ×  100 / (100 + 스킬가속)\n"
     "예)  기본CD 12s,  가속 25  →  12 × 100/125  =  9.6 s",
     2),

    ("최종 ATK",
     "(무기ATK  +  반지1ATK  +  반지2ATK)\n"
     "    ×  (1 + 공격력% / 100)   +   추가공격력",
     2),

    ("최종 DEF",
     "(투구DEF  +  갑옷DEF  +  장갑DEF  +  신발DEF  +  목걸이DEF  +  귀걸이DEF)\n"
     "    ×  (1 + 방어력% / 100)   +   추가방어력",
     2),

    ("캐릭터 최대 HP  (레벨 n)",
     "1500  +  (n − 1) × 100        [레벨 1: 1500  /  레벨 100: 11400]",
     1),

    ("캐릭터 최대 MP  (레벨 n)",
     "360  +  (n − 1) × 20          [레벨 1: 360   /  레벨 100: 2340]",
     1),

    ("MP 자연 회복",
     "15초마다  +3 MP  회복  (RegenMpVary = 3,  전 레벨 동일)",
     1),

    ("HP 자연 회복",
     "전투 외  5초마다  +5 HP  회복  (RegenHpVary = 5,  전 레벨 동일)",
     1),
]

for idx, row_data in enumerate(FORMULAS):
    label, formula, line_count = row_data
    row_h(ws5, r, max(18, line_count * 16 + 6))
    bg = BG_ROW_A if idx % 2 == 0 else BG_ROW_B

    cx(ws5, r, 1, label, bg=BG_HEAD, fg=CREAM, bold=True,
       h="left", v="top", border=brd())

    ws5.merge_cells(f"B{r}:E{r}")
    c = ws5.cell(row=r, column=2, value=formula)
    c.fill      = fill(BG_CODE if idx == 0 else bg)
    c.font      = Font(color=GOLD2 if idx == 0 else CREAM,
                       size=10, name="Consolas",
                       bold=(idx == 0))
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.border    = brd(accent=(idx == 0))
    r += 1

# ── 저장 ─────────────────────────────────────────────────────
wb.save(OUT)
print(f"✅ 저장: {OUT}")
