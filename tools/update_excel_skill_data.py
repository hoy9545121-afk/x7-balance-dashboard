"""
X7_무기_스킬.xlsx 업데이트 스크립트
  1. 단검 3T Q 추격자의 발톱: mp / dmg 수정 (v16 밸런스 반영)
  2. 단검 3T W 침묵의 일격: 버프 설명 +35% → +45%
  3. 단검 3T R 유혈의 장막: cd / dmg 수정 (v16 밸런스 반영)
  4. 무기 비교 개요: 활 추가 (설명 행 + 분당횟수 테이블)
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)

# ─── 1. 단검 시트 ─────────────────────────────────────────────────────────────
ws = wb['단검']

# Q 추격자의 발톱 (R77~R81): col E=mp, col G=dmg배율
# constants.py: mp=[35,39,42,46,50], dmg=[220,242,264,286,308] -> 배율=[2.2,2.42,2.64,2.86,3.08]
Q_MP  = [35, 39, 42, 46, 50]
Q_DMG = [2.2, 2.42, 2.64, 2.86, 3.08]
for i, row in enumerate(range(77, 82)):
    ws.cell(row=row, column=5).value = Q_MP[i]   # col E
    ws.cell(row=row, column=7).value = Q_DMG[i]  # col G

print("단검 Q (R77-R81) mp/dmg 수정 완료")

# W 침묵의 일격 (R82~R86): col H=버프설명 — +35% → +45%
W_BUFFS = [
    '공격속도 +45% / 5.0s  (사이클 13.0s)',
    '공격속도 +45% / 5.25s  (사이클 13.0s)',
    '공격속도 +45% / 5.5s  (사이클 13.5s)',
    '공격속도 +45% / 5.75s  (사이클 13.5s)',
    '공격속도 +45% / 6.0s  (사이클 14.0s)',
]
for i, row in enumerate(range(82, 87)):
    ws.cell(row=row, column=8).value = W_BUFFS[i]  # col H

print("단검 W (R82-R86) 버프 설명 +35%→+45% 수정 완료")

# R 유혈의 장막 (R92~R96): col D=cd, col G=dmg배율
# constants.py: cd=[40,38,36,34,32], dmg=[500,550,600,650,700] -> 배율=[5.0,5.5,6.0,6.5,7.0]
R_CD  = [40, 38, 36, 34, 32]
R_DMG = [5.0, 5.5, 6.0, 6.5, 7.0]
for i, row in enumerate(range(92, 97)):
    ws.cell(row=row, column=4).value = R_CD[i]   # col D
    ws.cell(row=row, column=7).value = R_DMG[i]  # col G

print("단검 R (R92-R96) cd/dmg 수정 완료")

# ─── 2. 무기 비교 개요 시트 ────────────────────────────────────────────────────
ws2 = wb['무기 비교 개요']

# 활 설명 행 추가 (R8 — 현재 비어 있음)
ws2.cell(row=8, column=2).value = '활'
ws2.cell(row=8, column=3).value = '원거리 치명타 — 충격 사격·연속 쏘기·급습·발묶음(CC)으로 원거리 딜링과 CC 특화'

print("무기 비교 개요 R8: 활 설명 추가 완료")

# 분당횟수 테이블 (R10~R14): 활 컬럼 추가
# 현재: B=커맨드, C=한손검, D=양손검, E=지팡이, F=단검, G=최고
# 변경: B=커맨드, C=한손검, D=양손검, E=지팡이, F=단검, G=활, H=최고
#
# DPS scores (from HTML DPS_DATA & simulator analysis):
#   활: Q=24, W=13, E=10, R=7
#   새 최고: Q=max(24,24,40,18,24)=40, W=max(0,12.5,16,0,13)=16,
#            E=max(12,12,9,12,10)=12, R=max(6,6,0,0,7)=7

# 헤더 행 (R10)
ws2.cell(row=10, column=7).value = '활'    # 기존 G='최고' → '활'
ws2.cell(row=10, column=8).value = '최고'  # 새 H

# Q 행 (R11): 활=24, 최고=40
ws2.cell(row=11, column=7).value = 24
ws2.cell(row=11, column=8).value = 40

# W 행 (R12): 활=13, 최고=16
ws2.cell(row=12, column=7).value = 13
ws2.cell(row=12, column=8).value = 16

# E 행 (R13): 활=10, 최고=12
ws2.cell(row=13, column=7).value = 10
ws2.cell(row=13, column=8).value = 12

# R 행 (R14): 활=7, 최고 기존 6→7 (활이 더 높으므로)
ws2.cell(row=14, column=7).value = 7
ws2.cell(row=14, column=8).value = 7

print("무기 비교 개요 분당횟수 테이블: 활 열 추가 + 최고 갱신 완료")

# ─── 저장 ────────────────────────────────────────────────────────────────────
wb.save(FILEPATH)
print("\n✅ 저장 완료:", FILEPATH)

# ─── 검증 ────────────────────────────────────────────────────────────────────
print("\n=== 검증 ===")
wb2 = openpyxl.load_workbook(FILEPATH, read_only=True, data_only=True)

ws_d = wb2['단검']
print("\n[ 단검 3T Q R77-R81 ]")
for r in range(77, 82):
    print(f"  R{r}: cd={ws_d.cell(r,4).value}, mp={ws_d.cell(r,5).value}, dmg={ws_d.cell(r,7).value}")

print("\n[ 단검 3T W R82-R86 ]")
for r in range(82, 87):
    print(f"  R{r}: buff={ws_d.cell(r,8).value}")

print("\n[ 단검 3T R R92-R96 ]")
for r in range(92, 97):
    print(f"  R{r}: cd={ws_d.cell(r,4).value}, dmg={ws_d.cell(r,7).value}")

ws_o = wb2['무기 비교 개요']
print("\n[ 무기 비교 개요 R8 ]")
print(f"  R8: {ws_o.cell(8,2).value} / {ws_o.cell(8,3).value}")
print("\n[ 무기 비교 개요 R10-R14 ]")
for r in range(10, 15):
    vals = [ws_o.cell(r, c).value for c in range(2, 9)]
    print(f"  R{r}: {vals}")

wb2.close()
