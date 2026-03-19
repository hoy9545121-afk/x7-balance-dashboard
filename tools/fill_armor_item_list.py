"""
X7_방어구_스킬.xlsx — 아이템 리스트 갱신

수행 내용:
1. 1T/2T/3T 방어구 시트에서 CID별 패시브 추출
2. 아이템 리스트 패시브1(col4) / 패시브2(col5) 채우기
3. 12020008 타입 오류 수정: LeatherGloves → LeatherShoes
4. 전체 CID 커버리지 검증 출력
"""

import re, os
from openpyxl import load_workbook
from copy import copy

# ── 스타일 복사 ────────────────────────────────────────────────
def copy_fmt(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.border      = copy(src.border)
        dst.alignment   = copy(src.alignment)
        dst.number_format = src.number_format

# ── 패시브 포맷: "Key (한국어)" ────────────────────────────────
def fmt_passive(key, ko):
    return f"{str(key).strip()} ({str(ko).strip()})"

# ════════════════════════════════════════════════════════════════
# STEP 1: 방어구 시트 3개에서 CID별 패시브 파싱
#   구조: 각 아이템의 첫 행에 CID, 이후 행(CID=None)은 추가 패시브
# ════════════════════════════════════════════════════════════════
def parse_passives_from_sheet(ws):
    """반환: {CID: [(key, korean), ...]}"""
    cid_passives = {}
    current_cid  = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        cid = row[8]   # col9
        ko  = row[1]   # col2
        key = row[2]   # col3

        # 새 아이템 시작
        if isinstance(cid, int):
            current_cid = cid
            cid_passives[current_cid] = []

        # 유효 패시브 행 (key, ko 모두 있어야 함)
        if current_cid is not None and ko and key:
            cid_passives[current_cid].append((str(key).strip(), str(ko).strip()))

    return cid_passives

wb = load_workbook("기획서/X7_방어구_스킬.xlsx")

all_passives = {}
for sheet in ["1T 방어구", "2T 방어구", "3T 방어구"]:
    parsed = parse_passives_from_sheet(wb[sheet])
    all_passives.update(parsed)
    print(f"[{sheet}] {len(parsed)}개 CID 파싱 완료")

print(f"\n[패시브 맵] 총 {len(all_passives)}개 CID\n")

# ════════════════════════════════════════════════════════════════
# STEP 2: 아이템 리스트 패시브1/패시브2 채우기
#   - 1개 패시브  → col4 단독, col5=None
#   - 2개 패시브  → col4, col5 각각
#   - 3개 패시브  → col4 = passive1, col5 = "passive2passive3" (연결)
# ════════════════════════════════════════════════════════════════
ws_list = wb["아이템 리스트"]

type_fixes   = 0
passive_ok   = 0
passive_miss = 0

for ri in range(2, ws_list.max_row + 1):
    cid = ws_list.cell(row=ri, column=1).value
    if not isinstance(cid, int):
        continue

    # ── STEP 2a: 타입 오류 수정 ──────────────────────────────
    if cid == 12020008:
        old_type = ws_list.cell(row=ri, column=3).value
        if old_type == "LeatherGloves":
            ws_list.cell(row=ri, column=3).value = "LeatherShoes"
            type_fixes += 1
            print(f"  [타입 수정] CID {cid}: LeatherGloves → LeatherShoes")

    # ── STEP 2b: 패시브 채우기 ───────────────────────────────
    passives = all_passives.get(cid, [])

    if not passives:
        passive_miss += 1
        print(f"  ⚠️  CID {cid}: 패시브 데이터 없음")
        continue

    passive_ok += 1

    # 패시브 포맷 문자열 생성
    fmt_list = [fmt_passive(k, ko) for k, ko in passives]

    if len(fmt_list) == 1:
        p1_val = fmt_list[0]
        p2_val = None
    elif len(fmt_list) == 2:
        p1_val = fmt_list[0]
        p2_val = fmt_list[1]
    else:  # 3개 이상
        p1_val = fmt_list[0]
        p2_val = "".join(fmt_list[1:])   # 두 번째 이후 연결

    c1 = ws_list.cell(row=ri, column=4, value=p1_val)
    c2 = ws_list.cell(row=ri, column=5, value=p2_val)

    # 서식: col1 기준으로 복사
    ref_cell = ws_list.cell(row=ri, column=1)
    copy_fmt(ref_cell, c1)
    copy_fmt(ref_cell, c2)

print(f"\n[패시브 채우기] ✅ {passive_ok}개  ❌ 미처리 {passive_miss}개")
if type_fixes:
    print(f"[타입 수정]     ✅ {type_fixes}건")

# ════════════════════════════════════════════════════════════════
# STEP 3: 저장
# ════════════════════════════════════════════════════════════════
TMP = "기획서/X7_방어구_스킬_tmp.xlsx"
wb.save(TMP)
try:
    os.replace(TMP, "기획서/X7_방어구_스킬.xlsx")
    print("\n✅ 저장: X7_방어구_스킬.xlsx")
    LOAD_PATH = "기획서/X7_방어구_스킬.xlsx"
except PermissionError:
    print(f"\n⚠️  Excel 열려있어 임시파일 저장: {TMP}")
    LOAD_PATH = TMP

# ════════════════════════════════════════════════════════════════
# STEP 4: 최종 검증 출력
# ════════════════════════════════════════════════════════════════
wb2  = load_workbook(LOAD_PATH)
ws2  = wb2["아이템 리스트"]

print("\n=== 아이템 리스트 최종 확인 ===")
print(f"{'CID':>10}  {'이름':30s}  {'타입':15s}  {'패시브1':45s}  {'패시브2'}")
print("-" * 140)

for row in ws2.iter_rows(min_row=2, values_only=True):
    cid, name, typ, p1, p2 = (row[i] if i < len(row) else None for i in range(5))
    if not isinstance(cid, int):
        continue
    p1s = str(p1) if p1 else "❌ 없음"
    p2s = str(p2) if p2 else "-"
    print(f"  {cid:>10}  {str(name):30s}  {str(typ):15s}  {p1s:45s}  {p2s}")

# ── 통계 요약 ──────────────────────────────────────────────────
total = no_passive = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if isinstance(row[0], int):
        total += 1
        if not row[3]:
            no_passive += 1

print(f"\n  총 아이템: {total}개  /  패시브 없음: {no_passive}개")
if no_passive == 0:
    print("  ✅ 모든 아이템 패시브 채워짐")
