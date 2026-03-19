"""
tools/rebuild_equip_sheets.py
⚔ 무기 공격력 / 🛡 방어구 방어력 시트를
'등급별 0~10강 전체 표시' 폼으로 재작성 (악세사리 시트 포맷과 동일)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r"C:\AI_simulator\기획서\X7_장비_능력치_기획서.xlsx"

# ── 데이터 ─────────────────────────────────────────────────────
TIERS       = ["Tier1","Tier2","Tier3","Tier4","Tier5","Tier6","Tier7"]
TIER_LABELS = ["T1","T2","T3","T4","T5","T6","T7"]
GRADES      = ["일반","고급","희귀","고대","영웅","유일","유물"]
GRADE_ORDER = {g: i for i, g in enumerate(GRADES)}

TIER_MAX_GRADE = {
    "Tier1":"고급","Tier2":"희귀","Tier3":"고대",
    "Tier4":"영웅","Tier5":"유일","Tier6":"유물","Tier7":"유물",
}

WPN_BASE  = {"Tier1":60,"Tier2":80,"Tier3":120,"Tier4":180,"Tier5":260,"Tier6":360,"Tier7":480}
WPN_GRADE = {"일반":0,"고급":20,"희귀":40,"고대":60,"영웅":80,"유일":100,"유물":120}
WPN_ENH   = {"Tier1":8,"Tier2":10,"Tier3":12,"Tier4":14,"Tier5":16,"Tier6":18,"Tier7":20}

ARM_BASE  = {"Tier1":30,"Tier2":40,"Tier3":60,"Tier4":90,"Tier5":130,"Tier6":180,"Tier7":240}
ARM_GRADE = {"일반":0,"고급":10,"희귀":20,"고대":30,"영웅":40,"유일":50,"유물":55}
ARM_ENH   = {"Tier1":4,"Tier2":5,"Tier3":6,"Tier4":7,"Tier5":8,"Tier6":9,"Tier7":10}

MAX_ENH   = 10
ENH_RANGE = list(range(MAX_ENH + 1))

# ── 컬러 ──────────────────────────────────────────────────────
BG_DARK  = "0D0D1A"
BG_PANEL = "141428"
BG_HEAD  = "1A1F3A"
BG_SUBH  = "0F2D5A"
BG_SUBH2 = "1A3A6A"
BG_ROW_A = "181830"
BG_ROW_B = "1E2040"
BG_NA    = "111118"

GOLD     = "E8C050"
GOLD2    = "F5D878"
GOLD_DIM = "9A7A28"
CREAM    = "EEE8D8"
DIM      = "7A6A5A"
NA_FG    = "2E2E3A"

TIER_C = {
    "Tier1":"50C870","Tier2":"7DCC50","Tier3":"C8C846",
    "Tier4":"E8C050","Tier5":"E68C32","Tier6":"D44040","Tier7":"A060E8",
}
GRADE_C = {
    "일반":"8A9BB0","고급":"50C870","희귀":"4090F0",
    "고대":"C89040","영웅":"A04BE8","유일":"E84B4B","유물":"E8C050",
}

# ── 헬퍼 ──────────────────────────────────────────────────────
def _s(style, color): return Side(style=style, color=color)
def _thin():  return _s("thin",   "252535")
def _mid():   return _s("medium", "383858")
def _gold():  return _s("medium", GOLD_DIM)
def _none():  return Side(style=None)

def brd(t=True, b=True, l=True, r=True, accent=False):
    s = _gold() if accent else _mid()
    return Border(top=s if t else _none(), bottom=s if b else _none(),
                  left=s if l else _none(), right=s if r else _none())

def brd_thin():
    return Border(top=_thin(), bottom=_thin(), left=_thin(), right=_thin())

def fill(c): return PatternFill("solid", fgColor=c)

def cx(ws, row, col, value="", bg=BG_PANEL, fg=CREAM,
       bold=False, size=10, h="center", v="center",
       wrap=False, border=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(value, str) and value.startswith("="):
        value = "\u200b" + value
    cell.value = value
    cell.fill  = fill(bg)
    cell.font  = Font(color=fg, bold=bold, size=size, name="맑은 고딕")
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border is not None:
        cell.border = border
    return cell

def row_h(ws, r, h): ws.row_dimensions[r].height = h
def col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def no_grid(ws): ws.sheet_view.showGridLines = False

def title_row(ws, row, text, span_end_col, size=13):
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 32)
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = fill(BG_HEAD)
    c.font  = Font(color=GOLD2, bold=True, size=size, name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=_gold())

def section_bar(ws, row, text, span_end_col, sub=False):
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 22)
    bg = BG_SUBH2 if sub else BG_SUBH
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = fill(bg)
    c.font  = Font(color=GOLD, bold=True, size=11, name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=_s("medium", GOLD_DIM), top=_s("thin","303050"))

def blank_row(ws, row, span_end_col):
    ws.merge_cells(f"A{row}:{get_column_letter(span_end_col)}{row}")
    row_h(ws, row, 8)
    ws.cell(row=row, column=1).fill = fill(BG_DARK)

def is_avail(tier, grade):
    return GRADE_ORDER[grade] <= GRADE_ORDER[TIER_MAX_GRADE[tier]]

# ── 레이아웃 상수 ─────────────────────────────────────────────
# A=티어(merged)  B=등급  C~M=+0강~+10강  N=+/단계
COL_TIER  = 1
COL_GRADE = 2
COL_ENH0  = 3                           # C
COL_STEP  = COL_ENH0 + MAX_ENH + 1     # N = 14
SPAN      = COL_STEP


def write_equip_sheet_new(ws, title, icon, base_d, grade_d, enh_d, stat_nm):
    no_grid(ws)
    col_w(ws, COL_TIER,  5)
    col_w(ws, COL_GRADE, 9)
    for e in range(MAX_ENH + 1):
        col_w(ws, COL_ENH0 + e, 7)
    col_w(ws, COL_STEP, 9)

    # ── 타이틀 ──────────────────────────────────────────────
    title_row(ws, 1, f"{icon}  {title} — {stat_nm} 수치표", SPAN, size=13)

    # ── 등급 색상 범례 ───────────────────────────────────────
    r = 3
    section_bar(ws, r, "등급 색상 범례", SPAN, sub=True); r += 1
    row_h(ws, r, 16)
    cx(ws, r, COL_TIER,  "등급 →", bg=BG_HEAD, fg=DIM, size=9, h="right")
    cx(ws, r, COL_GRADE, "",        bg=BG_HEAD)
    for gi, g in enumerate(GRADES):
        cx(ws, r, COL_ENH0 + gi, g, bg=BG_HEAD, fg=GRADE_C[g],
           bold=True, size=9, border=brd_thin())
    for c in range(COL_ENH0 + 7, SPAN + 1):   # 빈 나머지 열
        cx(ws, r, c, "", bg=BG_HEAD)
    r += 1

    # ── 섹션 A: 티어×등급×강화 전체 표 ─────────────────────
    blank_row(ws, r, SPAN); r += 1
    section_bar(ws, r,
        f"▶  티어 × 등급 × 강화  {stat_nm} 수치표  (★ = 해당 티어 최고 등급)",
        SPAN); r += 1

    # 헤더
    row_h(ws, r, 18)
    cx(ws, r, COL_TIER,  "티어", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    cx(ws, r, COL_GRADE, "등급", bg=BG_HEAD, fg=GOLD, bold=True, border=brd(accent=True))
    for e in ENH_RANGE:
        fg_ = GOLD if e == MAX_ENH else CREAM
        cx(ws, r, COL_ENH0 + e, f"+{e}강", bg=BG_HEAD, fg=fg_,
           bold=(e == MAX_ENH), size=9, border=brd(accent=(e == MAX_ENH)))
    cx(ws, r, COL_STEP, "+/단계", bg=BG_HEAD, fg=DIM, size=9, border=brd())
    r += 1

    # 데이터 행 — 티어별, 등급별
    for ti, tier in enumerate(TIERS):
        max_g    = TIER_MAX_GRADE[tier]
        avail_g  = [g for g in GRADES if is_avail(tier, g)]
        n_grades = len(avail_g)
        row_bg   = BG_ROW_A if ti % 2 == 0 else BG_ROW_B
        first_r  = r

        for gi_local, grade in enumerate(avail_g):
            row_h(ws, r, 16)
            is_max_g = (grade == max_g)

            # 티어 셀: 첫 등급 행에만 작성, 나머지는 세로 병합
            if gi_local == 0:
                if n_grades > 1:
                    ws.merge_cells(
                        f"A{first_r}:A{first_r + n_grades - 1}")
                cx(ws, first_r, COL_TIER, TIER_LABELS[ti],
                   bg=BG_HEAD, fg=TIER_C[tier],
                   bold=True, v="center", border=brd(accent=True))

            # 등급 셀
            grade_lbl = f"★{grade}" if is_max_g else grade
            cx(ws, r, COL_GRADE, grade_lbl,
               bg=BG_HEAD, fg=GRADE_C[grade],
               bold=is_max_g, size=9, border=brd())

            # 강화 수치
            base_val = base_d[tier] + grade_d[grade]
            for e in ENH_RANGE:
                val      = base_val + enh_d[tier] * e
                is_max_e = (e == MAX_ENH)
                if is_max_g and is_max_e:
                    fg_ = GOLD;     bold_ = True;  acc = True
                elif is_max_g:
                    fg_ = GOLD_DIM; bold_ = False; acc = False
                elif is_max_e:
                    fg_ = CREAM;    bold_ = False; acc = True
                else:
                    fg_ = CREAM;    bold_ = False; acc = False
                cx(ws, r, COL_ENH0 + e, val,
                   bg=row_bg, fg=fg_, bold=bold_,
                   border=brd(accent=acc))

            # +/단계
            cx(ws, r, COL_STEP, f"+{enh_d[tier]}",
               bg=BG_HEAD, fg=DIM, size=9, border=brd())
            r += 1

        # 티어 간 구분선
        blank_row(ws, r, SPAN); r += 1

    # ── 섹션 B: 티어 간 격차 요약 ────────────────────────────
    section_bar(ws, r,
        f"▶  티어 간 격차 요약  (★최고등급 +{MAX_ENH}강 기준)",
        SPAN); r += 1

    row_h(ws, r, 18)
    for ci, h in enumerate(["티어","최고등급",
                             f"최대 {stat_nm}","전 티어 대비 +","배율","비고"]):
        cx(ws, r, 1 + ci, h, bg=BG_HEAD, fg=GOLD, bold=True,
           border=brd(accent=True))
    r += 1

    prev = None
    for ti, tier in enumerate(TIERS):
        row_h(ws, r, 16)
        max_g   = TIER_MAX_GRADE[tier]
        max_val = base_d[tier] + grade_d[max_g] + enh_d[tier] * MAX_ENH
        diff    = max_val - prev if prev else "—"
        ratio   = f"×{max_val/prev:.2f}" if prev else "기준"
        note    = f"최고등급: {max_g}"
        row_bg  = BG_ROW_A if ti % 2 == 0 else BG_ROW_B

        cx(ws, r, 1, TIER_LABELS[ti], bg=BG_HEAD, fg=TIER_C[tier],
           bold=True, border=brd())
        cx(ws, r, 2, max_g, bg=BG_HEAD, fg=GRADE_C[max_g],
           size=9, border=brd())
        cx(ws, r, 3, max_val, bg=row_bg, fg=CREAM, bold=True, border=brd())
        cx(ws, r, 4, diff, bg=row_bg,
           fg=GOLD_DIM if diff != "—" else DIM, border=brd())
        cx(ws, r, 5, ratio, bg=row_bg,
           fg=GOLD if ratio != "기준" else DIM, border=brd())
        cx(ws, r, 6, note, bg=row_bg, fg=DIM, size=9,
           h="left", border=brd_thin())
        prev = max_val
        r += 1


# ── 기존 파일 로드 → 두 시트만 교체 ─────────────────────────
wb = openpyxl.load_workbook(FILE)
print("현재 시트:", wb.sheetnames)

for sn in ["⚔ 무기 공격력", "🛡 방어구 방어력"]:
    if sn in wb.sheetnames:
        del wb[sn]

# 순서: 📋(0) → ⚔(1) → 🛡(2) → 💍 → 📌 → ⚡
ws_wpn = wb.create_sheet("⚔ 무기 공격력",  1)
ws_arm = wb.create_sheet("🛡 방어구 방어력", 2)

write_equip_sheet_new(ws_wpn, "무기",   "⚔", WPN_BASE, WPN_GRADE, WPN_ENH, "ATK")
write_equip_sheet_new(ws_arm, "방어구", "🛡", ARM_BASE, ARM_GRADE, ARM_ENH, "DEF")

wb.save(FILE)
print("✅ 완료:", FILE)
print("최종 시트 순서:", wb.sheetnames)
