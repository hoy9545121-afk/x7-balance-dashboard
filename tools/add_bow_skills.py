"""
X7_무기_스킬..xlsx 에 활(Bow) 시트 추가 + 스킬 리스트 시트에 활 행 추가
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import copy

EXCEL_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬..xlsx"

# ── 색상 팔레트 (활 전용 녹색 계열) ──────────────────────────────
BOW_DARK   = "1A3A0A"   # 어두운 녹색 (헤더 배경)
BOW_MID    = "2D6014"   # 중간 녹색 (섹션 타이틀)
BOW_LIGHT  = "C6EFCE"   # 연한 녹색 (행 배경)
BOW_ACCENT = "375623"   # 진한 녹색 (강조 텍스트)
GOLD       = "C9A227"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F5F5F5"
GRAY_HEAD  = "D9D9D9"

def thin_border():
    s = Side(style='thin', color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="맑은 고딕", size=size, bold=bold, color=color)

def normal_font(size=10, bold=False, color="000000"):
    return Font(name="맑은 고딕", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 헬퍼: 셀 쓰기 ────────────────────────────────────────────────
def write(ws, row, col, value, font=None, fill_=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill_:  c.fill = fill_
    if align:  c.alignment = align
    if border: c.border = border
    return c

def merge_write(ws, r1, c1, r2, c2, value, font=None, fill_=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if font:  c.font = font
    if fill_: c.fill = fill_
    if align: c.alignment = align
    return c


def build_bow_sheet(wb):
    """활 시트 생성"""
    # 단검 다음에 삽입
    idx = wb.sheetnames.index("단검") + 1 if "단검" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("활", idx)

    # 열 너비
    col_widths = [6, 14, 20, 12, 8, 10, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 대제목 ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="🏹 활 (Bow) — 무기 스킬 기획서")
    c.font = Font(name="맑은 고딕", size=14, bold=True, color=GOLD)
    c.fill = fill(BOW_DARK)
    c.alignment = center()
    ws.row_dimensions[row].height = 28
    row += 1

    # ── 컨셉 설명 ─────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="컨셉: 원거리 저격·함정(1T) / 아군 지원·기동(2T) / 미구현(3T)  ·  CC 연쇄와 다중 발사체로 딜링+견제 동시 수행")
    c.font = Font(name="맑은 고딕", size=10, italic=True, color="C0C0C0")
    c.fill = fill(BOW_DARK)
    c.alignment = left()
    ws.row_dimensions[row].height = 18
    row += 1

    row += 1  # 빈 행

    # ── 공통 헤더 행 (재사용) ─────────────────────────────────────
    headers = ["커맨드", "스킬명", "유형", "쿨타임", "마나(Lv1)", "피해량(배율)", "특수 효과", "비고"]

    def section_header(ws, row, title, fill_color=BOW_MID):
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
        c.fill = fill(fill_color)
        c.alignment = center()
        ws.row_dimensions[row].height = 22
        return row + 1

    def col_headers(ws, row):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name="맑은 고딕", size=10, bold=True, color=BOW_ACCENT)
            c.fill = fill(BOW_LIGHT)
            c.alignment = center()
            c.border = thin_border()
        ws.row_dimensions[row].height = 18
        return row + 1

    def data_row(ws, row, vals, bg=GRAY_LIGHT):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = normal_font()
            c.fill = fill(bg)
            c.alignment = center() if ci != 8 else left()
            c.border = thin_border()
        ws.row_dimensions[row].height = 18
        return row + 1

    # ================================================================
    # 1티어
    # ================================================================
    row = section_header(ws, row, "■ 1티어 (Lv.1만 확보 — 기본 스킬셋)")
    row = col_headers(ws, row)

    skills_1t = [
        # cmd, 스킬명, 유형, cd, mp(Lv1), 피해량, 특수효과, 비고
        ("Q", "저격",         "방향 투사체",        "5s",  30,  "165% (+25% vs 기절대상)",
         "느려짐 → 기절 연쇄",    "GroupId 300401 · 단일/느려진 대상 추가피해"),
        ("W", "부채 사격",    "방향 범위 (부채꼴)", "12s", 30,  "225%",
         "느려짐 + 밀어냄",       "GroupId 300451 · 광역 CC"),
        ("E", "불화살 소나기","위치지정 범위",       "15s", 30,  "60%×5발 = 300% (합산)",
         "—",                     "GroupId 300501 · Pos 타입 · 5회 분할 투사"),
        ("R", "폭탄 화살",    "방향 투사체+폭발",   "50s", 100, "500%×2 = 1000% (합산)",
         "기절",                  "GroupId 300551 · 폭발 2중 타격"),
    ]
    bg_colors = [WHITE, GRAY_LIGHT, WHITE, GRAY_LIGHT]
    for i, s in enumerate(skills_1t):
        row = data_row(ws, row, s, bg_colors[i])

    # 메모
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="※ 1T는 Lv.1만 확보. Q의 느려짐→기절 연쇄: Q 피격 후 느려진 대상을 다시 Q로 타격 시 기절 적용.")
    c.font = Font(name="맑은 고딕", size=9, italic=True, color="666666")
    c.alignment = left()
    row += 2

    # ================================================================
    # 2티어
    # ================================================================
    row = section_header(ws, row, "■ 2티어 (Lv.1만 확보 — 아군 지원 + 기동)")
    row = col_headers(ws, row)

    skills_2t = [
        ("Q", "섬광 화살",    "방향 투사체 (양방향)", "4s",  25, "150% (적) / 강화평타 200% (아군)",
         "아군: 공속↑·방어↑ 버프 / 적: 피해",  "GroupId 300601 · 아군에 맞으면 강화평타 발동"),
        ("W", "천상의 그물",  "방향 범위",             "10s", 30, "200%",
         "느려짐",                               "GroupId 300651 · 광역 그물"),
        ("E", "재정비 도약",  "이동 (아군 대상)",      "0s(서브)", 15, "0% (피해 없음)",
         "이동 + 이속 버프",                     "GroupId 300701 · 쿨다운 없음 서브 스킬"),
        ("R", "금빛 성광",    "방향 공격",             "30s", 80, "400%",
         "캐스팅 1s + 보호막",                   "GroupId 300751 · 채널링 1초 후 발사"),
    ]
    for i, s in enumerate(skills_2t):
        row = data_row(ws, row, s, bg_colors[i])

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="※ 2T는 Lv.1만 확보. 섬광 화살(Q)은 아군/적 양방향 투사체: 아군 관통 시 강화평타 발동, 적 타격 시 피해.")
    c.font = Font(name="맑은 고딕", size=9, italic=True, color="666666")
    c.alignment = left()
    row += 2

    # ================================================================
    # 3티어
    # ================================================================
    row = section_header(ws, row, "■ 3티어 — 미구현 (계획 중)", fill_color="888888")

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="3T 활 스킬(GroupId 300801~300951)은 아직 데이터가 없습니다. 추후 기획 및 구현 예정.")
    c.font = Font(name="맑은 고딕", size=10, italic=True, color="555555")
    c.fill = fill("EEEEEE")
    c.alignment = left()
    ws.row_dimensions[row].height = 20
    row += 2

    # ================================================================
    # 스킬 계수 요약 표 (참고)
    # ================================================================
    row = section_header(ws, row, "▶ 스킬 계수 요약 (참고 — 1T/2T Lv.1 기준)")
    summary_headers = ["티어", "커맨드", "스킬명", "CD", "MP", "피해 계수(×)", "유형", "특이사항"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color=BOW_ACCENT)
        c.fill = fill(BOW_LIGHT)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    summary_data = [
        ("1T","Q","저격",        5,  30, 1.65, "방향","조건부 +0.25× (기절 대상)"),
        ("1T","W","부채 사격",   12, 30, 2.25, "방향","광역 CC (느려짐+밀어냄)"),
        ("1T","E","불화살 소나기",15,30, 3.00, "Pos", "60%×5발"),
        ("1T","R","폭탄 화살",   50,100,10.00, "방향","500%×2 / 기절"),
        ("2T","Q","섬광 화살",    4, 25, 1.50, "방향","아군 강화평타 200%"),
        ("2T","W","천상의 그물", 10, 30, 2.00, "방향","느려짐"),
        ("2T","E","재정비 도약",  0, 15, 0.00, "이동","피해 없음, 이속 버프"),
        ("2T","R","금빛 성광",   30, 80, 4.00, "방향","캐스팅 1s / 보호막"),
    ]
    bg2 = [WHITE, GRAY_LIGHT] * 5
    for i, s in enumerate(summary_data):
        for ci, v in enumerate(s, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = normal_font()
            c.fill = fill(bg2[i])
            c.alignment = center() if ci != 8 else left()
            c.border = thin_border()
        ws.row_dimensions[row].height = 17
        row += 1

    return ws


def add_bow_to_skill_list(wb):
    """스킬 리스트 시트에 활 행 추가"""
    if "스킬 리스트" not in wb.sheetnames:
        print("스킬 리스트 시트 없음 — 건너뜀")
        return

    ws = wb["스킬 리스트"]
    last_row = ws.max_row + 1

    # 현재 마지막 행의 스타일 참조용
    def add_row(ws, row, vals):
        bg = WHITE if row % 2 == 1 else "F0F7E8"
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name="맑은 고딕", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        ws.row_dimensions[row].height = 17

    bow_skills = [
        # 장비 구분, 커맨드, GroupId, 스킬명, 상태, 비고
        ("1티어 활", "Q", 300401, "저격",          "Lv.1만", "cd=5s / MP=30 / dmg=165%+25%"),
        ("1티어 활", "W", 300451, "부채 사격",     "Lv.1만", "cd=12s / MP=30 / dmg=225% / 느려짐+밀어냄"),
        ("1티어 활", "E", 300501, "불화살 소나기", "Lv.1만", "cd=15s / MP=30 / 60%×5=300% / Pos"),
        ("1티어 활", "R", 300551, "폭탄 화살",     "Lv.1만", "cd=50s / MP=100 / 500%×2=1000% / 기절"),
        ("2티어 활", "Q", 300601, "섬광 화살",     "Lv.1만", "cd=4s / MP=25 / dmg=150%(적) + 강화평타200%(아군)"),
        ("2티어 활", "W", 300651, "천상의 그물",   "Lv.1만", "cd=10s / MP=30 / dmg=200% / 느려짐"),
        ("2티어 활", "E", 300701, "재정비 도약",   "Lv.1만", "cd=0s(서브) / MP=15 / 이동스킬 / 피해없음"),
        ("2티어 활", "R", 300751, "금빛 성광",     "Lv.1만", "cd=30s / MP=80 / dmg=400% / 캐스팅1s / 보호막"),
        ("3티어 활", "Q", 300801, "(미구현)",      "미구현", "-"),
        ("3티어 활", "W", 300851, "(미구현)",      "미구현", "-"),
        ("3티어 활", "E", 300901, "(미구현)",      "미구현", "-"),
        ("3티어 활", "R", 300951, "(미구현)",      "미구현", "-"),
    ]

    for i, row_data in enumerate(bow_skills):
        add_row(ws, last_row + i, list(row_data))

    print(f"스킬 리스트 시트에 {len(bow_skills)}행 추가 완료 (행 {last_row}~{last_row+len(bow_skills)-1})")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print("시트 목록:", wb.sheetnames)

    # 1) 활 시트 추가
    if "활" in wb.sheetnames:
        del wb["활"]
        print("기존 활 시트 삭제 후 재생성")
    build_bow_sheet(wb)
    print("활 시트 생성 완료")

    # 2) 스킬 리스트에 행 추가
    add_bow_to_skill_list(wb)

    # 3) 저장
    try:
        wb.save(EXCEL_PATH)
        print(f"저장 완료: {EXCEL_PATH}")
    except PermissionError:
        alt = EXCEL_PATH.replace(".xlsx", "_v2.xlsx")
        wb.save(alt)
        print(f"원본 잠김 — 대체 저장: {alt}")


if __name__ == "__main__":
    main()
