"""3번 스킬셋 섹션 + 스킬 리스트 하단 덤프"""
import openpyxl

XLSX_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

# 스킬 리스트 나머지
ws = wb["스킬 리스트"]
print("[스킬 리스트] 행 31~62:")
for r in range(31, ws.max_row + 1):
    row_vals = []
    for c in range(1, 8):
        v = ws.cell(row=r, column=c).value
        row_vals.append(str(v)[:25] if v is not None else "")
    if any(row_vals):
        print(f"  R{r:2d}: {' | '.join(row_vals)}")

# 각 무기 시트 3번 스킬셋 찾기
for sheet_name in ["한손검", "양손검", "활", "지팡이", "단검"]:
    ws = wb[sheet_name]
    print(f"\n[{sheet_name}] 3번 스킬셋 섹션 찾기 (행 40~{ws.max_row}):")
    in_3rd = False
    for r in range(40, ws.max_row + 1):
        row_vals = []
        for c in range(1, 13):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:18] if v is not None else "")
        combined = " | ".join(row_vals)
        if "3번" in combined:
            in_3rd = True
        if in_3rd and any(row_vals):
            print(f"  R{r:3d}: {combined}")
        # 3번 시작 후 80행 이상이면 중단
        if in_3rd and r > (ws.max_row - 2):
            break
