"""
지팡이 3T Q 얼음 화살 쿨타임 수정
  Excel cd [3, 2.5, 2.5, 2, 2] → constants.py [4.0, 3.5, 3.5, 3.0, 3.0]
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['지팡이']

# R125~R129: 지팡이 3T Q Lv1~5, col D=cd
NEW_CD = [4.0, 3.5, 3.5, 3.0, 3.0]
for i, row in enumerate(range(125, 130)):
    old = ws.cell(row, 4).value
    ws.cell(row, 4).value = NEW_CD[i]
    print(f"  R{row} Lv{i+1}: cd {old} → {NEW_CD[i]}")

wb.save(FILEPATH)
print("✅ 저장 완료")

# 검증
wb2 = openpyxl.load_workbook(FILEPATH, read_only=True, data_only=True)
ws2 = wb2['지팡이']
print("\n검증:")
for r in range(125, 130):
    print(f"  R{r}: cd={ws2.cell(r,4).value}, mp={ws2.cell(r,5).value}, dmg={ws2.cell(r,7).value}")
wb2.close()
