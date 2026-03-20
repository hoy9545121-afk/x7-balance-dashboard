"""
X7_무기_스킬.xlsx 3번 스킬셋 Lv.1~5 전체 수정
Source: 게임 데이터 전체_원본.csv/skill_info.csv (CSV 기준)
- 양손검/활/지팡이 3T: CSV에서 Lv.1~5 전체 추출 → 그대로 적용
- 한손검/단검 3T: CSV Lv.1만 → 기존 xlsx Lv.2~5 설계 유지 (단, 음수 진행 등 이상 값 교정)

열 배치 (B=2, C=3, D=4, E=5, G=7):
  B=커맨드, C=레벨, D=쿨타임(s), E=마나소모, G=피해량(배율)
"""
import openpyxl

XLSX_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"

# ─── 1. 수정 데이터 정의 (CSV 기준) ───────────────────────────────────────

# 양손검 3T — CSV: 200000~200155 전체 레벨 확보
# 행: Q=77~81, W=82~86, E=87~91, R=92~96
YSONGEOM_3T = {
    "sheet": "양손검",
    "skills": {
        "Q": {
            "rows": [77, 78, 79, 80, 81],
            "cd":   [5.0, 4.5, 4.0, 3.5, 3.0],
            "mp":   [10,   11,  13,  14,  15],
        },
        "W": {
            "rows": [82, 83, 84, 85, 86],
            "cd":   [12.0, 11.0, 10.0, 9.0, 8.0],
            "mp":   [30,   34,   38,   41,  45],
        },
        "E": {
            "rows": [87, 88, 89, 90, 91],
            "cd":   [15.0, 15.0, 15.0, 15.0, 15.0],
            "mp":   [15,   16,   17,   18,   20],
        },
        "R": {
            "rows": [92, 93, 94, 95, 96],
            "cd":   [40.0, 38.0, 36.0, 34.0, 32.0],
            "mp":   [50,   55,   60,   65,   70],
        },
    }
}

# 활 3T — CSV: 300001~300305 전체 레벨
# 행: Q=83~87, W=88~92, E=93~97, R=98~102
BOW_3T = {
    "sheet": "활",
    "skills": {
        "Q": {
            "rows": [83, 84, 85, 86, 87],
            "cd":   [5.0, 3.0, 3.0, 3.0, 3.0],
            "mp":   [10,  30,  30,  30,  30],
        },
        "W": {
            "rows": [88, 89, 90, 91, 92],
            "cd":   [12.0, 10.0, 10.0, 10.0, 10.0],
            "mp":   [35,   30,   30,   30,   30],
        },
        "E": {
            "rows": [93, 94, 95, 96, 97],
            "cd":   [20.0, 15.0, 15.0, 15.0, 15.0],
            "mp":   [55,   30,   30,   30,   30],
        },
        "R": {
            "rows": [98, 99, 100, 101, 102],
            "cd":   [45.0, 30.0, 30.0, 30.0, 30.0],
            "mp":   [180,  50,   50,   50,   50],
        },
    }
}

# 지팡이 3T — CSV: 400001~400155 전체 레벨
# 행: Q=125~129, W=130~134, E=135~139, R=140~144
# 주의: E(심판의 낙뢰) CD는 RemovalEffect 기반 40~32s 유지 (CSV skill_info cd=20은 내부값)
STAFF_3T = {
    "sheet": "지팡이",
    "skills": {
        "Q": {
            "rows": [125, 126, 127, 128, 129],
            "cd":   [3.0, 2.5, 2.5, 2.0, 2.0],   # CSV
            "mp":   [30,  34,  38,  41,  45],
        },
        "W": {
            "rows": [130, 131, 132, 133, 134],
            "cd":   [15.0, 14.0, 13.0, 12.0, 11.0],
            "mp":   [50,   55,   60,   65,   70],
        },
        "E": {
            "rows": [135, 136, 137, 138, 139],
            # CD: RemovalEffect 기반 유지 (변경 없음)
            "cd":   [None, None, None, None, None],  # None = 변경 없음
            "mp":   [50,   65,   80,   95,  110],
        },
        "R": {
            "rows": [140, 141, 142, 143, 144],
            "cd":   [20.0, 19.0, 18.0, 17.0, 16.0],
            "mp":   [50,   55,   60,   65,   70],
        },
    }
}

# 한손검 3T — CSV Lv.1만 확보, Lv.2~5는 기존 xlsx 설계값 유지
# Lv.1만 CD/MP 교정
# 행: Q=77~81, W=82~86, E=87~91, R=92~96
SWORD_3T = {
    "sheet": "한손검",
    "skills": {
        "Q": {
            "rows": [77],           # Lv.1만
            "cd":   [6.0],
            "mp":   [15],           # CSV: 100601 mp=15
        },
        "W": {
            "rows": [82],
            "cd":   [12.0],
            "mp":   [35],           # CSV: 100651 mp=35
        },
        "E": {
            "rows": [87],
            "cd":   [20.0],
            "mp":   [55],           # CSV: 100701 mp=55
        },
        "R": {
            "rows": [92],
            "cd":   [45.0],
            "mp":   [100],          # 2-node: 별도 노드에서 관리, 기획값 100 유지
        },
    }
}

# 단검 3T — CSV Lv.1만, Lv.2~5 기존 xlsx 설계값 유지 + 이상값 교정
# 행: Q=77~81, W=82~86, E=87~91, R=92~96
DAGGER_3T_LV1 = {
    "sheet": "단검",
    "skills": {
        "Q": {
            "rows": [77],
            "cd":   [6.0],
            "mp":   [50],           # CSV: 500601 mp=50
        },
        "W": {
            "rows": [82],
            "cd":   [None],         # CD는 RemovalEffect 기반, xlsx 유지
            "mp":   [20],           # CSV: 500651 mp=20
        },
        "E": {
            "rows": [87],
            "cd":   [15.0],
            "mp":   [40],           # CSV: 500701 mp=40 ✓
        },
        "R": {
            "rows": [92],
            "cd":   [40.0],         # v16 변경값 유지 (원래 45→40)
            "mp":   [150],          # CSV: 500751 mp=150
        },
    }
}

# 단검 Lv.2~5 개별 교정 (기존 설계 기반, 이상값만 수정)
# Q Lv.2~5: 기존 39/42/46/50 → CSV Lv.1=50이므로 50 이상으로 재설계
# W Lv.2~5: 기존 23/25/28/30 → Lv.1=20이면 모두 ≥20 이므로 OK
# E Lv.5: 기존 30 → Lv.1=40보다 낮음 → 55로 교정
# R Lv.2~5: 기존 165/180/195/210 → Lv.1=150이므로 OK
DAGGER_3T_LV2_5_FIXES = {
    "sheet": "단검",
    "fixes": [
        # Q Lv.2~5: 재설계 (50→55→60→65→70)
        {"row": 78, "col_e": 55},
        {"row": 79, "col_e": 60},
        {"row": 80, "col_e": 65},
        {"row": 81, "col_e": 70},
        # W Lv.2~5: OK (20→23→25→28→30 유지)
        # E Lv.5: 30→55 교정
        {"row": 91, "col_e": 55},
        # R Lv.2~5: OK (150→165→180→195→210 유지)
    ]
}

# ─── 2. 적용 함수 ─────────────────────────────────────────────────────────

def apply_skill_data(ws, skill_name, rows, cds, mps, col_d=4, col_e=5):
    for i, (row, cd, mp) in enumerate(zip(rows, cds, mps)):
        lv = i + 1
        changed = []
        if cd is not None:
            cell = ws.cell(row=row, column=col_d)
            if cell.value != cd:
                changed.append(f"CD {cell.value}→{cd}")
                cell.value = cd
        if mp is not None:
            cell = ws.cell(row=row, column=col_e)
            if cell.value != mp:
                changed.append(f"MP {cell.value}→{mp}")
                cell.value = mp
        status = f"  Lv.{lv} R{row}: {'  '.join(changed) if changed else 'OK'}"
        print(status)

# ─── 3. 실행 ──────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX_PATH)
total_fixed = 0

# 양손검
print("\n=== 양손검 3T ===")
ws = wb[YSONGEOM_3T["sheet"]]
for cmd, data in YSONGEOM_3T["skills"].items():
    print(f"  [{cmd}]")
    apply_skill_data(ws, cmd, data["rows"], data["cd"], data["mp"])

# 활
print("\n=== 활 3T ===")
ws = wb[BOW_3T["sheet"]]
for cmd, data in BOW_3T["skills"].items():
    print(f"  [{cmd}]")
    apply_skill_data(ws, cmd, data["rows"], data["cd"], data["mp"])

# 지팡이
print("\n=== 지팡이 3T ===")
ws = wb[STAFF_3T["sheet"]]
for cmd, data in STAFF_3T["skills"].items():
    print(f"  [{cmd}]")
    apply_skill_data(ws, cmd, data["rows"], data["cd"], data["mp"])

# 한손검
print("\n=== 한손검 3T (Lv.1만) ===")
ws = wb[SWORD_3T["sheet"]]
for cmd, data in SWORD_3T["skills"].items():
    print(f"  [{cmd}]")
    apply_skill_data(ws, cmd, data["rows"], data["cd"], data["mp"])

# 단검 Lv.1
print("\n=== 단검 3T (Lv.1) ===")
ws = wb[DAGGER_3T_LV1["sheet"]]
for cmd, data in DAGGER_3T_LV1["skills"].items():
    print(f"  [{cmd}]")
    apply_skill_data(ws, cmd, data["rows"], data["cd"], data["mp"])

# 단검 Lv.2~5 이상값 교정
print("\n=== 단검 3T (Lv.2~5 이상값 교정) ===")
ws = wb[DAGGER_3T_LV2_5_FIXES["sheet"]]
for fix in DAGGER_3T_LV2_5_FIXES["fixes"]:
    row = fix["row"]
    cell = ws.cell(row=row, column=5)
    old = cell.value
    new = fix["col_e"]
    if old != new:
        print(f"  R{row} MP: {old}→{new}")
        cell.value = new
    else:
        print(f"  R{row}: OK")

# 저장
try:
    wb.save(XLSX_PATH)
    print(f"\n저장 완료: {XLSX_PATH}")
except PermissionError:
    alt = XLSX_PATH.replace(".xlsx", "_v2.xlsx")
    wb.save(alt)
    print(f"원본 잠김 — 대체 저장: {alt}")
