"""
성장_경제_밸런스_통합기획서.xlsx 생성 v2
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
XP 재설계: 등비 수열 (매 레벨 일정 비율 상승)
  req_xp(lv) = A × r^(lv-1)   (A=4300, r=1.065)
몬스터 XP: 난이도 가중치 설계 형식
"""
import os, shutil, math
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH  = r"C:\AI_simulator\기획서\성장_경제_밸런스_통합기획서.xlsx"
SYNC_PATH = r"C:\Users\hoy5343\Documents\카카오톡 받은 파일\성장_경제_밸런스_통합기획서.xlsx"

# ── 색상 ────────────────────────────────────────────────────────────
C = dict(
    BG="FF090008", CARD="FF111018", SECT="FF161620",
    HDR="FF23222E", SUB="FF1C1B28",
    GOLD="FFE8C050", GOLD2="FFD4A830", GOLD_BG="FF2A2010",
    CREAM="FFF0E8D8", T2="FFB8A898", T3="FF7A6A5A",
    GREEN="FF50C870", BLUE="FF4090F0",
    PURPLE="FFA060E8", RED="FFD43030", ORANGE="FFE68C32",
    ROW_A="FF141220", ROW_B="FF111018",
    WARN_BG="FF2A1A10", WARN_FG="FFFF8040",
    TIER1="FF4BC860", TIER2="FF82CD5A", TIER3="FFC8C846",
    TIER4="FFE8C050", TIER5="FFE68C32", TIER6="FFD44040", TIER7="FFA060E8",
)
TIER_C = [None, C["TIER1"],C["TIER2"],C["TIER3"],C["TIER4"],C["TIER5"],C["TIER6"],C["TIER7"]]

# ── 스타일 헬퍼 ─────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def cx(ws, row, col, value="", bold=False, size=10, color="FFF0E8D8",
       bg=None, align="left", valign="center", wrap=False,
       border_top=False, border_bot=False, border_left=False, border_right=False,
       num_fmt=None, thick_top=False, thick_bot=False, font_name="Noto Sans KR"):
    if isinstance(value, str) and value.startswith("="):
        value = "\u200b" + value
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = Font(bold=bold, size=size, color=color, name=font_name)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg: c.fill = fill(bg)
    if num_fmt: c.number_format = num_fmt
    bdr = {}
    if thick_top:   bdr["top"]    = Side(style="medium", color="FFE8C050")
    elif border_top:bdr["top"]    = Side(style="thin",   color="FF333344")
    if thick_bot:   bdr["bottom"] = Side(style="medium", color="FFE8C050")
    elif border_bot:bdr["bottom"] = Side(style="thin",   color="FF333344")
    if border_left: bdr["left"]   = Side(style="thin",   color="FF333344")
    if border_right:bdr["right"]  = Side(style="thin",   color="FF333344")
    if bdr: c.border = Border(**bdr)
    return c

def mc(ws, r1, c1, r2, c2, **kw):
    cell = cx(ws, r1, c1, **kw)
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return cell

def blank(ws, row, c1=1, c2=13, bg=None):
    for c in range(c1, c2+1): ws.cell(row,c).fill = fill(bg or C["BG"])
    ws.row_dimensions[row].height = 6

def sec_bar(ws, row, title, c1=2, c2=11):
    mc(ws, row, c1, row, c2, value=title, bold=True, size=10,
       color=C["GOLD"], bg=C["GOLD_BG"], align="left", thick_top=True)
    ws.row_dimensions[row].height = 22

def hdr_row(ws, row, cols, labels, bg=None):
    bg = bg or C["HDR"]
    for (cs,ce), lbl in zip(cols, labels):
        mc(ws, row, cs, row, ce, value=lbl, bold=True, size=9,
           color=C["T2"], bg=bg, align="center",
           thick_top=True, thick_bot=True)
    ws.row_dimensions[row].height = 22

# ── XP 공식 ─────────────────────────────────────────────────────────
XPM_BASE = 500   # 기준 분당 XP (T1 Normal)
A_XP     = 4300  # lv1 기본 요구 XP
R_XP     = 1.065 # 등비 공비 (매 레벨 6.5% 상승)

def req_xp_formula(lv):
    """등비 수열 공식: req_xp(lv) = A × r^(lv-1)  (10 단위 반올림 — 비율 오차 최소화)"""
    raw = A_XP * (R_XP ** (lv - 1))
    return max(100, round(raw / 10) * 10)

def cum_xp_formula(lv):
    """lv1 ~ lv 누계 XP"""
    return A_XP * (R_XP**lv - 1) / (R_XP - 1)

def cum_h_formula(lv, xpm=XPM_BASE):
    return cum_xp_formula(lv) / (xpm * 60)

# ── 레벨별 데이터 생성 ──────────────────────────────────────────────
# 지역 매핑 (장비 티어 기준 — lv61+ 미정)
AREA_INFO = [
    (1,10,  "이니스 섬",    1),
    (11,20, "솔즈리드 반도",2),
    (21,30, "릴리엇 구릉지",3),
    (31,40, "가랑돌 평원",  4),
    (41,50, "마리아 노플",  5),
    (51,60, "긴 모래톱",    6),
    (61,100,"(미정)",       7),
]
def area_tier(lv):
    for a,b,name,t in AREA_INFO:
        if a<=lv<=b: return name,t
    return "(미정)",7

level_rows = []
prev_cum = 0
for lv in range(1, 101):
    req = req_xp_formula(lv)
    cum = prev_cum + req
    name, tier = area_tier(lv)
    ratio = "—" if lv == 1 else "×1.065"
    level_rows.append(dict(
        lv=lv, area=name, tier=tier, req=req,
        cum=cum, ratio=ratio,
        cum_h=round(cum/(XPM_BASE*60), 2),
        time_min=round(req/XPM_BASE, 1),
    ))
    prev_cum = cum

# 누계 시간 앵커 (수식 기반 재계산)
ANCHORS_NEW = [(lv, round(cum_h_formula(lv))) for lv in [10,20,30,40,50,60,70,80,90,100]]
ANCHORS_OLD = [(10,2),(20,6),(30,12),(35,20),(40,30),(45,50),(50,90),(55,170),(60,330),
               (70,600),(80,800),(90,950),(100,1200)]

# 티어별 Normal 몬스터 XP (CSV 데이터)
TIER_MON_XP = {1:10,2:15,3:22,4:33,5:49,6:73,7:109}

# 숙련도 (원본 그대로)
PROF_ROWS = [
    (1,"Lv1","캐릭터 10레벨 완료","2h",  36000,  36000,"2T 진입 가능"),
    (2,"Lv1","캐릭터 20레벨 완료","4h",  72000,  72000,"3T 진입 가능"),
    (3,"Lv1","캐릭터 30레벨 완료","6h",  27700, 108000,""),
    (3,"Lv2","","",                      36000,   None, ""),
    (3,"Lv3","","",                      44300,   None, "4T 진입 가능"),
    (4,"Lv1","캐릭터 40레벨 완료","18h", 83100, 324000,""),
    (4,"Lv2","","",                     108000,   None, ""),
    (4,"Lv3","","",                     132900,   None, "5T 진입 가능"),
    (5,"Lv1","캐릭터 50레벨 완료","60h",166200,1080000,""),
    (5,"Lv2","","",                     191100,   None, ""),
    (5,"Lv3","","",                     216000,   None, ""),
    (5,"Lv4","","",                     240900,   None, ""),
    (5,"Lv5","","",                     265800,   None, "6T 진입 가능"),
    (6,"Lv1","캐릭터 55레벨 완료","80h",221500,1440000,""),
    (6,"Lv2","","",                     254800,   None, ""),
    (6,"Lv3","","",                     288000,   None, ""),
    (6,"Lv4","","",                     321200,   None, ""),
    (6,"Lv5","","",                     354500,   None, "7T 진입 가능"),
    (7,"Lv1","캐릭터 60레벨 완료","160h",246200,2880000,""),
    (7,"Lv2","","",264600,None,""),
    (7,"Lv3","","",283100,None,""),
    (7,"Lv4","","",301500,None,""),
    (7,"Lv5","","",320000,None,""),
    (7,"Lv6","","",338500,None,""),
    (7,"Lv7","","",356900,None,""),
    (7,"Lv8","","",375400,None,""),
    (7,"Lv9","","",393800,None,"7T 만렙"),
]

ORIG_XP = {
    1:4300,10:7700,20:15400,21:12900,30:23100,31:34300,35:61700,
    36:42900,60:1234300,61:578600,70:1041400,71:428600,
    80:771400,81:321400,90:578600,91:535700,100:964300,
}

# ══════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ══ Sheet 1: 기획 의도 ══════════════════════════════════════════════
ws = wb.create_sheet("📋 기획 의도")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = C["GOLD2"]
ws.column_dimensions["A"].width = 1.5
ws.column_dimensions["B"].width = 18
for ch in "CDEFGHIJK":
    ws.column_dimensions[ch].width = 13
ws.column_dimensions["K"].width = 18

for r in range(1, 90): [ws.cell(r,c).__setattr__("fill", fill(C["BG"])) for c in range(1,14)]

R = 1
blank(ws, R); R+=1
mc(ws,R,2,R,11, value="성장 경제 밸런스 — 통합 기획서",
   bold=True, size=16, color=C["GOLD"], bg=C["CARD"], align="center"); ws.row_dimensions[R].height=36; R+=1
mc(ws,R,2,R,11, value=f"캐릭터 레벨 & 숙련도 경험치 설계  v2.0  |  등비 수열 XP 공식 적용 (A={A_XP:,}, r={R_XP})",
   size=9, color=C["T2"], bg=C["CARD"], align="center"); ws.row_dimensions[R].height=18; R+=1
blank(ws,R); R+=1

# ── 1. 개요 및 설계 원칙
sec_bar(ws, R, "1. 개요 및 설계 원칙"); R+=1
intro = [
    ("캐릭터 경험치",
     "매 레벨 일정 비율(×1.065)로 상승하는 등비 수열 적용. 어떤 구간에서도 동일한 성장감 제공."),
    ("숙련도 경험치",
     "티어 내 점진 상승 (첫 레벨 대비 마지막 레벨 ≈ 1.6배). 분당 300 XP 고정 (티어·난이도 무관)."),
    ("캐릭터 & 숙련도 동기화",
     "지역 완료 레벨 달성 시 해당 지역 숙련도 티어 만렙 달성 — 콘텐츠 진입 조건 동기화."),
    ("몬스터 경험치 설계",
     "티어별 베이스 XP × 난이도 가중치. Normal 기준 파밍 효율이 가장 높고, 높은 난이도는 XP 보정 제공."),
]
for k,v in intro:
    ws.row_dimensions[R].height = 30
    cx(ws,R,2, k, bold=True, size=9, color=C["GOLD"], bg=C["SECT"])
    mc(ws,R,3,R,11, value=v, size=9, color=C["CREAM"], bg=C["SECT"], wrap=True, valign="center"); R+=1
blank(ws,R); R+=1

# ── 2. XP 공식
sec_bar(ws, R, "2. 캐릭터 XP 공식 (등비 수열)"); R+=1
formula_lines = [
    ("공식",          f"req_xp(lv) = {A_XP:,} × {R_XP}^(lv-1)   →   매 레벨 정확히 ×{R_XP} 상승"),
    ("공비 선택 이유", f"전 100레벨에 걸쳐 균일한 성장감 제공. r={R_XP} 기준 1→100레벨 약 ×{R_XP**99:.0f} 배율."),
    ("누계 공식",      f"cumXP(lv) = {A_XP:,} × ({R_XP}^lv - 1) / ({R_XP} - 1)"),
    ("기준 분당 XP",   f"{XPM_BASE:,} XP/분 (T1 Normal 몬스터 기준). 고티어 지역 파밍 시 실질 레벨업 단축."),
    ("lv1 필요 XP",   f"{req_xp_formula(1):,} XP  (≈ {req_xp_formula(1)/XPM_BASE:.0f}분)"),
    ("lv100 필요 XP", f"{req_xp_formula(100):,} XP  (≈ {req_xp_formula(100)/XPM_BASE/60:.1f}h / T1 기준)"),
    ("100레벨 총 XP", f"{sum(r['req'] for r in level_rows):,} XP  ≈ {sum(r['req'] for r in level_rows)/XPM_BASE/60:.0f}h (T1 기준)"),
]
for k,v in formula_lines:
    ws.row_dimensions[R].height = 26
    cx(ws,R,2, k, bold=True, size=9, color=C["GOLD"], bg=C["SECT"])
    mc(ws,R,3,R,11, value=v, size=9, color=C["CREAM"], bg=C["SECT"], wrap=True, valign="center"); R+=1
blank(ws,R); R+=1

# ── 3. 앵커 비교
sec_bar(ws, R, "3. 레벨 앵커 타임 — v1(구형) vs v2(등비 공식)"); R+=1
ws.row_dimensions[R].height = 22
hdr_row(ws, R, [(2,2),(3,4),(5,6),(7,8),(9,11)],
        ["완료 레벨","레벨 달성 시간\n[등비공식 v2]","해당 레벨\n필요 XP","레벨 달성 시간\n[구형 v1]","비고"])
R+=1
old_map = dict(ANCHORS_OLD)
for lv, h_new in ANCHORS_NEW:
    h_old = old_map.get(lv, "—")
    rrow = level_rows[lv-1]
    bg = C["ROW_A"] if (ANCHORS_NEW.index((lv,h_new)))%2==0 else C["ROW_B"]
    ws.row_dimensions[R].height = 20
    cx(ws,R,2, f"{lv}레벨", bold=True, size=9, color=C["GOLD"], bg=bg, align="center")
    mc(ws,R,3,R,4, value=f"≈{h_new}h", bold=True, size=9, color=C["GREEN"], bg=bg, align="center")
    mc(ws,R,5,R,6, value=rrow["req"], size=9, color=C["CREAM"], bg=bg, align="right", num_fmt="#,##0")
    mc(ws,R,7,R,8, value=f"{h_old}h" if h_old!="—" else "—",
       size=9, color=C["T3"], bg=bg, align="center")
    mc(ws,R,9,R,11, value="", bg=bg); R+=1
blank(ws,R); R+=1
mc(ws,R,2,R,11,
   value="※ v2는 구간 전환 없이 전 레벨 동일 공비 적용. 초반(1~30레벨)은 구형과 거의 동일, 중반(40~60)이 빨라지고 후반(70~100)이 길어짐.",
   size=8, color=C["WARN_FG"], bg=C["WARN_BG"], wrap=True); ws.row_dimensions[R].height=24; R+=1
blank(ws,R); R+=1

# ── 4. 몬스터 경험치 설계
sec_bar(ws, R, "4. 몬스터 경험치 설계"); R+=1
ws.row_dimensions[R].height = 30
mc(ws,R,2,R,11,
   value="몬스터 경험치 = 티어별 베이스 XP  ×  난이도 가중치\n"
         "캐릭터 XP는 몬스터 티어에 따라 증가하므로, 현재 콘텐츠보다 높은 티어 몬스터를 처치하면 빠른 레벨업 가능.",
   size=9, color=C["CREAM"], bg=C["SECT"], wrap=True, valign="center"); R+=1
blank(ws,R); R+=1

# 4.1 티어별 베이스 XP
sec_bar(ws, R, "4.1 티어별 베이스 XP (Normal 기준)", c1=2, c2=11); R+=1
ws.row_dimensions[R].height = 22
hdr_row(ws, R, [(2,2),(3,3),(4,5),(6,7),(8,9),(10,11)],
        ["장비 티어","베이스 XP\n/킬","분당 XP\n(분당 10킬 기준)","T1 대비\n배율","추천 캐릭터\n레벨 구간","비고"])
R+=1
for ti in range(1,8):
    bg = C["ROW_A"] if ti%2==0 else C["ROW_B"]
    tc = TIER_C[ti]
    base = TIER_MON_XP[ti]
    xpm  = base * 10   # 분당 10킬
    mult = round(base/TIER_MON_XP[1], 1)
    lv_range = {1:"1~10",2:"11~20",3:"21~30",4:"31~40",5:"41~50",6:"51~60",7:"61+"}[ti]
    note = "기준" if ti==1 else f"T1 대비 {mult}배 빠른 레벨업"
    ws.row_dimensions[R].height = 20
    cx(ws,R,2, f"T{ti}", bold=True, size=10, color=tc, bg=bg, align="center")
    cx(ws,R,3, base, bold=True, size=9, color=C["CREAM"], bg=bg, align="center")
    mc(ws,R,4,R,5, value=xpm, size=9, color=C["GOLD"], bg=bg, align="center", num_fmt="#,##0")
    mc(ws,R,6,R,7, value=f"×{mult}", size=9,
       color=C["GREEN"] if ti>1 else C["T2"], bg=bg, align="center")
    mc(ws,R,8,R,9, value=lv_range, size=9, color=C["T2"], bg=bg, align="center")
    mc(ws,R,10,R,11, value=note, size=9, color=C["T2"], bg=bg); R+=1
blank(ws,R); R+=1

# 4.2 난이도 가중치
sec_bar(ws, R, "4.2 난이도별 경험치 가중치 (참고)", c1=2, c2=11); R+=1
ws.row_dimensions[R].height = 42
mc(ws,R,2,R,11,
   value="모든 티어에 동일하게 적용. 체력과 경험치가 같은 배율로 오르므로 시간당 XP 효율은 모든 난이도가 동일.\n"
         "→ Normal 순살이 실질적으로 가장 효율적 (같은 XP를 더 안전하고 빠르게 획득). 고난이도는 XP 목적이 아닌 드랍·도전 목적.",
   size=9, color=C["CREAM"], bg=C["SECT"], wrap=True, valign="center"); R+=1
ws.row_dimensions[R].height = 22
diff_data = [
    ("Normal",    "×1.0", "기본 파밍 추천 — 빠른 처치로 XP 극대화"),
    ("Strong",    "×1.25","체력·XP 동일 배율 → 시간당 XP 동일"),
    ("Commander", "×1.5", "체력·XP 동일 배율 → 시간당 XP 동일"),
    ("Elite",     "×2.5", "체력·XP 동일 배율 → 시간당 XP 동일"),
    ("Hero",      "×4",   "체력·XP 동일 배율 → 시간당 XP 동일"),
    ("Boss",      "×8",   "체력·XP 동일 배율 → 시간당 XP 동일  ※ 완제품 드랍 및 특수 보상이 목적"),
]
hdr_row(ws, R, [(2,3),(4,5),(6,11)],
        ["난이도","경험치 가중치","설계 의도"])
R+=1
for i,(diff, wt, note) in enumerate(diff_data):
    bg = C["ROW_A"] if i%2==0 else C["ROW_B"]
    ws.row_dimensions[R].height = 20
    mc(ws,R,2,R,3, value=diff, bold=(diff=="Normal"), size=9,
       color=C["GOLD"] if diff=="Normal" else C["CREAM"], bg=bg)
    mc(ws,R,4,R,5, value=wt, bold=True, size=9, color=C["GOLD"], bg=bg, align="center")
    mc(ws,R,6,R,11, value=note, size=9, color=C["T2"], bg=bg); R+=1
blank(ws,R); R+=1
mc(ws,R,2,R,11,
   value="→ 예시: T3 Normal 10킬/분 = 220 XP/분 / T3 Boss 1마리 처치(5분) × 55 XP = 11 XP/분  ※ Boss는 XP가 아닌 드랍이 목적",
   size=9, color=C["WARN_FG"], bg=C["WARN_BG"]); ws.row_dimensions[R].height=22; R+=1

# ── 5. 숙련도 설계
blank(ws,R); R+=1
sec_bar(ws, R, "5. 숙련도 설계 원칙"); R+=1
prof_pts = [
    ("기준 XP/분",  "300 XP/분 (캐릭터 경험치의 60%). 티어·난이도 무관 고정값."),
    ("시간 게이팅", "높은 스펙으로도 숙련도 가속 불가 → 정량적 진입 조건. BM으로 소량 가속 허용 예정."),
    ("티어 동기화", "지역 완료 레벨 달성 시 해당 지역 숙련도 티어 만렙 → 다음 지역 콘텐츠 바로 입장 가능."),
    ("티어 내 레벨", "레벨마다 보너스 능력치 부여. 티어 내 XP는 점진 상승 (첫 레벨 대비 마지막 ≈ 1.6배)."),
]
for k,v in prof_pts:
    ws.row_dimensions[R].height=26
    cx(ws,R,2, k, bold=True, size=9, color=C["GOLD"], bg=C["SECT"])
    mc(ws,R,3,R,11, value=v, size=9, color=C["CREAM"], bg=C["SECT"], wrap=True, valign="center"); R+=1

# ══ Sheet 2: 캐릭터 레벨 테이블 ════════════════════════════════════
ws2 = wb.create_sheet("📊 캐릭터 레벨")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = C["GREEN"]
ws2.column_dimensions["A"].width = 1.5
ws2.column_dimensions["B"].width = 6
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 5
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 8
ws2.column_dimensions["I"].width = 10
ws2.column_dimensions["J"].width = 12
ws2.column_dimensions["K"].width = 10

for r in range(1,120): [ws2.cell(r,c).__setattr__("fill",fill(C["BG"])) for c in range(1,14)]

R2=1; blank(ws2,R2,1,12); R2+=1
mc(ws2,R2,2,R2,11, value="캐릭터 레벨 경험치 테이블 v2.0",
   bold=True, size=14, color=C["GOLD"], bg=C["CARD"], align="center"); ws2.row_dimensions[R2].height=32; R2+=1
mc(ws2,R2,2,R2,11,
   value=f"공식: req_xp(lv) = {A_XP:,} × {R_XP}^(lv-1)  |  매 레벨 정확히 ×{R_XP} 상승  |  기준 {XPM_BASE} XP/분 (T1 Normal)",
   size=9, color=C["WARN_FG"], bg=C["WARN_BG"], align="center"); ws2.row_dimensions[R2].height=18; R2+=1
blank(ws2,R2,1,12); R2+=1

# 헤더
ws2.row_dimensions[R2].height = 30
hdr_pairs = [
    (2,2,"레벨"),  (3,3,"지역"),  (4,4,"T"),
    (5,5,"소요\n(분)"),(6,7,"필요 경험치\n(등비 공식)"),(8,9,"누계 경험치"),
    (10,10,"상승폭"),(11,11,"레벨 달성\n시간(h)"),
]
for cs,ce,lbl in hdr_pairs:
    mc(ws2,R2,cs,R2,ce, value=lbl, bold=True, size=9, color=C["T2"], bg=C["HDR"],
       align="center", valign="center", wrap=True, thick_top=True, thick_bot=True)
R2+=1

ANCHOR_LVS = {lv for lv,_ in ANCHORS_NEW}
for row in level_rows:
    lv = row["lv"]
    is_anc = lv in ANCHOR_LVS
    bg = C["GOLD_BG"] if is_anc else (C["ROW_A"] if lv%2==0 else C["ROW_B"])
    tc = TIER_C[row["tier"]]
    ws2.row_dimensions[R2].height = 18

    cx(ws2,R2,2, lv, bold=is_anc, size=9, color=C["GOLD"] if is_anc else C["CREAM"],
       bg=bg, align="center", thick_top=is_anc, thick_bot=is_anc)
    cx(ws2,R2,3, row["area"], size=8, color=C["T2"], bg=bg, align="left",
       thick_top=is_anc, thick_bot=is_anc)
    cx(ws2,R2,4, f"T{row['tier']}", bold=True, size=9, color=tc, bg=bg, align="center",
       thick_top=is_anc, thick_bot=is_anc)
    cx(ws2,R2,5, row["time_min"], size=9, color=C["CREAM"], bg=bg, align="right",
       num_fmt="#,##0.0", thick_top=is_anc, thick_bot=is_anc)
    mc(ws2,R2,6,R2,7, value=row["req"], bold=is_anc, size=9,
       color=C["GOLD"] if is_anc else C["CREAM"], bg=bg, align="right",
       num_fmt="#,##0", thick_top=is_anc, thick_bot=is_anc)
    mc(ws2,R2,8,R2,9, value=row["cum"], size=9, color=C["T2"], bg=bg, align="right",
       num_fmt="#,##0", thick_top=is_anc, thick_bot=is_anc)
    # 상승폭 (항상 ~×1.065)
    cx(ws2,R2,10, row["ratio"], size=9, color=C["GREEN"], bg=bg, align="center",
       thick_top=is_anc, thick_bot=is_anc)
    cx(ws2,R2,11, row["cum_h"], size=9, color=C["T3"], bg=bg, align="right",
       num_fmt="0.0", thick_top=is_anc, thick_bot=is_anc)
    R2+=1

blank(ws2,R2,1,12); R2+=1
mc(ws2,R2,2,R2,11,
   value=f"※ 상승폭은 전 레벨 대비 ×{R_XP} (6.5%) 고정. 레벨 달성 시간은 T1 Normal 기준 ({XPM_BASE} XP/분). 고티어 지역 파밍 시 실제 소요 시간 대폭 단축.",
   size=8, color=C["T2"], bg=C["CARD"]); ws2.row_dimensions[R2].height=18

# ══ Sheet 3: 숙련도 ════════════════════════════════════════════════
ws3 = wb.create_sheet("🏆 숙련도")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = C["PURPLE"]
ws3.column_dimensions["A"].width = 1.5
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 8
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 8
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 20

for r in range(1,55): [ws3.cell(r,c).__setattr__("fill",fill(C["BG"])) for c in range(1,12)]

R3=1; blank(ws3,R3,1,11); R3+=1
mc(ws3,R3,2,R3,9, value="숙련도 티어 & 레벨별 경험치 테이블",
   bold=True, size=14, color=C["GOLD"], bg=C["CARD"], align="center"); ws3.row_dimensions[R3].height=32; R3+=1
mc(ws3,R3,2,R3,9, value="기준 300 XP/분 (티어·난이도 무관 고정)  |  지역 완료 시점 = 해당 티어 만렙",
   size=9, color=C["T2"], bg=C["CARD"], align="center"); ws3.row_dimensions[R3].height=18; R3+=1
blank(ws3,R3); R3+=1

ws3.row_dimensions[R3].height = 22
hdr_row(ws3, R3,
        [(2,2),(3,3),(4,5),(6,6),(7,7),(8,8),(9,9)],
        ["티어","레벨","달성 목표","기간(h)","레벨 필요 XP","티어 총 XP","비고"])
R3+=1
cur_t=0
for row in PROF_ROWS:
    t,lv,goal,period,lv_xp,tier_xp,note = row
    new_t = (t!=cur_t); cur_t=t
    tc = TIER_C[t]; bg = C["ROW_A"] if t%2==0 else C["ROW_B"]
    ws3.row_dimensions[R3].height=20
    cx(ws3,R3,2, f"T{t}" if new_t else "", bold=new_t, size=10, color=tc, bg=bg,
       align="center", thick_top=new_t)
    cx(ws3,R3,3, lv, size=9, color=C["CREAM"], bg=bg, align="center", thick_top=new_t)
    mc(ws3,R3,4,R3,5, value=goal, size=9, color=C["T2"], bg=bg, thick_top=new_t)
    cx(ws3,R3,6, period, size=9, color=C["CREAM"], bg=bg, align="center", thick_top=new_t)
    cx(ws3,R3,7, lv_xp, bold=True, size=9, color=C["GOLD"], bg=bg, align="right",
       num_fmt="#,##0", thick_top=new_t)
    cx(ws3,R3,8, tier_xp if new_t else "", size=9, color=C["T2"], bg=bg, align="right",
       num_fmt="#,##0", thick_top=new_t)
    cx(ws3,R3,9, note, size=9, color=C["GREEN"] if "가능" in note or "만렙" in note else C["T2"],
       bg=bg, thick_top=new_t); R3+=1

# ══ Sheet 4: 수치 요약 ══════════════════════════════════════════════
ws4 = wb.create_sheet("📈 수치 요약")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = C["ORANGE"]
ws4.column_dimensions["A"].width = 1.5
for ch in "BCDEFGHIJK":
    ws4.column_dimensions[ch].width = 13

for r in range(1,55): [ws4.cell(r,c).__setattr__("fill",fill(C["BG"])) for c in range(1,13)]

R4=1; blank(ws4,R4,1,12); R4+=1
mc(ws4,R4,2,R4,11, value="핵심 수치 요약",
   bold=True, size=14, color=C["GOLD"], bg=C["CARD"], align="center"); ws4.row_dimensions[R4].height=32; R4+=1
blank(ws4,R4); R4+=1

sec_bar(ws4, R4, "캐릭터 XP 공식 핵심 수치", c2=11); R4+=1
summary = [
    ("공식",            f"req_xp(lv) = A × r^(lv-1)"),
    ("A (lv1 XP)",     f"{A_XP:,} XP"),
    ("r (공비)",        f"{R_XP}  (매 레벨 6.5% 상승)"),
    ("lv1 필요 XP",    f"{req_xp_formula(1):,} XP  =  {req_xp_formula(1)/XPM_BASE:.0f}분"),
    ("lv50 필요 XP",   f"{req_xp_formula(50):,} XP  =  {req_xp_formula(50)/XPM_BASE/60:.1f}h"),
    ("lv100 필요 XP",  f"{req_xp_formula(100):,} XP  =  {req_xp_formula(100)/XPM_BASE/60:.1f}h"),
    ("100레벨 총 XP",  f"{sum(r['req'] for r in level_rows):,} XP  ≈  {sum(r['req'] for r in level_rows)/XPM_BASE/60:.0f}h  (T1 기준)"),
    ("단조 증가 여부",  "항상 증가 (전 레벨 ×1.065) — 감소 구간 없음"),
]
for i,(k,v) in enumerate(summary):
    bg = C["ROW_A"] if i%2==0 else C["ROW_B"]
    ws4.row_dimensions[R4].height=22
    mc(ws4,R4,2,R4,4, value=k, bold=True, size=9, color=C["GOLD"], bg=bg)
    mc(ws4,R4,5,R4,11, value=v, size=9, color=C["CREAM"], bg=bg); R4+=1
blank(ws4,R4); R4+=1

sec_bar(ws4, R4, "앵커 타임 요약 (v2 기준)", c2=11); R4+=1
ws4.row_dimensions[R4].height=22
hdr_row(ws4, R4, [(2,3),(4,5),(6,8),(9,11)],
        ["레벨","레벨 달성 시간\n(v2 등비공식)","필요 XP\n(해당 레벨)","레벨 달성 시간\n(v1 구형)"])
R4+=1
for i,(lv,h_new) in enumerate(ANCHORS_NEW):
    h_old = old_map.get(lv,"—")
    bg = C["ROW_A"] if i%2==0 else C["ROW_B"]
    rrow = level_rows[lv-1]
    ws4.row_dimensions[R4].height=20
    mc(ws4,R4,2,R4,3, value=f"{lv}레벨", bold=True, size=9, color=C["GOLD"], bg=bg, align="center")
    mc(ws4,R4,4,R4,5, value=f"≈{h_new}h", bold=True, size=9, color=C["GREEN"], bg=bg, align="center")
    mc(ws4,R4,6,R4,8, value=rrow["req"], size=9, color=C["CREAM"], bg=bg, align="right", num_fmt="#,##0")
    mc(ws4,R4,9,R4,11, value=f"{h_old}h" if h_old!="—" else "—",
       size=9, color=C["T3"], bg=bg, align="center"); R4+=1

# ── 저장 ────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
try:
    wb.save(OUT_PATH)
    shutil.copy2(OUT_PATH, SYNC_PATH)
    print(f"✅ 저장: {OUT_PATH}")
    print(f"✅ 동기화: {SYNC_PATH}")
except PermissionError:
    alt = OUT_PATH.replace(".xlsx","_v2.xlsx")
    wb.save(alt)
    print(f"⚠️  원본 잠김 → 대체 저장: {alt}")

print(f"\n── 등비 수열 XP 미리보기 (공비 ×{R_XP}) ──")
print(f"{'레벨':>5}  {'req_xp':>12}  {'ratio':>8}  {'cum_h':>8}")
for row in level_rows:
    lv = row["lv"]
    if lv in {1,10,20,30,40,50,60,70,80,90,100}:
        print(f"  Lv{lv:3d}  {row['req']:>12,}  {row['ratio']:>8}  {row['cum_h']:>8.1f}h")
