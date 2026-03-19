"""
스킬 MP Lv1 (+ 단검 E Lv5) 일괄 수정
constants.py 기준으로 Excel 미반영 값 수정
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)

# {sheet: [(row, col, new_value, label), ...]}
FIXES = {
    '양손검': [
        (77, 5, 6,  'Q Lv1 mp'),
        (82, 5, 18, 'W Lv1 mp'),
        (87, 5, 9,  'E Lv1 mp'),
        (92, 5, 30, 'R Lv1 mp'),
    ],
    '한손검': [
        (77, 5, 9,   'Q Lv1 mp'),
        (82, 5, 20,  'W Lv1 mp'),
        (87, 5, 30,  'E Lv1 mp'),
        (92, 5, 100, 'R Lv1 mp'),
    ],
    '활': [
        (83, 5, 6,   'Q Lv1 mp'),
        (88, 5, 20,  'W Lv1 mp'),
        (93, 5, 30,  'E Lv1 mp'),
        (98, 5, 100, 'R Lv1 mp'),
    ],
    '지팡이': [
        (125, 5, 18, 'Q Lv1 mp'),
        (130, 5, 30, 'W Lv1 mp'),
        (135, 5, 30, 'E Lv1 mp'),
        (140, 5, 30, 'R Lv1 mp'),
    ],
    '단검': [
        (77, 5, 20,  'Q Lv1 mp'),
        (82, 5, 12,  'W Lv1 mp'),
        (87, 5, 24,  'E Lv1 mp'),
        (91, 5, 30,  'E Lv5 mp'),  # 단검 E Lv5 특이값
        (92, 5, 90,  'R Lv1 mp'),
    ],
}

total = 0
for sheet, changes in FIXES.items():
    ws = wb[sheet]
    for row, col, new_val, label in changes:
        old = ws.cell(row, col).value
        ws.cell(row, col).value = new_val
        print(f"  [{sheet}] R{row} {label}: {old} → {new_val}")
        total += 1

wb.save(FILEPATH)
print(f"\n✅ {total}개 수정 완료 — 저장됨")

# 검증
wb2 = openpyxl.load_workbook(FILEPATH, read_only=True, data_only=True)
print("\n=== 검증 ===")
for sheet, changes in FIXES.items():
    ws2 = wb2[sheet]
    for row, col, new_val, label in changes:
        actual = ws2.cell(row, col).value
        mark = '✅' if actual == new_val else '❌'
        print(f"  {mark} [{sheet}] R{row} {label}: {actual}")
wb2.close()
