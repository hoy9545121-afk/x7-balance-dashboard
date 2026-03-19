"""
경제 기획서 상세 수치 추출
"""
import openpyxl

BASE = r'C:\AI_simulator\기획서'

def dump(path, sheet, r1, r2):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"⚠ 없음: {sheet}")
        wb.close()
        return
    ws = wb[sheet]
    print(f"\n=== [{sheet}] ===")
    for r in range(r1, r2+1):
        row = [ws.cell(r, c).value for c in range(1, 16)]
        if any(v is not None for v in row):
            clean = [v for v in row if v is not None]
            print(f"  R{r}: {clean}")
    wb.close()

G = fr'{BASE}\X7_성장_경제_밸런스_통합기획서.xlsx'
D = fr'{BASE}\X7_장비_제작_및_드랍.xlsx'

# 기획 의도 전체 (몬스터 XP 테이블 포함)
dump(G, '📋 기획 의도', 1, 80)

# 수치 요약 전체
dump(G, '📈 수치 요약', 1, 60)

# 장비 아이템 리스트 전체
dump(D, '📋 장비 아이템 리스트', 1, 100)

# 수량 계산 근거
dump(D, '📐 수량 계산 근거', 1, 60)

# 드랍률 설계 전체
dump(D, '🎲 드랍률 설계', 1, 80)

# Craft 수량 반영안
dump(D, '✅ Craft 수량 반영안', 1, 80)

# 전리품 제작
dump(D, '🗡️ 전리품 제작', 1, 60)

# 던전 코어 제작
dump(D, '🏰 던전 코어 제작', 1, 60)

# 채집 시간 기준
dump(D, '📊 채집 시간 기준', 1, 40)

# 하우징 채집 효과
dump(D, '🏡 하우징 채집 효과', 1, 40)
