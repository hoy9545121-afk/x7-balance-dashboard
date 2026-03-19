"""A안 vs B안 비교 시트 추가"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BG_DARK   = '001A1F3A'; BG_MED    = '001E2040'; BG_DARK2  = '00181830'
BG_HEADER = '000F2D5A'; BG_SECTION= '001A3A6A'
FC_TITLE  = '00F5D878'; FC_SECTION= '00E8C050'; FC_VALUE  = '00EEE8D8'
FC_DIM    = '007A6A5A'; FC_NOTE   = '009A7A28'
GREEN = '0044DD88'; RED   = '00FF6666'; YELLOW= '00FFDD55'
ORANGE= '00FF9933';  BLUE  = '004499FF'; PURPLE= '00BB88FF'

TIER_FC = ['008A9BB0','004BE8A0','004B8CE8','00C89040','00A04BE8','00E84B4B','00E8B84B']

def c(ws, row, col, val, fc=FC_VALUE, bg=BG_DARK, bold=False, sz=10,
      ha='center', va='center', wrap=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name='맑은 고딕', color=fc, bold=bold, size=sz)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    return cell


wb = openpyxl.load_workbook('기획서/X7_숙련도_옵션_기획서.xlsx')

# 기존 시트 있으면 제거
if '⚖ A안 vs B안 비교' in wb.sheetnames:
    del wb['⚖ A안 vs B안 비교']
ws = wb.create_sheet('⚖ A안 vs B안 비교')

for col, w in [('A',6),('B',24),('C',30),('D',30),('E',22)]:
    ws.column_dimensions[col].width = w

# ── 1. 제목 ───────────────────────────────────────────────────
ws.row_dimensions[1].height = 34
c(ws,1,1,'⚖  숙련도 옵션 밸런스 설계안 비교  (내부 검토용)',FC_TITLE,BG_DARK,True,15,'left')
ws.merge_cells('A1:E1')
ws.row_dimensions[2].height = 18
c(ws,2,1,'작성일: 2026-03-19   |   목적: 양손검 공격력 옵션 설계 방향 결정',FC_DIM,BG_DARK,sz=9,ha='left')
ws.merge_cells('A2:E2')

# ── 2. 문제 정의 ──────────────────────────────────────────────
ws.row_dimensions[4].height = 22
c(ws,4,1,'① 문제 정의',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells('A4:E4')

PROB = [
    ('근본 원인',
     'flat 수치(+ATK)는 티어가 높아질수록 기준 ATK 대비 % 기여도가 줄어든다.\n'
     '반면 % 보너스(공격속도, 치명타확률)는 어느 티어에서나 동일한 % DPS 효과를 낸다.\n'
     '→ 고티어에서 스탯 간 가치가 역전된다.'),
    ('현행 수치 (T7, CC=10% 기준)',
     'ATK +2 flat  =  0.33% DPS     ← 기준\n'
     '공격속도 +1%  =  1.00% DPS     (ATK의 3배)\n'
     '치명타확률 +1% = 1.67% DPS     (ATK의 5배)\n'
     '스킬가속   +1  =  0.99% DPS     (ATK의 3배)\n'
     '→ 양손검 공격력 옵션이 타 옵션 대비 1/3 이하 가치로 역전됨'),
]
for i, (k, v) in enumerate(PROB):
    r = 5 + i
    ws.row_dimensions[r].height = 72
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws,r,1,k,'00FF9966',bg,True,10,'left')
    c(ws,r,2,v,FC_VALUE,bg,sz=9,ha='left',wrap=True)
    ws.merge_cells(f'B{r}:E{r}')

# ── 3. 개요 비교 ──────────────────────────────────────────────
ws.row_dimensions[8].height = 22
c(ws,8,1,'② 설계안 개요',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells('A8:E8')

ws.row_dimensions[9].height = 20
for col, txt, fc in [(1,'구분',FC_SECTION),(2,'현행',YELLOW),(3,'A안',GREEN),(4,'B안',BLUE),(5,'비고',FC_DIM)]:
    c(ws,9,col,txt,fc,BG_SECTION,True,10)

OVERVIEW = [
    ('양손검 옵션', 'AttackVary\n(공격력 flat)',
     'AttackVary 유지\n(변경 없음)',
     'AttackVaryper로 변경\n(공격력 %)',
     '두 안의 핵심 차이'),
    ('T5~T6\n치명타확률', '+1%/step',
     '+0.5%/step (하향)',
     '+0.5%/step (동일)',
     '두 안 공통 적용'),
    ('T7\n공격속도',     '+1%/step',
     '+0.5%/step (하향)',
     '+1%/step 유지\n(ATK%와 이미 동가)',
     'T7만 차이 있음'),
    ('변경 범위',        '—',
     '최소\n(일부 % 수치 하향만)',
     '중간\n(ATK 옵션 타입 전환)',
     ''),
]
for i, row in enumerate(OVERVIEW):
    r = 10 + i
    ws.row_dimensions[r].height = 46
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    fcs = [FC_SECTION, YELLOW, GREEN, BLUE, FC_DIM]
    for ci, val in enumerate(row):
        c(ws,r,ci+1,val,fcs[ci],bg,sz=9,ha='left' if ci > 0 else 'center',wrap=True)

# ── 4. DPS 등가 검증 ──────────────────────────────────────────
ws.row_dimensions[15].height = 22
c(ws,15,1,'③ 단계별 지급 수치 비교  (CC=10% 기준 DPS% 병기)',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells('A15:G15')

# 열 너비 재설정 (컬럼 F, G 추가)
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 18

ws.row_dimensions[16].height = 20
for col, txt, fc in [(1,'티어',FC_SECTION),(2,'스탯',FC_SECTION),
                     (3,'A안  실제 수치',GREEN),(4,'A안 DPS%',GREEN),
                     (5,'B안  실제 수치',BLUE),(6,'B안 DPS%',BLUE),
                     (7,'판정',FC_SECTION)]:
    c(ws,16,col,txt,fc,BG_SECTION,True,9)

REF = {1:(80,12),2:(120,33),3:(180,61),4:(260,80),5:(360,110),6:(480,140),7:(600,177)}

def base_dps(atk, def_, cc=10, cd=200):
    return atk * (500/(def_+500)) * (1+(cc/100)*(cd/100)) * 0.91

def pct(b, n): return (n/b - 1)*100

# A안: flat ATK (3% 기준), CC·Spd는 현행 WEAPON_DATA 기준
# (atk_flat, cc%, spd%, accel)
A_STEP = {1:(2,1.8,3.0,3), 2:(4,1.8,3.0,3), 3:(2,0.6,1.0,1), 4:(2,0.6,1.0,1),
          5:(2,0.4,0.6,1), 6:(3,0.4,0.6,1), 7:(2,0.2,0.3,1)}
# B안: ATK% (3% 기준), CC·Spd는 A안과 동일 (옵션 변경 없음)
# (atk_pct%, cc%, spd%, accel)
B_STEP = {1:(3,1.8,3.0,3), 2:(3,1.8,3.0,3), 3:(1,0.6,1.0,1), 4:(1,0.6,1.0,1),
          5:(1,0.4,0.6,1), 6:(1,0.4,0.6,1), 7:(1,0.2,0.3,1)}

# (스탯명, A안 실제수치 문자열, A안 DPS%, B안 실제수치 문자열, B안 DPS%)
ROWS = []
for t in range(1,8):
    atk0, d = REF[t]
    b0 = base_dps(atk0, d)
    sa = A_STEP[t]; sb = B_STEP[t]

    # ATK
    a_atk_v = f'+{sa[0]} ATK'
    b_atk_v = f'+{sb[0]}% ATK'
    a_atk_d = pct(b0, base_dps(atk0+sa[0], d))
    b_atk_d = pct(b0, base_dps(atk0*(1+sb[0]/100), d))
    ROWS.append((t,'양손검 — 공격력', a_atk_v, a_atk_d, b_atk_v, b_atk_d))

    # CritChance
    cc_v = f'+{sa[1]:.1f}% 치명타확률'
    a_cc_d = pct(b0, base_dps(atk0,d,10+sa[1]))
    b_cc_d = pct(b0, base_dps(atk0,d,10+sb[1]))
    ROWS.append((t,'단검 — 치명타확률', cc_v, a_cc_d, cc_v, b_cc_d))

    # AtkSpd
    a_spd_v = f'+{sa[2]:.1f}% 공격속도'
    b_spd_v = f'+{sb[2]:.1f}% 공격속도'
    ROWS.append((t,'활 — 공격속도', a_spd_v, sa[2], b_spd_v, sb[2]))

    # SkillAccel
    acc_v = f'+{sa[3]} 스킬가속'
    acc_d = sa[3]/(100+sa[3])*100
    ROWS.append((t,'지팡이 — 스킬가속', acc_v, acc_d, acc_v, acc_d))

def judge(a, b):
    diff = abs(a - b)
    if diff < 0.25: return '✅ 동등'
    elif a > b:     return f'A우세\n(+{diff:.2f}%p)'
    else:           return f'B우세\n(+{diff:.2f}%p)'

prev_t = None
for i, (t, stat, a_val, a_d, b_val, b_d) in enumerate(ROWS):
    r = 17 + i
    ws.row_dimensions[r].height = 20
    bg = BG_DARK2 if t % 2 == 1 else BG_MED
    t_lbl = f'T{t}' if t != prev_t else ''
    tfc = TIER_FC[t-1] if t_lbl else FC_DIM
    c(ws,r,1,t_lbl,tfc,bg,bool(t_lbl),10)
    c(ws,r,2,stat,FC_VALUE,bg,sz=9,ha='left')

    # A안: 실제 수치 + DPS%
    diff = abs(a_d - b_d)
    fa = GREEN if diff < 0.25 else (ORANGE if a_d > b_d else GREEN)
    fb = GREEN if diff < 0.25 else (ORANGE if b_d > a_d else GREEN)
    c(ws,r,3,a_val, fa, bg, True, 10)
    c(ws,r,4,f'({a_d:.2f}% DPS)', FC_DIM, bg, sz=8)
    c(ws,r,5,b_val, fb, bg, True, 10)
    c(ws,r,6,f'({b_d:.2f}% DPS)', FC_DIM, bg, sz=8)
    jfc = GREEN if '✅' in judge(a_d,b_d) else YELLOW
    c(ws,r,7,judge(a_d,b_d),jfc,bg,sz=9,wrap=True)
    prev_t = t

# ── 5. 장단점 ─────────────────────────────────────────────────
RB = 17 + len(ROWS) + 2
ws.row_dimensions[RB-1].height = 22
c(ws,RB-1,1,'④ 장단점 비교',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells(f'A{RB-1}:G{RB-1}')

ws.row_dimensions[RB].height = 22
for col, txt, fc in [(2,'관점',FC_SECTION),(3,'A안  (flat 유지)',GREEN),
                     (5,'B안  (ATK%로 변경)',BLUE)]:
    c(ws,RB,col,txt,fc,BG_SECTION,True,10)
ws.merge_cells(f'C{RB}:D{RB}')
ws.merge_cells(f'E{RB}:G{RB}')

PROS = [
    ('장점', '유저 체감',
     '"+3 공격력" — 수치가 직접 눈에 보여\n체감과 이해가 직관적임',
     '수식 완결성:\nflat/% 이분법 해소,\nT7까지 완전 동가 달성'),
    ('장점', '밸런스\n보수성',
     '기존 시스템 최소 변경,\n캡 여유 풍부 (ATK flat은 캡 별도)',
     'ATK%=공격속도%=피해증가%\n완벽한 선형 등가'),
    ('단점', '고티어\n불균형',
     'T7: +2 ATK = 0.33% DPS\nvs 스킬가속 +1 = 0.99% DPS\n→ 3배 차이 여전히 잔존',
     'ATK% 캡 여유 감소:\n41% 소진 → 잔여 9%p\n향후 버프·장비 옵션 설계 제약'),
    ('단점', '유저 설명\n부담',
     '"왜 공격속도가 공격력보다 세냐"\n유저 문의·불만 가능성',
     '"+1% 공격력" 체감 불분명,\n실제 증가 수치 안 보여\n만족도 저하 우려'),
]
for i, (cat, viewpt, a_val, b_val) in enumerate(PROS):
    r = RB+1+i
    ws.row_dimensions[r].height = 62
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    cat_fc = GREEN if cat == '장점' else RED
    c(ws,r,1,cat,cat_fc,bg,True,9)
    c(ws,r,2,viewpt,FC_SECTION,bg,True,9,'left',wrap=True)
    fc_pro = GREEN if cat == '장점' else RED
    c(ws,r,3,a_val,fc_pro,bg,sz=9,ha='left',wrap=True)
    ws.merge_cells(f'C{r}:D{r}')
    c(ws,r,5,b_val,BLUE if cat == '장점' else RED,bg,sz=9,ha='left',wrap=True)
    ws.merge_cells(f'E{r}:G{r}')

# ── 6. 논의 포인트 ────────────────────────────────────────────
RD = RB+1+len(PROS)+2
ws.row_dimensions[RD-1].height = 22
c(ws,RD-1,1,'⑤ 팀 논의 포인트',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells(f'A{RD-1}:G{RD-1}')

ws.row_dimensions[RD].height = 20
for col, txt, fc in [(2,'논의 주제',FC_SECTION),(3,'A안 관점',GREEN),
                     (5,'B안 관점',BLUE),(7,'결정 기준',YELLOW)]:
    c(ws,RD,col,txt,fc,BG_SECTION,True,9)
ws.merge_cells(f'C{RD}:D{RD}')
ws.merge_cells(f'E{RD}:F{RD}')

DISC = [
    ('Q1', '유저 체감 vs 수식 정확성\n어느 쪽을 우선하는가?',
     '체감 명확 → A안 유리',
     '수식 완결 → B안 유리',
     '게임 단계·유저층 성향'),
    ('Q2', 'T7 잔류 불균형\n(+2 ATK = 0.33% vs 스킬가속 +1 = 0.99%)\n허용 가능한 오차인가?',
     'A안 선택 시 핵심 결정사항.\n엔드콘텐츠 이슈로 유예 가능?',
     'B안은 해당 없음 (완전 해소)',
     'T7 콘텐츠 오픈 시점'),
    ('Q3', 'ATK% 캡(50%) 소진이\n향후 설계에 문제가 되는가?',
     'flat ATK 별도 캡 → 여유 충분',
     'ATK% 41% 소진 → 잔여 9%p\n타 시스템 설계 시 주의',
     '향후 버프·장비 옵션 계획'),
    ('Q4', '향후 flat ATK vs ATK%\n어느 방향으로 장비 옵션을 추가할 것인가?',
     'flat 추가 예정 → A안 일관성 유지',
     'ATK% 추가 예정 → B안과 시너지',
     '장기 장비 설계 방향'),
]
for i, (q, topic, a_v, b_v, crit) in enumerate(DISC):
    r = RD+1+i
    ws.row_dimensions[r].height = 58
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws,r,1,q,YELLOW,bg,True,10)
    c(ws,r,2,topic,FC_VALUE,bg,sz=9,ha='left',wrap=True)
    c(ws,r,3,a_v,GREEN,bg,sz=9,ha='left',wrap=True)
    ws.merge_cells(f'C{r}:D{r}')
    c(ws,r,5,b_v,BLUE,bg,sz=9,ha='left',wrap=True)
    ws.merge_cells(f'E{r}:F{r}')
    c(ws,r,7,crit,YELLOW,bg,sz=9,ha='left',wrap=True)

# ── 7. 검토 의견 ──────────────────────────────────────────────
RR = RD+1+len(DISC)+2
ws.row_dimensions[RR-1].height = 22
c(ws,RR-1,1,'⑥ 검토 의견',FC_SECTION,BG_HEADER,True,12,'left')
ws.merge_cells(f'A{RR-1}:G{RR-1}')

ws.row_dimensions[RR].height = 100
c(ws,RR,1,
  '▶ 단기 권장 (현재 T1~T5 콘텐츠 단계):\n'
  '   A안 — 변경 범위 최소, 유저 체감 명확.\n'
  '   T7 잔류 불균형(+2 ATK = 0.33% vs 스킬가속 +1 = 0.99%)은 T7 콘텐츠 오픈 전까지 유예 가능.\n\n'
  '▶ 장기 검토 (T6~T7 콘텐츠 오픈 시점):\n'
  '   B안 전환 고려 — 고티어 밸런스 완결.\n'
  '   단, ATK% 캡(잔여 9%p) 정책 및 유저 UI 설명 방식 병행 논의 필요.\n\n'
  '▶ 현행 적용 완료:\n'
  '   T5~T7 치명타확률·공격속도 하향 (CC: T1/T2=1.8%, T3/T4=0.6%, T5/T6=0.4%, T7=0.2%)',
  FC_VALUE, BG_MED, sz=10, ha='left', va='top', wrap=True)
ws.merge_cells(f'A{RR}:G{RR}')

wb.save('기획서/X7_숙련도_옵션_기획서.xlsx')
print(f'완료 — 시트 추가됨: ⚖ A안 vs B안 비교  (row {RR}까지)')
