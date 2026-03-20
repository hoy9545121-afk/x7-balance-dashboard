"""
X7_무기_스킬.xlsx 내용 수정 (constants.py 기준):
- 각 무기 시트 3번 스킬셋 Lv.1 MP 교정
- 스킬 리스트 시트 비고 교정
"""
import openpyxl

XLSX_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"

# ── constants.py 기준 3번 스킬셋 Lv.1 데이터 ──────────────────
# (cmd, cd, mp, dmg_ratio)  dmg_ratio = dmg% / 100 (xlsx 단위)
SKILLS_3T_LV1 = {
    "한손검": {
        "Q": (6.0, 10, 1.5),
        "W": (12.0, 20, 2.0),
        "E": (20.0, 30, 2.5),
        "R": (45.0, 100, 5.0),
    },
    "양손검": {
        "Q": (5.0, 15, 2.0),
        "W": (12.0, 30, 2.5),
        "E": (15.0, 25, 3.0),
        "R": (40.0, 100, 4.0),
    },
    "활": {
        "Q": (5.0, 10, 2.2),
        "W": (12.0, 20, 2.8),
        "E": (20.0, 30, 3.5),
        "R": (45.0, 100, 5.0),
    },
    "지팡이": {
        "Q": (4.0, 20, 2.0),
        "W": (15.0, 40, 4.0),
        "E": (40.0, 50, 6.0),
        "R": (20.0, 100, 3.0),
    },
    "단검": {
        "Q": (6.0, 15, 2.2),
        "W": (13.0, 30, 1.2),
        "E": (15.0, 40, 2.5),
        "R": (40.0, 100, 5.0),
    },
}

# 각 무기 시트의 3번 스킬셋 Lv.1 행 번호 (실측)
# col: B=2(cmd), C=3(lv), D=4(cd), E=5(mp), G=7(dmg)
WEAPON_LV1_ROWS = {
    "한손검": {"Q": 77, "W": 82, "E": 87, "R": 92},
    "양손검": {"Q": 77, "W": 82, "E": 87, "R": 92},
    "활":    {"Q": 83, "W": 88, "E": 93, "R": 98},
    "지팡이": {"Q": 125, "W": 130, "E": 135, "R": 140},
    "단검":  {"Q": 77, "W": 82,  "E": 87,  "R": 92},
}

# 스킬 리스트 3번 비고 수정 (행번호: 올바른 비고 문자열)
SKILL_LIST_G_FIXES = {
    # 한손검 3번
    11: "cd=6s/MP=10/dmg=150%(치명타)",
    12: "cd=12s/MP=20/dmg=200%+방어버프(+30%/5s)",
    13: "cd=20s/MP=30/dmg=250%",
    14: "cd=45s/MP=100/dmg=500%",
    # 양손검 3번
    23: "cd=5s/MP=15/dmg=200%~(강화평타 부여)",
    24: "cd=12s/MP=30/dmg=250%~",
    25: "cd=15s/6-hit/MP=25/dmg=300%~",
    26: "cd=40s/MP=100/dmg=400%~",
    # 지팡이 3번
    35: "cd=4~3s/MP=20~30/dmg=200%~280%",
    38: "cd=20s/CC(부유)/MP=100/dmg=300%~420%",
    # 단검 3번
    47: "cd=6~4s/MP=15~20/dmg=220%~308%(치명타)",
    48: "cd=13s/MP=30~35/버프(공속+45%)+dmg=120%~168%",
    49: "cd=15~11s/MP=40~55/dmg=250%~350%",
    50: "cd=40~32s/MP=100~140/dmg=500%~700%",
    # 활 3번
    59: "cd=5~3s/MP=10~15/dmg=220%~308%(치명타)",
    60: "cd=12~8s/MP=20~30/dmg=280%~392%",
    61: "cd=20~16s/MP=30~45/공속+50%~70%+dmg=350%~490%",
    62: "cd=45~37s/MP=100~140/CC(속박)+dmg=500%~700%",
}

# ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH)

# ── 1. 각 무기 시트 Lv.1 MP/CD/DMG 교정 ──────────────────────
print("=== 3번 스킬셋 Lv.1 값 교정 ===")
fixed_count = 0

for weapon, skill_rows in WEAPON_LV1_ROWS.items():
    ws = wb[weapon]
    expected = SKILLS_3T_LV1[weapon]
    for cmd, lv1_row in skill_rows.items():
        exp_cd, exp_mp, exp_dmg = expected[cmd]
        issues = []

        # CD (col D=4)
        cell_cd = ws.cell(row=lv1_row, column=4)
        cur_cd = cell_cd.value
        try:
            if cur_cd is not None and abs(float(cur_cd) - exp_cd) > 0.01:
                issues.append(f"CD {cur_cd}→{exp_cd}")
                cell_cd.value = exp_cd
        except (TypeError, ValueError):
            pass

        # MP (col E=5)
        cell_mp = ws.cell(row=lv1_row, column=5)
        cur_mp = cell_mp.value
        try:
            if cur_mp is not None and int(float(cur_mp)) != exp_mp:
                issues.append(f"MP {cur_mp}→{exp_mp}")
                cell_mp.value = exp_mp
        except (TypeError, ValueError):
            pass

        # DMG (col G=7)
        cell_dmg = ws.cell(row=lv1_row, column=7)
        cur_dmg = cell_dmg.value
        try:
            if cur_dmg is not None and abs(float(cur_dmg) - exp_dmg) > 0.001:
                issues.append(f"DMG {cur_dmg}→{exp_dmg}")
                cell_dmg.value = exp_dmg
        except (TypeError, ValueError):
            pass

        if issues:
            print(f"  ✏ {weapon} {cmd} R{lv1_row}: {', '.join(issues)}")
            fixed_count += len(issues)
        else:
            print(f"  ✓ {weapon} {cmd} R{lv1_row}: OK")

print(f"\n수정된 셀 수: {fixed_count}")

# ── 2. 스킬 리스트 비고 교정 ──────────────────────────────────
print("\n=== 스킬 리스트 비고 교정 ===")
ws_list = wb["스킬 리스트"]
list_fixed = 0
for row, new_val in SKILL_LIST_G_FIXES.items():
    cell = ws_list.cell(row=row, column=7)  # G열
    old_val = cell.value
    if old_val != new_val:
        print(f"  R{row:2d} G: '{str(old_val)[:40]}'\n     → '{new_val}'")
        cell.value = new_val
        list_fixed += 1
    else:
        print(f"  R{row:2d} G: OK")

print(f"\n스킬 리스트 수정 셀: {list_fixed}")

# ── 3. 저장 ───────────────────────────────────────────────────
try:
    wb.save(XLSX_PATH)
    print(f"\n저장 완료: {XLSX_PATH}")
except PermissionError:
    alt = XLSX_PATH.replace(".xlsx", "_fixed.xlsx")
    wb.save(alt)
    print(f"\n원본 잠김 — 대체 저장: {alt}")
