# 방어구 옵션 밸런싱 Excel 생성기
# Input:  C:/Users/hoy5343/Downloads/방어구_옵션_밸런싱_최종_V2.csv
# Output: C:/Users/hoy5343/Downloads/방어구_옵션_밸런싱_최종_V2.xlsx
import csv, re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN_PATH  = r"C:\Users\hoy5343\Downloads\방어구_옵션_밸런싱_최종_V2.csv"
OUT_PATH = r"C:\AI_simulator\기획서\방어구_옵션_밸런싱_최종_V2.xlsx"

# ── 색상 ──────────────────────────────────────────────────
TIER_BG    = {"1T": "2C3E50", "2T": "1A5276", "3T": "1B2631"}
TIER_FG    = {"1T": "E8B84B", "2T": "AED6F1", "3T": "F0B27A"}
TIER_LABEL = {
    "1T": "1티어  —  기본 제작",
    "2T": "2티어  —  던전 입문",
    "3T": "3티어  —  본격 성장",
}

ACQ_BG = {
    "기본제작":   "4D5656",
    "전리품제작": "1E6B3E",
    "던전코어":   "1A4F72",
    "드랍(Boss)": "7B241C",
}
ACQ_ICON = {
    "기본제작":   "🔨",
    "전리품제작": "💎",
    "던전코어":   "🔮",
    "드랍(Boss)": "👺",
}

MAT_HDR_BG = {"판금": "5D6D7E", "가죽": "935116", "천": "6C3483"}
MAT_ROW_A  = {"판금": "EBF5FB", "가죽": "FEF5E7", "천": "F5EEF8"}  # 짝수 행
MAT_ROW_B  = {"판금": "D6EAF8", "가죽": "FDEBD0", "천": "E8DAEF"}  # 홀수 행
MAT_ICON   = {"판금": "⚙", "가죽": "🟤", "천": "🟣"}

# 스킬 레벨별 헤더 색상
REINFORCE_HDR_BG = ["808B96", "AED6F1", "A9DFBF", "52BE80", "1E8449", "F4D03F"]
REINFORCE_HDR_FG = ["FFFFFF", "1A5276", "145A32", "FFFFFF", "FFFFFF", "7D6608"]

# 스킬 레벨별 셀 색상
REINFORCE_ROW_BG = ["F2F3F4", "D6EAF8", "D5F5E3", "A9DFBF", "52BE80", "F9E79F"]
REINFORCE_ROW_FG = ["555555", "1A5276", "145A32", "0B5345", "145A32", "7D6608"]
REINFORCE_BOLD   = [False,    False,    False,    False,    True,     True   ]

PARTS_ICON = {"아머": "🛡", "헬멧": "⛑", "장갑": "🥊", "신발": "👟"}
COL_WIDTHS = [8, 22, 28, 9, 9, 9, 9, 9, 9]

ACQ_ORDER = ["기본제작", "전리품제작", "던전코어", "드랍(Boss)"]
MAT_ORDER = ["판금", "가죽", "천"]


# ── 유틸 ────────────────────────────────────────────────
def make_fill(color):
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def group_border(is_last):
    """아이템 그룹 마지막 행이면 굵은 하단 테두리, 아니면 얇은 테두리."""
    t = Side(style="thin",   color="CCCCCC")
    b = Side(style="medium", color="888888") if is_last else Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=b)

def sc(cell, bg=None, fg="000000", bold=False, size=10,
       h="center", v="center", italic=False, wrap=False, border=None):
    if bg:
        cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border = border if border is not None else thin_border()

def parse_option(raw):
    """'DamageDownVaryper (받는 피해 감소)' → ('받는 피해 감소', 'DamageDownVaryper')"""
    m = re.match(r"^(\S+)\s+\((.+)\)\s*$", raw.strip())
    if m:
        return m.group(2), m.group(1)   # (kr_name, en_key)
    return raw.strip(), raw.strip()


# ── 시트 작성 ────────────────────────────────────────────
def write_sheet(wb, tier, rows):
    ws = wb.create_sheet(title=f"{tier} 방어구")

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # 타이틀
    ws.merge_cells(f"A{r}:I{r}")
    c = ws[f"A{r}"]
    c.value = f"  ⚔️  X7  방어구 옵션 밸런싱   /   {TIER_LABEL[tier]}"
    sc(c, bg=TIER_BG[tier], fg=TIER_FG[tier], bold=True, size=13, h="left")
    ws.row_dimensions[r].height = 32
    r += 1

    # 컬럼 헤더
    col_headers = ["파츠", "옵션명 (한국어)", "옵션 Key (영문)",
                   "Lv.1", "Lv.2", "Lv.3", "Lv.4", "Lv.5"]
    for i, h in enumerate(col_headers, 1):
        col = get_column_letter(i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        if i <= 3:
            sc(cell, bg="2C3E50", fg="FFFFFF", bold=True, size=9)
        else:
            k = i - 4
            sc(cell, bg=REINFORCE_HDR_BG[k], fg=REINFORCE_HDR_FG[k], bold=True, size=9)
    ws.row_dimensions[r].height = 20
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    # 데이터 그룹화
    grouped = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row["획득"]][row["재질"]].append(row)

    for acq in ACQ_ORDER:
        if acq not in grouped:
            continue

        # 획득 방법 헤더
        ws.merge_cells(f"A{r}:I{r}")
        c = ws[f"A{r}"]
        c.value = f"  {ACQ_ICON.get(acq, '📦')}  획득 방법 :  {acq}"
        sc(c, bg=ACQ_BG.get(acq, "555555"), fg="FFFFFF", bold=True, size=11, h="left")
        ws.row_dimensions[r].height = 24
        r += 1

        for mat in MAT_ORDER:
            if mat not in grouped[acq]:
                continue
            mat_rows = grouped[acq][mat]

            # 재질 서브헤더
            ws.merge_cells(f"A{r}:I{r}")
            c = ws[f"A{r}"]
            c.value = f"        {MAT_ICON.get(mat, '')}  재질 :  {mat}"
            sc(c, bg=MAT_HDR_BG[mat], fg="FFFFFF", bold=True, size=10, h="left")
            ws.row_dimensions[r].height = 20
            r += 1

            # ── 아이템 그룹 단위로 묶기 ──────────────────────────────
            prev_grp = None
            item_groups = []
            for row_data in mat_rows:
                grp = row_data.get("그룹", "")
                if grp != prev_grp:
                    item_groups.append([])
                    prev_grp = grp
                item_groups[-1].append(row_data)

            for item_idx, item_rows in enumerate(item_groups):
                row_bg = MAT_ROW_B[mat] if item_idx % 2 else MAT_ROW_A[mat]
                n = len(item_rows)

                # 파츠 셀 세로 병합 (옵션 2개 이상)
                if n > 1:
                    ws.merge_cells(f"A{r}:A{r + n - 1}")

                for sub_idx, row_data in enumerate(item_rows):
                    is_last = (sub_idx == n - 1)
                    brd = group_border(is_last)

                    kr_name, en_key = parse_option(row_data["옵션명"])
                    parts = row_data["파츠"]
                    icon = PARTS_ICON.get(parts, "")

                    # 파츠: 첫 행에만 표시 (병합 셀)
                    if sub_idx == 0:
                        ws[f"A{r}"].value = f"{icon}  {parts}"
                    sc(ws[f"A{r}"], bg=row_bg, h="center", v="center", size=10, border=brd)

                    ws[f"B{r}"].value = kr_name
                    sc(ws[f"B{r}"], bg=row_bg, h="left", size=10, border=brd)

                    ws[f"C{r}"].value = en_key
                    sc(ws[f"C{r}"], bg=row_bg, fg="888888", h="left", size=8,
                       italic=True, border=brd)

                    is_varyper = "Varyper" in en_key
                    for k in range(5):
                        val = row_data.get(f"{k}강", "").strip()
                        col = get_column_letter(4 + k)
                        if val:
                            try:
                                nv = int(val)
                                if is_varyper:
                                    pct = nv / 10
                                    display_val = f"{pct:.0f}%" if pct == int(pct) else f"{pct:.1f}%"
                                else:
                                    display_val = nv
                            except ValueError:
                                display_val = val
                        else:
                            display_val = ""
                        ws[f"{col}{r}"].value = display_val
                        sc(ws[f"{col}{r}"],
                           bg=REINFORCE_ROW_BG[k], fg=REINFORCE_ROW_FG[k],
                           bold=REINFORCE_BOLD[k], size=10, border=brd)

                    ws.row_dimensions[r].height = 18
                    r += 1

        r += 1  # 획득 그룹 사이 빈 행


# ── 설계 의도 시트 ──────────────────────────────────────
def write_design_intent_sheet(wb):
    ws = wb.create_sheet(title="밸런스 설계 의도", index=0)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    def sec_header(r, num, title):
        ws.merge_cells(f"A{r}:G{r}")
        c = ws[f"A{r}"]
        c.value = f"  {num}.  {title}"
        sc(c, bg="2C3E50", fg="E8B84B", bold=True, size=12, h="left")
        ws.row_dimensions[r].height = 28

    def sub_header(r, title):
        ws.merge_cells(f"B{r}:G{r}")
        c = ws[f"B{r}"]
        c.value = f"  ▪  {title}"
        sc(c, bg="34495E", fg="AED6F1", bold=True, size=10, h="left")
        ws.row_dimensions[r].height = 22

    def body_row(r, text, bg="EBF5FB"):
        ws.merge_cells(f"B{r}:G{r}")
        c = ws[f"B{r}"]
        c.value = text
        sc(c, bg=bg, fg="1A2A3A", h="left", size=10, wrap=True)
        ws.row_dimensions[r].height = 36

    def bullet_row(r, label, desc, lbg="D6EAF8", dbg="EBF5FB"):
        ws[f"B{r}"].value = f"▶  {label}"
        sc(ws[f"B{r}"], bg=lbg, fg="1A5276", bold=True, size=10, h="left")
        ws.merge_cells(f"C{r}:G{r}")
        c = ws[f"C{r}"]
        c.value = desc
        sc(c, bg=dbg, fg="1A2A3A", h="left", size=10, wrap=True)
        ws.row_dimensions[r].height = 32

    def tbl_header(r, cols, bgs):
        for i, (col_ltr, h, bg) in enumerate(zip(cols, bgs[0], bgs[1])):
            ws[f"{col_ltr}{r}"].value = h
            sc(ws[f"{col_ltr}{r}"], bg=bg, fg="FFFFFF", bold=True, size=9)
        ws.row_dimensions[r].height = 20

    def blank(r):
        ws.row_dimensions[r].height = 10

    r = 1

    # ── 타이틀 ──
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "  🛡️  방어구 옵션 수치 밸런싱  —  설계 리포트"
    sc(c, bg="1B2631", fg="F0B27A", bold=True, size=15, h="left")
    ws.row_dimensions[r].height = 38
    r += 1; blank(r); r += 1

    # ── 1. 설계 목적 및 방향 ──
    sec_header(r, "1", "설계 목적 및 방향"); r += 1
    body_row(r,
        "본 리포트는 캐릭터 기본 능력치와 장비 사양을 바탕으로, "
        "유저에게 스킬 레벨업의 성취감과 상위 티어 장비 획득의 동기를 부여하기 위해 작성되었습니다."
    ); r += 1

    bullets_1 = [
        ("스킬 레벨업의 유의미한 체감",
         "모든 스킬 레벨에서 이전 레벨 대비 약 15%의 성장(최소 10% 이상)을 보장하여 "
         "스킬 Lv.5 달성 시 Lv.1 대비 2배의 성능을 냅니다."),
        ("티어 간 위계 확립",
         "하위 티어의 최대 강화 장비보다 상위 티어의 기본 장비가 더 높은 가치를 지니도록 설계하여 "
         "지속적인 파밍을 유도합니다."),
        ("능력치 가치 평준화",
         "공격력, 방어력, 체력 등 각 옵션이 전투력 향상에 기여하는 효율을 "
         "동일 수준(약 3~5%)으로 맞췄습니다."),
    ]
    for label, desc in bullets_1:
        bullet_row(r, label, desc); r += 1

    r += 1

    # ── 2. 전투 공식 기반 가치 분석 ──
    sec_header(r, "2", "전투 공식 기반 가치 분석"); r += 1
    body_row(r,
        "전투 효율 조율 상수 K=500을 기준으로, 각 능력치 1포인트가 주는 "
        "실질적인 생존 및 공격 기여도를 분석하여 수치를 배분했습니다."
    ); r += 1

    # 가치 분석 표
    ws[f"B{r}"].value = "능력치 종류"
    sc(ws[f"B{r}"], bg="1A5276", fg="FFFFFF", bold=True, size=9)
    ws[f"C{r}"].value = "기준 수치 (1T Lv.1)"
    sc(ws[f"C{r}"], bg="1A5276", fg="FFFFFF", bold=True, size=9)
    ws.merge_cells(f"D{r}:G{r}")
    ws[f"D{r}"].value = "기대 효율"
    sc(ws[f"D{r}"], bg="1A5276", fg="FFFFFF", bold=True, size=9)
    ws.row_dimensions[r].height = 20; r += 1

    analysis_rows = [
        ("공격력 (AttackVary)",                  "5",   "최종 데미지 약 4.1% 상승"),
        ("방어력 (DefenseVary)",                 "20",  "유효 생명력(EHP) 약 2.7% 상승"),
        ("최대 생명력 (MaxHpVary)",              "75",  "전체 생존력 약 3.1% 상승"),
        ("피해 증가 / 치명타 확률 (per)",        "30",  "실질 DPS/생존력 약 3.0% 상승"),
    ]
    for idx, (name, base, eff) in enumerate(analysis_rows):
        bg = "F2F3F4" if idx % 2 == 0 else "EBF5FB"
        ws[f"B{r}"].value = name
        sc(ws[f"B{r}"], bg=bg, fg="1A2A3A", h="left", size=10)
        ws[f"C{r}"].value = base
        sc(ws[f"C{r}"], bg=bg, fg="1A2A3A", size=10)
        ws.merge_cells(f"D{r}:G{r}")
        ws[f"D{r}"].value = eff
        sc(ws[f"D{r}"], bg=bg, fg="1A2A3A", h="left", size=10)
        ws.row_dimensions[r].height = 18; r += 1

    r += 1

    # ── 3. 세부 설계 규칙 ──
    sec_header(r, "3", "세부 설계 규칙"); r += 1

    sub_header(r, "3.1  스킬 레벨 성장 로직"); r += 1
    bullet_row(r, "성장 계수",
        "1.15ⁿ (15% 복리 성장 적용) — 모든 스킬 레벨에서 이전 레벨 대비 최소 10% 이상 성장을 보장합니다.",
        lbg="D5F5E3", dbg="EAFAF1"); r += 1
    bullet_row(r, "라운딩 룰",
        "유저가 수치를 직관적으로 인지할 수 있도록 0 또는 5 단위로 끝자리를 보정합니다.",
        lbg="D5F5E3", dbg="EAFAF1"); r += 1

    r += 1
    sub_header(r, "3.2  티어별 스케일링 (Tier Scaling)"); r += 1
    body_row(r,
        "상위 티어 장비의 선호도를 보장하기 위해 티어별 베이스 멀티플라이어를 적용했습니다.",
        bg="EAFAF1"); r += 1

    tier_bullets = [
        ("1T  (×1.0, 기준)",    "기본 제작 장비. 초반 플레이의 핵심 성장 도구."),
        ("2T  (×1.6)",           "1T 스킬 Lv.5 장비와 유사한 수준에서 시작 → 파밍 동기 부여."),
        ("3T  (×2.5)",           "압도적인 기본 성능 차이 발생 → 지속적인 던전 탐색 유도."),
    ]
    tier_bgs = [("D5F5E3","EAFAF1"), ("A9DFBF","D5F5E3"), ("52BE80","A9DFBF")]
    for (label, desc), (lbg, dbg) in zip(tier_bullets, tier_bgs):
        bullet_row(r, label, desc, lbg=lbg, dbg=dbg); r += 1

    r += 1

    # ── 4. 티어별 주요 수치 배치 결과 ──
    sec_header(r, "4", "티어별 주요 수치 배치 결과"); r += 1

    hdr_cols = ["B","C","D","E","F","G"]
    hdr_vals = ["티어","옵션명","Lv.1 (기본)","Lv.2","Lv.4","Lv.5 (MAX)"]
    hdr_bgs  = ["1A5276","1A5276","808B96","52BE80","1E8449","F4D03F"]
    hdr_fgs  = ["FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF","7D6608"]
    for col, val, bg, fg in zip(hdr_cols, hdr_vals, hdr_bgs, hdr_fgs):
        ws[f"{col}{r}"].value = val
        sc(ws[f"{col}{r}"], bg=bg, fg=fg, bold=True, size=9)
    ws.row_dimensions[r].height = 20; r += 1

    result_rows = [
        ("1T","방어력",   "20","24","30","34"),
        ("1T","최대 생명력","75","85","115","130"),
        ("2T","방어력",   "32","36","48","55"),
        ("2T","최대 생명력","120","140","185","210"),
        ("3T","방어력",   "50","55","75","85"),
        ("3T","최대 생명력","190","220","290","330"),
    ]
    tier_row_bg = {"1T": ("EBF5FB","D6EAF8"), "2T": ("FEF9E7","FDEBD0"), "3T": ("F5EEF8","E8DAEF")}
    for idx, (tier, opt, v0, v1, v3, v5) in enumerate(result_rows):
        bg = tier_row_bg[tier][idx % 2]
        vals = [tier, opt, v0, v1, v3, v5]
        reinforce_bgs = [bg, bg, "F2F3F4","A9DFBF","52BE80","F9E79F"]
        reinforce_fgs = ["333333","333333","555555","0B5345","145A32","7D6608"]
        for col, val, rbg, rfg in zip(hdr_cols, vals, reinforce_bgs, reinforce_fgs):
            ws[f"{col}{r}"].value = val
            sc(ws[f"{col}{r}"], bg=rbg, fg=rfg,
               bold=(col in ["F","G"]), size=10,
               h="left" if col in ["B","C"] else "center")
        ws.row_dimensions[r].height = 18; r += 1

    r += 1

    # ── 5. 최종 제언 ──
    sec_header(r, "5", "최종 제언"); r += 1
    body_row(r,
        "본 밸런싱안은 특정 옵션이 버려지지 않고 유저의 플레이 스타일에 따라 고르게 선택될 수 있도록 설계되었습니다. "
        "특히 1T 스킬 Lv.5 방어력(34)보다 3T 스킬 Lv.1 방어력(50)이 높게 설정되어 있어, "
        "고레벨 저티어 장비보다 저레벨 고티어 장비가 실전에서 더 유리하게 작용합니다.",
        bg="FEF9E7"
    ); r += 1

    r += 1

    # ── 6. 사용 능력치 목록 및 티어별 최대 수치 ──
    sec_header(r, "6", "사용 능력치 목록 및 티어별 최대 수치  (스킬 Lv.5 기준 / 단일 부위)"); r += 1
    body_row(r,
        "단일 부위 장착 기준입니다. 같은 능력치 옵션이 여러 부위에 존재할 경우 중복 적용되어 수치가 합산됩니다."
    ); r += 1

    # 테이블 헤더
    stat_tbl_cols = [
        ("B", "능력치명 (한국어)"),
        ("C", "Key (영문)"),
        ("D", "적용 부위"),
        ("E", "1T  Lv.5"),
        ("F", "2T  Lv.5"),
        ("G", "3T  Lv.5"),
    ]
    for col, title in stat_tbl_cols:
        ws[f"{col}{r}"].value = title
        sc(ws[f"{col}{r}"], bg="2C3E50", fg="FFFFFF", bold=True, size=9)
    ws.row_dimensions[r].height = 20; r += 1

    # 카테고리별 능력치 데이터
    # (카테고리명, 헤더BG, 홀수행BG, 짝수행BG, 스탯목록)
    # [%] 표기 = 천분률(1000=100%) 환산값 / 나머지는 원시값
    # 수치: 1T Lv.5 → 2T Lv.5 → 3T Lv.5 (단일 부위 기준)
    STAT_SECTIONS = [
        ("⚔  공격", "922B21", "FDEDEC", "FADBD8", [
            ("공격력",          "AttackVary",              "아머·헬멧·장갑·신발",    9,    14,    22),
            ("공격 속도 [%]",   "AtkSpeedVaryper",         "아머·헬멧·장갑·신발", "3.5%","5.6%","8.7%"),
            ("치명타 확률 [%]", "CriVaryper",              "아머·헬멧·장갑·신발", "2.1%","3.3%","5.2%"),
            ("치명타 피해 [%]", "CriDamageVaryper",        "아머·헬멧·장갑·신발",  "-","16.6%","26.1%"),
            ("피해 증가 [%]",   "DamageUpVaryper",         "장갑·신발",            "-",   "-",   "7%"),
            ("PVE 피해 증가 [%]","PVEDamageUpVaryper",     "아머·장갑·신발",       "-",   "-",   "7%"),
        ]),
        ("🛡  방어 / 생존", "1A5276", "EBF5FB", "D6EAF8", [
            ("방어력",               "DefenseVary",           "아머·장갑·신발",        34,    55,    85),
            ("최대 생명력",          "MaxHpVary",             "아머·헬멧·장갑·신발",  130,   210,   330),
            ("받는 피해 감소 [%]",   "DamageDownVaryper",     "아머·헬멧·장갑·신발", "1.7%","2.8%","4.4%"),
            ("상태이상 저항력",      "SCNegativeRecoveryVary","헬멧·장갑",             "-",   "-",    87),
        ]),
        ("💚  회복", "145A32", "EAFAF1", "D5F5E3", [
            ("생명력 자연 회복",  "RegenHpVary",           "헬멧·신발",               9,    14,    22),
            ("치유력 [%]",        "HealAmpVaryper",        "아머·헬멧·장갑·신발",   "-",   "-", "7.5%"),
            ("받는 치유량",       "HealAcceptVary",        "장갑",                   "-",   "-",   131),
        ]),
        ("💙  마나 / 스킬", "6C3483", "F5EEF8", "E8DAEF", [
            ("최대 마력",          "MaxMpVary",             "아머",                   34,    55,    85),
            ("마력 자연 회복",     "RegenMpVary",           "아머·헬멧·장갑·신발",     7,     9,    14),
            ("스킬 가속 ※",       "SkillCooldownAccVary",  "아머·헬멧·장갑·신발",    "-",   14,    21),
        ]),
    ]
    # [%] = 천분률÷10  /  스킬 가속 ※ = 쿨타임 충전 속도 (100+n)배, CD×100/(100+n) 공식

    TIER_COL_BG = {"E": "D6EAF8", "F": "D5F5E3", "G": "F9E79F"}
    TIER_COL_FG = {"E": "1A5276", "F": "145A32", "G": "7D6608"}

    for cat_name, cat_bg, row_bg_a, row_bg_b, stats in STAT_SECTIONS:
        # 카테고리 구분 헤더
        ws.merge_cells(f"B{r}:G{r}")
        c = ws[f"B{r}"]
        c.value = f"   {cat_name}"
        sc(c, bg=cat_bg, fg="FFFFFF", bold=True, size=10, h="left")
        ws.row_dimensions[r].height = 22; r += 1

        for idx, (kr, en, parts, t1, t2, t3) in enumerate(stats):
            row_bg = row_bg_a if idx % 2 == 0 else row_bg_b

            ws[f"B{r}"].value = kr
            sc(ws[f"B{r}"], bg=row_bg, fg="1A2A3A", bold=True, size=10, h="left")

            ws[f"C{r}"].value = en
            sc(ws[f"C{r}"], bg=row_bg, fg="888888", size=8, italic=True, h="left")

            ws[f"D{r}"].value = parts
            sc(ws[f"D{r}"], bg=row_bg, fg="444444", size=9, h="center", wrap=True)

            for col, val in [("E", t1), ("F", t2), ("G", t3)]:
                is_dash = (val == "-")
                cell = ws[f"{col}{r}"]
                cell.value = val
                sc(cell,
                   bg="DDDDDD" if is_dash else TIER_COL_BG[col],
                   fg="AAAAAA" if is_dash else TIER_COL_FG[col],
                   bold=(col == "G" and not is_dash),
                   size=10)

            ws.row_dimensions[r].height = 18; r += 1

        r += 1  # 카테고리 사이 빈 행

    # 범례
    ws.merge_cells(f"B{r}:G{r}")
    c = ws[f"B{r}"]
    c.value = "  ※  [%] 수치 = 천분률 ÷ 10  (예: 100천분률 = 10%)  /  스킬 가속 ※ : CD × 100 / (100 + n) 공식 적용, n=25 시 CD -20%"
    sc(c, bg="F8F9FA", fg="777777", size=8, italic=True, h="left")
    ws.row_dimensions[r].height = 16


# ── 메인 ────────────────────────────────────────────────
def main():
    rows = []
    with open(IN_PATH, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tier in ["1T", "2T", "3T"]:
        tier_rows = [r for r in rows if r["티어"] == tier]
        if tier_rows:
            write_sheet(wb, tier, tier_rows)

    write_design_intent_sheet(wb)

    wb.save(OUT_PATH)
    print(f"✅ 저장 완료: {OUT_PATH}")
    print(f"  시트: {' / '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
