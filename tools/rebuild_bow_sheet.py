"""
활(Bow) 시트를 단검 시트와 동일한 서식으로 완전 재구성
- 다크 네이비 배경
- Q/W/E/R 색상 코딩
- DPS 자동계산 수식 (I=DPS, J=분당횟수, K=1분누적, L=1분마나)
- 열 구조: B~L (단검과 동일)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"

# ── 색상 상수 (단검 시트와 동일한 팔레트) ──────────────────────────
C_TITLE_BG   = "1F2D3D"   # 타이틀 배경 (다크 네이비)
C_TITLE_FG   = "E8B84B"   # 타이틀 텍스트 (골드)
C_DESC_BG    = "2C3E50"   # 설명 배경
C_DESC_FG    = "BDC3C7"   # 설명 텍스트
C_SEC_BG     = "34495E"   # 섹션 헤더 배경
C_SEC_FG     = "ECF0F1"   # 섹션 헤더 텍스트
C_COL_BG     = "2C3E50"   # 컬럼 헤더 배경
C_COL_FG     = "FFFFFF"   # 컬럼 헤더 텍스트
C_WHITE      = "FFFFFF"
C_GRAY_LV    = "999999"   # Lv.2~5 미편집 회색 텍스트

# Q/W/E/R 스킬별 행 배경 (Lv.1 / Lv.2~5)
SKILL_COLOR = {
    "Q": ("D5F5E3", "EAFAF1"),
    "W": ("D6EAF8", "EBF5FB"),
    "E": ("F9EBEA", "FDEDEC"),
    "R": ("FEF9E7", "FEFDE7"),
}

def side(style='thin', color="CCCCCC"):
    return Side(style=style, color=color)

def border():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, size=10, color="000000", name="맑은 고딕"):
    return Font(name=name, bold=bold, size=size, color=color)

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def wc(ws, row, col, val, fnt=None, fl=None, al=None, brd=True):
    c = ws.cell(row=row, column=col, value=val)
    if fnt: c.font = fnt
    if fl:  c.fill = fl
    if al:  c.alignment = al
    if brd: c.border = border()
    return c


def merge(ws, r1, c1, r2, c2, val, fnt=None, fl=None, al=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=val)
    if fnt: c.font = fnt
    if fl:  c.fill = fl
    if al:  c.alignment = al
    return c


# 열 인덱스 (B=2, C=3, ...)
B, C, D, E, F, G, H, I, J, K, L = 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
LAST_COL = L  # B~L = 11열

COL_WIDTHS = {
    1: 2,    # A (여백)
    B: 6,    # 커맨드
    C: 5,    # 레벨
    D: 9,    # 쿨타임
    E: 9,    # 마나소모
    F: 14,   # 공격유형
    G: 12,   # 피해량
    H: 36,   # 버프/디버프
    I: 9,    # DPS
    J: 9,    # 분당횟수
    K: 12,   # 1분누적피해
    L: 12,   # 1분마나소모
}

HEADERS = ["커맨드","레벨","쿨타임(s)","마나소모","공격유형","피해량(배율)","버프/디버프","DPS","분당횟수","1분누적피해","1분마나소모"]


# ── 스킬 데이터 ───────────────────────────────────────────────────
# (cmd, 스킬명, cd_s, mp, 유형, dmg_ratio, 버프메모, [lv2~5 cd/mp/dmg 오버라이드 or None])
# Lv.2~5는 "편집 필요" placeholder

T1_SKILLS = [
    ("Q", "저격",          5,  30, "방향 투사체",   1.65, "느려짐 → 기절 연쇄 / 기절 대상 추가 25%"),
    ("W", "부채 사격",    12,  30, "방향 범위",     2.25, "느려짐 + 밀어냄 / 부채꼴 범위"),
    ("E", "불화살 소나기",15,  30, "위치지정 범위", 3.0,  "60%×5발 합산 / Pos 타입"),
    ("R", "폭탄 화살",    50, 100, "방향 공격",    5.0,   "기절 / 500% (자체/폭발 중 1개 적용)"),
]

T2_SKILLS = [
    ("Q", "섬광 화살",     4,  25, "방향 공격",    1.5,  "아군: 강화평타 200% + 공속·방어 버프 / 적: 150% 피해"),
    ("W", "천상의 그물",  10,  30, "방향 범위",    2.0,  "느려짐 / 광역 그물"),
    ("E", "재정비 도약",   0,  15, "이동(서브)",   0.0,  "이동 + 이속 버프 / 피해 없음 / 쿨다운 없음"),
    ("R", "금빛 성광",    30,  80, "방향 공격",    4.0,  "캐스팅 1s 후 발사 / 피격 대상 보호막"),
]


def write_tier_section(ws, start_row, tier_label, tier_note, skills, is_lv1_only=True):
    """하나의 티어 섹션 기록. 반환값: 다음 사용할 행 번호"""
    r = start_row

    # 섹션 헤더
    merge(ws, r, B, r, LAST_COL, tier_label,
          fnt=font(bold=True, size=12, color=C_SEC_FG),
          fl=fill(C_SEC_BG),
          al=align())
    ws.row_dimensions[r].height = 22
    r += 1

    # 노트 행
    merge(ws, r, B, r, LAST_COL, tier_note,
          fnt=font(size=10, color=C_DESC_FG),
          fl=fill(C_TITLE_BG),
          al=left())
    ws.row_dimensions[r].height = 16
    r += 1

    # 공백 5행 (단검과 동일 구조)
    for _ in range(5):
        for col in range(B, LAST_COL+1):
            ws.cell(row=r, column=col).fill = fill(C_TITLE_BG)
        r += 1

    # 컬럼 헤더
    for ci, h in enumerate(HEADERS):
        col = B + ci
        wc(ws, r, col, h,
           fnt=font(bold=True, size=10, color=C_COL_FG),
           fl=fill(C_COL_BG),
           al=align())
    ws.row_dimensions[r].height = 18
    r += 1

    # 스킬 데이터
    for cmd, name, cd, mp, stype, dmg, buff in skills:
        base_color, sub_color = SKILL_COLOR[cmd]

        for lv in range(1, 6):
            is_lv1 = (lv == 1)
            bg = base_color if is_lv1 else sub_color
            txt_color = "000000" if is_lv1 else C_GRAY_LV
            row_fnt = font(bold=is_lv1, size=10, color=txt_color)
            fl_c = fill(bg)

            # 값 결정
            if is_lv1:
                lv_cd   = cd
                lv_mp   = mp
                lv_dmg  = dmg
                lv_type = stype
                lv_buff = buff
                lv_name = name
            else:
                lv_cd   = cd
                lv_mp   = mp
                lv_dmg  = None   # 편집 필요
                lv_type = ""
                lv_buff = "Lv.2~5 편집 필요" if is_lv1_only else ""
                lv_name = ""

            # 셀 쓰기
            cmd_val = f"{cmd}  {lv_name}" if is_lv1 else ""
            wc(ws, r, B, cmd_val, fnt=font(bold=True, size=10, color=txt_color), fl=fl_c, al=align())
            wc(ws, r, C, f"Lv.{lv}",  fnt=row_fnt, fl=fl_c, al=align())
            wc(ws, r, D, lv_cd if lv_cd > 0 else "서브", fnt=row_fnt, fl=fl_c, al=align())
            wc(ws, r, E, lv_mp,        fnt=row_fnt, fl=fl_c, al=align())
            wc(ws, r, F, lv_type,      fnt=row_fnt, fl=fl_c, al=align())
            wc(ws, r, G, lv_dmg,       fnt=row_fnt, fl=fl_c, al=align())
            wc(ws, r, H, lv_buff,      fnt=row_fnt, fl=fl_c, al=left())

            # DPS 수식 (Lv.1이고 CD>0일 때만)
            if is_lv1 and lv_cd > 0 and lv_dmg and lv_dmg > 0:
                i_ref = f"G{r}"
                d_ref = f"D{r}"
                wc(ws, r, I, f"={i_ref}/{d_ref}",         fnt=row_fnt, fl=fl_c, al=align())
                wc(ws, r, J, f"=60/{d_ref}",              fnt=row_fnt, fl=fl_c, al=align())
                wc(ws, r, K, f"=J{r}*{i_ref}",           fnt=row_fnt, fl=fl_c, al=align())
                wc(ws, r, L, f"=J{r}*E{r}",              fnt=row_fnt, fl=fl_c, al=align())
            else:
                for col in [I, J, K, L]:
                    wc(ws, r, col, None, fnt=row_fnt, fl=fl_c, al=align())

            ws.row_dimensions[r].height = 16
            r += 1

        # 스킬 간 구분선 (빈 행)
        for col in range(B, LAST_COL+1):
            ws.cell(row=r, column=col).fill = fill("F0F0F0")
        r += 1

    # 섹션 후 빈 행 2개
    for _ in range(2):
        r += 1

    return r


def build_bow_sheet(wb):
    # 기존 활 시트 삭제
    if "활" in wb.sheetnames:
        del wb["활"]

    # 단검 다음 위치에 삽입
    idx = wb.sheetnames.index("단검") + 1 if "단검" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("활", idx)

    # 열 너비 설정
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Row 1: 타이틀 ────────────────────────────────────────────
    merge(ws, 1, B, 1, LAST_COL, "X7  활  스킬 밸런스",
          fnt=font(bold=True, size=16, color=C_TITLE_FG),
          fl=fill(C_TITLE_BG),
          al=align())
    ws.row_dimensions[1].height = 30

    # ── Row 2: 설명 ──────────────────────────────────────────────
    merge(ws, 2, B, 2, LAST_COL,
          "무기 컨셉: 원거리 저격·함정(1T) / 아군 지원·기동(2T) / 미구현(3T)  ·  CC 연쇄와 다중 발사체로 딜링+견제 동시 수행",
          fnt=font(size=11, color=C_DESC_FG),
          fl=fill(C_DESC_BG),
          al=left())
    ws.row_dimensions[2].height = 18

    # ── Row 3: 공백 ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── 1티어 섹션 (Row 4~) ──────────────────────────────────────
    next_row = write_tier_section(
        ws, 4,
        "▶ 1티어  |  Lv.1 데이터만 확보 (Lv.2~5 편집 필요)",
        "GroupId 300401~300551  ·  Q 저격 CC 연쇄(느려짐→기절)  ·  R 폭탄화살 계수 500% (자체/폭발 중 1개 적용)",
        T1_SKILLS,
        is_lv1_only=True,
    )

    # ── 2티어 섹션 ──────────────────────────────────────────────
    next_row = write_tier_section(
        ws, next_row,
        "▶ 2티어  |  ⚖ 일부 균형 추정값 적용  ·  Lv.1 데이터 확보 (Lv.2~5 편집 필요)",
        "GroupId 300601~300751  ·  Q 섬광화살 아군/적 양방향  ·  E 재정비도약 이동스킬(CD없음)",
        T2_SKILLS,
        is_lv1_only=True,
    )

    # ── 3티어 섹션 (미구현) ─────────────────────────────────────
    r = next_row
    merge(ws, r, B, r, LAST_COL,
          "▶ 3티어  |  ⏳ 미구현 (계획 중)",
          fnt=font(bold=True, size=12, color=C_SEC_FG),
          fl=fill("555555"),
          al=align())
    ws.row_dimensions[r].height = 22
    r += 1

    merge(ws, r, B, r, LAST_COL,
          "GroupId 300801~300951  ·  아직 스킬 데이터가 없습니다. 기획 및 CSV 구현 이후 업데이트 예정.",
          fnt=font(size=10, color="AAAAAA"),
          fl=fill("333333"),
          al=left())
    ws.row_dimensions[r].height = 18
    r += 1

    # ── 하단 계수 요약 테이블 ──────────────────────────────────
    r += 2
    merge(ws, r, B, r, LAST_COL, "■ 스킬 계수 요약 (1T/2T Lv.1 기준)",
          fnt=font(bold=True, size=11, color=C_TITLE_FG),
          fl=fill(C_SEC_BG),
          al=align())
    ws.row_dimensions[r].height = 20
    r += 1

    summary_headers = ["티어","커맨드","스킬명","쿨타임(s)","마나","피해계수(×)","유형","특이사항"]
    sh_cols = [B, C, D, E, F, G, H, I]
    for ci, h in enumerate(summary_headers):
        wc(ws, r, sh_cols[ci], h,
           fnt=font(bold=True, size=10, color=C_COL_FG),
           fl=fill(C_COL_BG),
           al=align())
    ws.row_dimensions[r].height = 17
    r += 1

    summary_data = [
        ("1T","Q","저격",         5,  30, 1.65, "방향 투사체", "느려짐→기절 연쇄 / 기절대상 +25%"),
        ("1T","W","부채 사격",   12,  30, 2.25, "방향 범위",   "느려짐+밀어냄 / 광역 CC"),
        ("1T","E","불화살 소나기",15, 30, 3.0,  "위치지정",    "60%×5발"),
        ("1T","R","폭탄 화살",   50, 100, 5.0,  "방향 공격",   "500% / 기절"),
        ("2T","Q","섬광 화살",    4,  25, 1.5,  "방향 공격",   "아군 강화평타200%+버프 / 적 150%"),
        ("2T","W","천상의 그물", 10,  30, 2.0,  "방향 범위",   "느려짐"),
        ("2T","E","재정비 도약",  0,  15, 0.0,  "이동(서브)",  "피해없음 / 이속버프"),
        ("2T","R","금빛 성광",   30,  80, 4.0,  "방향 공격",   "캐스팅 1s / 보호막"),
    ]
    bg_cycle = ["FAFAFA", "F0F0F0"]
    for i, row_data in enumerate(summary_data):
        bg = bg_cycle[i % 2]
        for ci, v in enumerate(row_data):
            wc(ws, r, sh_cols[ci], v,
               fnt=font(size=10),
               fl=fill(bg),
               al=align() if ci != 7 else left())
        ws.row_dimensions[r].height = 16
        r += 1

    print(f"활 시트 완성 (총 {r-1}행)")
    return ws


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print("기존 시트:", wb.sheetnames)
    build_bow_sheet(wb)
    try:
        wb.save(EXCEL_PATH)
        print(f"저장 완료: {EXCEL_PATH}")
    except PermissionError:
        alt = EXCEL_PATH.replace(".xlsx", "_v2.xlsx")
        wb.save(alt)
        print(f"원본 잠김 — 대체 저장: {alt}")


if __name__ == "__main__":
    main()
