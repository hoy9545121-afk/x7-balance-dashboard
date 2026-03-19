"""
Excel 기획서 스킬 데이터 불일치 수정
  1. 한손검 3T R 광휘의 심판: dmg 450→500% (5레벨 전체)
  2. 단검 3T W 침묵의 일격: dmg 0→120% (5레벨 전체) + 공격유형 텍스트 수정
  3. 단검 3T Q 추격자의 발톱: 공격유형에 (치명타) 추가

constants.py 기준값:
  한손검 R: dmg [500,550,600,650,700] → 배율 [5.0,5.5,6.0,6.5,7.0]
  단검 W: dmg [120,132,144,156,168] → 배율 [1.2,1.32,1.44,1.56,1.68]
  단검 Q: crit=True → 공격유형 '돌진 공격 (치명타)'
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)

# ─── 1. 한손검 R 광휘의 심판 (R92~R96): col G=dmg배율 ─────────────────────────
ws_1h = wb['한손검']
# constants.py: dmg=[500,550,600,650,700] → 배율=[5.0,5.5,6.0,6.5,7.0]
R_DMG_1H = [5.0, 5.5, 6.0, 6.5, 7.0]
for i, row in enumerate(range(92, 97)):
    ws_1h.cell(row=row, column=7).value = R_DMG_1H[i]

print("한손검 R (R92-R96) dmg 4.5→5.0 계열 수정 완료")

# ─── 2. 단검 W 침묵의 일격 (R82~R86): col F=공격유형, col G=dmg배율 ──────────
ws_dag = wb['단검']
# constants.py: dmg=[120,132,144,156,168] → 배율=[1.2,1.32,1.44,1.56,1.68]
W_DMG_DAG = [1.2, 1.32, 1.44, 1.56, 1.68]
for i, row in enumerate(range(82, 87)):
    ws_dag.cell(row=row, column=6).value = '버프 + 공격 (공격속도)'  # 공격 포함으로 수정
    ws_dag.cell(row=row, column=7).value = W_DMG_DAG[i]

print("단검 W (R82-R86) dmg 0→1.2 계열 + 공격유형 수정 완료")

# ─── 3. 단검 Q 추격자의 발톱 (R77~R81): col F=공격유형 ─────────────────────────
# crit=True 반영 → '돌진 공격 (치명타)'
for row in range(77, 82):
    ws_dag.cell(row=row, column=6).value = '돌진 공격 (치명타)'

print("단검 Q (R77-R81) 공격유형 '돌진 공격 (치명타)' 수정 완료")

# ─── 저장 ────────────────────────────────────────────────────────────────────
wb.save(FILEPATH)
print("\n✅ 저장 완료:", FILEPATH)

# ─── 검증 ────────────────────────────────────────────────────────────────────
print("\n=== 검증 ===")
wb2 = openpyxl.load_workbook(FILEPATH, read_only=True, data_only=True)

ws_1h2 = wb2['한손검']
print("\n[ 한손검 3T R 광휘의 심판 R92-R96 ]")
for r in range(92, 97):
    print(f"  R{r}: cd={ws_1h2.cell(r,4).value}, mp={ws_1h2.cell(r,5).value}, dmg={ws_1h2.cell(r,7).value}")

ws_d2 = wb2['단검']
print("\n[ 단검 3T Q 추격자의 발톱 R77-R81 ]")
for r in range(77, 82):
    print(f"  R{r}: type={ws_d2.cell(r,6).value}, dmg={ws_d2.cell(r,7).value}")

print("\n[ 단검 3T W 침묵의 일격 R82-R86 ]")
for r in range(82, 87):
    print(f"  R{r}: type={ws_d2.cell(r,6).value}, dmg={ws_d2.cell(r,7).value}, buff={ws_d2.cell(r,8).value}")

wb2.close()
