"""
장비_제작_및_드랍.xlsx — 📋 장비 아이템 리스트 전수 수정
방어구_스킬.xlsx 아이템 리스트 기준으로 패시브/CID/이름 교정

오류 유형:
  A. 재질 컬럼 올바름 + CID/패시브 틀림 → 재질 유지, 재질에 맞는 CID/패시브로 교체
  B. 위와 같음이지만 보스드랍 신발에 해당 (A와 동일 처리)

수정 행 목록 (총 19행):
  2T 전리품 가죽: row 20, 21, 22, 23
  2T 전리품 천:   row 25, 26, 27
  2T 던전코어 아머: row 29, 31, 32, 34
  3T 전리품 가죽 장갑: row 46, 47
  3T 보스드랍 판금 신발: row 67, 68
  3T 보스드랍 가죽 신발: row 71, 73
  3T 보스드랍 천  신발: row 77, 78
"""

import os
from openpyxl import load_workbook
from copy import copy


def copy_fmt(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.border      = copy(src.border)
        dst.alignment   = copy(src.alignment)
        dst.number_format = src.number_format


# ──────────────────────────────────────────────────────────
# 수정 데이터: {row: (p1, p2, p3, cid, name)}
#   p3=None → col7 변경 없음 (2T 아이템)
#   p3=문자열 → 3T 세 번째 패시브
# ──────────────────────────────────────────────────────────
FIXES = {
    # ── 2T 전리품 가죽 ─────────────────────────────────────
    20: ("생명력 자연 회복", "치명타 확률",  None, 12020002, "2T 가죽 투구(전리품 제작)"),
    21: ("공격력",          "최대 생명력",   None, 12020011, "2T 가죽 투구(전리품 제작)"),
    22: ("마력 자연 회복",  "최대 생명력",   None, 12020005, "2T 가죽 장갑(전리품 제작)"),
    23: ("공격 속도",       "스킬 가속",     None, 12020008, "2T 가죽 신발(전리품 제작)"),

    # ── 2T 전리품 천 ──────────────────────────────────────
    25: ("치명타 확률",     "공격력",        None, 12020003, "2T 천 투구(전리품 제작)"),
    26: ("마력 자연 회복",  "마력 자연 회복", None, 12020006, "2T 천 장갑(전리품 제작)"),
    27: ("공격력",          "마력 자연 회복", None, 12020009, "2T 천 신발(전리품 제작)"),

    # ── 2T 던전 코어 아머 ─────────────────────────────────
    29: ("최대 생명력",     "치명타 피해",   None, 12020016, "2T 판금 갑옷(던전 코어)"),
    31: ("받는 피해 감소",  "최대 생명력",   None, 12020014, "2T 가죽 갑옷(던전 코어)"),
    32: ("최대 마력",       "치명타 확률",   None, 12020017, "2T 가죽 갑옷(던전 코어)"),
    34: ("최대 생명력",     "스킬 가속",     None, 12020015, "2T 천 갑옷(던전 코어)"),

    # ── 3T 전리품 가죽 장갑 ───────────────────────────────
    46: ("최대 생명력",    "받는 피해 감소", "받는 치유량",   12030008, "3T 가죽 장갑(전리품 제작)"),
    47: ("공격력",         "마력 자연 회복", "최대 생명력",   12030011, "3T 가죽 장갑(전리품 제작)"),

    # ── 3T 보스드랍 판금 신발 ────────────────────────────
    67: ("공격력",         "최대 생명력",   "받는 피해 감소", 12030034, "3T 판금 신발(드랍)"),
    68: ("스킬 가속",      "공격력",        "공격 속도",      12030037, "3T 판금 신발(드랍)"),

    # ── 3T 보스드랍 가죽 신발 ────────────────────────────
    71: ("스킬 가속",      "피해 증가",     "공격 속도",      12030032, "3T 가죽 신발(드랍)"),
    73: ("마력 자연 회복", "공격력",        "PVE 피해 증가",  12030038, "3T 가죽 신발(드랍)"),

    # ── 3T 보스드랍 천 신발 ──────────────────────────────
    77: ("방어력",         "스킬 가속",     "PVE 피해 증가",  12030033, "3T 천 신발(드랍)"),
    78: ("공격력",         "치명타 피해",   "공격 속도",      12030036, "3T 천 신발(드랍)"),
}

# ──────────────────────────────────────────────────────────
wb = load_workbook("기획서/X7_장비_제작_및_드랍.xlsx")
ws = wb["📋 장비 아이템 리스트"]

print("=== 수정 적용 ===")
for ri, (p1, p2, p3, cid, name) in sorted(FIXES.items()):
    ref = ws.cell(row=ri, column=1)

    old_cid  = ws.cell(row=ri, column=9).value
    old_name = ws.cell(row=ri, column=10).value
    old_p1   = ws.cell(row=ri, column=5).value
    old_p2   = ws.cell(row=ri, column=6).value

    # 패시브 쓰기 (col5, col6)
    c5 = ws.cell(row=ri, column=5, value=p1); copy_fmt(ref, c5)
    c6 = ws.cell(row=ri, column=6, value=p2); copy_fmt(ref, c6)

    # p3 있으면 col7도 갱신
    if p3 is not None:
        c7 = ws.cell(row=ri, column=7, value=p3); copy_fmt(ref, c7)

    # CID / 이름
    c9  = ws.cell(row=ri, column=9,  value=cid);  copy_fmt(ref, c9)
    c10 = ws.cell(row=ri, column=10, value=name); copy_fmt(ref, c10)

    tier = ws.cell(row=ri, column=1).value
    mat  = ws.cell(row=ri, column=3).value
    part = ws.cell(row=ri, column=4).value

    print(f"  row{ri:3d} {tier} {str(mat):3s} {str(part):4s}  "
          f"CID: {str(old_cid):>12} → {str(cid):>12}  "
          f"P: ({old_p1}/{old_p2}) → ({p1}/{p2}{'/' + p3 if p3 else ''})")

print(f"\n총 {len(FIXES)}행 수정 완료")

# ──────────────────────────────────────────────────────────
# 저장
# ──────────────────────────────────────────────────────────
TMP = "기획서/X7_장비_제작_및_드랍_tmp.xlsx"
wb.save(TMP)
try:
    os.replace(TMP, "기획서/X7_장비_제작_및_드랍.xlsx")
    print("\n✅ 저장: X7_장비_제작_및_드랍.xlsx")
    LOAD = "기획서/X7_장비_제작_및_드랍.xlsx"
except PermissionError:
    print(f"\n⚠️  Excel 열려있음 → 임시파일: {TMP}")
    LOAD = TMP

# ──────────────────────────────────────────────────────────
# 최종 검증
# ──────────────────────────────────────────────────────────
wb2 = load_workbook(LOAD)
ws2 = wb2["📋 장비 아이템 리스트"]

total = ok = dup = no_cid = 0
cid_seen = {}
for row in ws2.iter_rows(min_row=4, values_only=True):
    if row[0] not in ("1T", "2T", "3T"):
        continue
    total += 1
    cid  = row[8]
    name = row[9]
    if not cid or "⚠️" in str(name):
        no_cid += 1
        print(f"  ❌ 미매칭: row data={row[:7]}")
    elif cid in cid_seen:
        dup += 1
        print(f"  ⚠️ 중복 CID {cid}: {cid_seen[cid][1]} vs {name}")
    else:
        cid_seen[cid] = (row, name)
        ok += 1

print(f"\n=== 최종 검증 ===")
print(f"  총 아이템: {total}  /  정상: {ok}  /  미매칭: {no_cid}  /  중복CID: {dup}")
if ok == 75:
    print("  ✅ 75개 전부 매칭 완료!")
