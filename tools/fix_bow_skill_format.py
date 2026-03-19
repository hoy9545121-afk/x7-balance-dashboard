"""
활 스킬 양식 표준화 + 전 시트 "N티어" → "N번 스킬셋" 치환
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)

# ═══════════════════════════════════════════════════════
# STEP 1: "N티어" → "N번 스킬셋" 전체 치환
# ═══════════════════════════════════════════════════════
tier_map = [('1티어', '1번 스킬셋'), ('2티어', '2번 스킬셋'), ('3티어', '3번 스킬셋')]

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                v = cell.value
                for old, new in tier_map:
                    v = v.replace(old, new)
                cell.value = v

# 활 시트 요약표(rows 128~141)의 "1T"/"2T"/"3T" 약어도 교체
ws_bow = wb['활']
for r in range(128, 142):
    cell = ws_bow.cell(row=r, column=2)
    if isinstance(cell.value, str):
        v = cell.value
        # 단어 경계 확인 (셀 내 "1T" 단독 or 앞에 공백)
        v = v.replace('1T', '1번').replace('2T', '2번').replace('3T', '3번')
        cell.value = v

# ═══════════════════════════════════════════════════════
# STEP 2: 활 시트 1번/2번 스킬셋 구조 표준화
# ═══════════════════════════════════════════════════════
# 컬럼 인덱스 (1-based)
C_CMD, C_LV, C_CD, C_MP, C_TYPE, C_DMG, C_BUFF = 2, 3, 4, 5, 6, 7, 8
C_DPS, C_HPM, C_CUM, C_MPM = 9, 10, 11, 12

COL_HEADER = ['커맨드', '레벨', '쿨타임(s)', '마나소모', '공격유형',
              '피해량(배율)', '버프/디버프', 'DPS', '분당횟수', '1분누적피해', '1분마나소모']
OV_HEADER  = ['커맨드', '스킬명', '유형', '설명']
NOTE_ROW   = ('※ 피해량 배율 1.0 = ATK×100%  |  노란 셀 = Lv.1 확인 데이터, '
              '연한 셀 = 편집 필요  |  D(쿨타임)·E(마나)·G(피해량) 직접 수정 → 오른쪽 분석 자동 갱신')

# 스킬 데이터: (cmd, name, type, desc, cd, mp, dmg, buff)
skills_1 = [
    ('Q', '저격', '방향 투사체',
     '저격으로 단일 대상에게 피해를 입힙니다. 느려짐 → 기절 연쇄 / 기절 대상 추가 25%',
     5, 30, 1.65, '느려짐 → 기절 연쇄 / 기절 대상 추가 25%'),
    ('W', '부채 사격', '방향 범위',
     '부채꼴 범위로 화살을 쏩니다. 느려짐 + 밀어냄',
     12, 30, 2.25, '느려짐 + 밀어냄 / 부채꼴 범위'),
    ('E', '불화살 소나기', '위치지정 범위',
     '지정 위치에 불화살을 5발 퍼붓습니다. 60%×5발',
     15, 30, 3.0, '60%×5발 합산 / Pos 타입'),
    ('R', '폭탄 화살', '방향 공격',
     '폭탄 화살을 발사하여 강력한 피해를 입힙니다. 500% / 기절',
     50, 100, 5.0, '기절 / 500% (자체/폭발 중 1개 적용)'),
]

skills_2 = [
    ('Q', '섬광 화살', '방향 공격',
     '아군에게는 강화평타 200%+버프, 적에게는 150% 피해를 입힙니다.',
     4, 25, 1.5, '아군: 강화평타 200% + 공속·방어 버프 / 적: 150% 피해'),
    ('W', '천상의 그물', '방향 범위',
     '광역 그물을 발사하여 적을 느리게 합니다.',
     10, 30, 2.0, '느려짐 / 광역 그물'),
    ('E', '재정비 도약', '이동 (서브)',
     '지정 위치로 도약합니다. 피해 없음 / 이동속도 버프',
     '서브', 15, 0, '이동 + 이속 버프 / 피해 없음 / 쿨다운 없음'),
    ('R', '금빛 성광', '방향 공격',
     '캐스팅 1초 후 강력한 광선을 발사합니다. 피격 대상 보호막 부여',
     30, 80, 4.0, '캐스팅 1s 후 발사 / 피격 대상 보호막'),
]


def write_section(ws, start_row, section_num, status_text, groupid_note, skills):
    """
    표준 34행 구조:
      row+0 : section header
      row+1 : GroupId note
      row+2 : blank
      row+3 : overview table header
      row+4~7 : Q/W/E/R overview (4행)
      row+8 : blank
      row+9 : ⚔️ 스킬 강화 정보 subheader
      row+10 : column header
      row+11~30 : skill data (5행 × 4스킬 = 20행)
      row+31 : blank
      row+32 : notes
      row+33 : blank
    합계 34행, returns start_row + 34
    """
    r = start_row

    # row+0: section header
    ws.cell(row=r, column=C_CMD).value = f'▶ {section_num}번 스킬셋  |  {status_text}'
    r += 1
    # row+1: GroupId note
    ws.cell(row=r, column=C_CMD).value = groupid_note
    r += 1
    # row+2: blank
    r += 1
    # row+3: overview table header
    for i, h in enumerate(OV_HEADER):
        ws.cell(row=r, column=C_CMD + i).value = h
    r += 1
    # row+4~7: skill overview
    for cmd, name, stype, desc, *_ in skills:
        ws.cell(row=r, column=C_CMD).value = cmd
        ws.cell(row=r, column=C_CMD + 1).value = name
        ws.cell(row=r, column=C_CMD + 2).value = stype
        ws.cell(row=r, column=C_CMD + 3).value = desc
        r += 1
    # row+8: blank
    r += 1
    # row+9: ⚔️ subheader
    ws.cell(row=r, column=C_CMD).value = f'⚔️ 스킬 강화 정보 — {section_num}번 스킬셋'
    r += 1
    # row+10: column header
    for i, h in enumerate(COL_HEADER):
        ws.cell(row=r, column=C_CMD + i).value = h
    r += 1
    # row+11~30: skill data (5행 × 4스킬)
    for cmd, name, stype, desc, cd, mp, dmg, buff in skills:
        for lv in range(1, 6):
            ws.cell(row=r, column=C_CMD).value  = cmd if lv == 1 else None
            ws.cell(row=r, column=C_LV).value   = lv
            ws.cell(row=r, column=C_CD).value   = cd
            ws.cell(row=r, column=C_MP).value   = mp
            ws.cell(row=r, column=C_TYPE).value = stype
            ws.cell(row=r, column=C_DMG).value  = dmg
            ws.cell(row=r, column=C_BUFF).value = buff if lv == 1 else 'Lv.2~5 편집 필요'
            # DPS 공식: CD가 숫자일 때만
            if isinstance(cd, (int, float)) and cd > 0:
                ws.cell(row=r, column=C_DPS).value = f'=G{r}/(D{r}+1)'
                ws.cell(row=r, column=C_HPM).value = f'=60/(D{r}+1)'
                ws.cell(row=r, column=C_CUM).value = f'=J{r}*G{r}'
                ws.cell(row=r, column=C_MPM).value = f'=J{r}*E{r}'
            r += 1
    # row+31: blank
    r += 1
    # row+32: notes
    ws.cell(row=r, column=C_CMD).value = NOTE_ROW
    r += 1
    # row+33: blank
    r += 1

    return r  # = start_row + 34


# 병합 셀 해제 (rows 4~71 내, unmerge_cells 사용)
merges_to_remove = [
    str(m) for m in list(ws_bow.merged_cells.ranges)
    if m.min_row >= 4 and m.max_row <= 71
]
for m in merges_to_remove:
    ws_bow.unmerge_cells(m)

# 기존 rows 4~71 초기화
for row in range(4, 72):
    for col in range(1, 13):
        ws_bow.cell(row=row, column=col).value = None

# 1번 스킬셋 (rows 4~37)
write_section(
    ws_bow, 4, 1,
    'Lv.1 데이터만 확보 (Lv.2~5 편집 필요)',
    'GroupId 300401~300551  ·  Q 저격 CC 연쇄(느려짐→기절)  ·  R 폭탄화살 계수 500% (자체/폭발 중 1개 적용)',
    skills_1,
)

# 2번 스킬셋 (rows 38~71)
write_section(
    ws_bow, 38, 2,
    '⚖ 일부 균형 추정값 적용  ·  Lv.1 데이터 확보 (Lv.2~5 편집 필요)',
    'GroupId 300601~300751  ·  Q 섬광화살 아군/적 양방향  ·  E 재정비도약 이동스킬(CD없음)',
    skills_2,
)

# ═══════════════════════════════════════════════════════
# STEP 3: 스킬 리스트 시트 — 활 행 컬럼 위치 수정 (A-F → B-G)
# ═══════════════════════════════════════════════════════
ws_list = wb['스킬 리스트']

STATUS_MAP = {
    'Lv.1만':  '🟡 Lv.1만 확보',
    '미구현':  '미구현',
}

# 활 행 데이터 읽기 (rows 51~62, cols A~F = 1~6)
bow_data = {}
for r in range(51, 63):
    row_vals = {}
    for c in range(1, 7):
        v = ws_list.cell(row=r, column=c).value
        if v is not None:
            row_vals[c] = v
    bow_data[r] = row_vals

# 기존 활 행 초기화
for r in range(51, 63):
    for c in range(1, 8):
        ws_list.cell(row=r, column=c).value = None

# 표준 컬럼(B-G)에 재배치
for r, data in bow_data.items():
    if 1 in data:  # 장비 구분 A→B
        ws_list.cell(row=r, column=2).value = data[1]
    if 2 in data:  # 커맨드 B→C
        ws_list.cell(row=r, column=3).value = data[2]
    if 3 in data:  # GroupId C→D (number→text)
        ws_list.cell(row=r, column=4).value = str(int(data[3]))
    if 4 in data:  # 스킬명 D→E
        ws_list.cell(row=r, column=5).value = data[4]
    if 5 in data:  # 데이터 상태 E→F (full format)
        status = str(data[5])
        ws_list.cell(row=r, column=6).value = STATUS_MAP.get(status, status)
    if 6 in data:  # 비고 F→G (공백 정규화)
        ws_list.cell(row=r, column=7).value = str(data[6]).replace(' / ', '/')

# ═══════════════════════════════════════════════════════
# 저장
# ═══════════════════════════════════════════════════════
wb.save(FILEPATH)
print("✅ 완료")
print("  - 전 시트 '1/2/3티어' → '1/2/3번 스킬셋' 치환")
print("  - 활 1번/2번 스킬셋: 개요 테이블 + ⚔️ 서브헤더 + 표준 포맷 재구성")
print("  - 스킬 리스트 활 행: A-F → B-G 컬럼 이동 + GroupId text 변환")
