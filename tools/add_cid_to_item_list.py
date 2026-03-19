"""
X7_장비_제작_및_드랍.xlsx — 📋 장비 아이템 리스트 시트에 Cid / 장비 이름 컬럼 추가

매칭 전략:
  방어구_스킬 아이템 리스트의 타입코드(PlateHelm 등)가 잘못 배정되어 있으므로
  타입코드 대신 이름에 포함된 획득방법 키워드를 파싱하여 사용.

  매칭 키 우선순위:
  1. (tier, acquisition, passives_sorted)  — tier + 획득방법 + 패시브 3키
  2. (tier, passives_sorted)              — tier + 패시브만 (최후 수단)
  3. 미매칭                               — "⚠️ CID 미등록"
"""

import re, os
import openpyxl
from openpyxl import load_workbook
from copy import copy

# ── 획득방법 정규화 ────────────────────────────────────────────────
# 방어구_스킬 이름에서 추출되는 패턴 → 장비 리스트 컬럼값으로 통일
ACQUISITION_NAME_MAP = {
    "기본템":        "기본 제작",
    "전리품 제작":   "전리품 제작",
    "전리품제작":    "전리품 제작",
    "던전 코어":     "던전 코어 제작",
    "던전코어":      "던전 코어 제작",
    "드랍":          "보스 드랍",
}

MATERIAL_MAP = {"Plate": "판금", "Leather": "가죽", "Cloth": "천"}
PART_MAP     = {"Helm": "헬멧", "Armor": "아머", "Gloves": "장갑", "Shoes": "신발"}

def parse_type(type_code):
    """'PlateHelm' → ('판금', '헬멧')"""
    for m_en, m_ko in MATERIAL_MAP.items():
        if type_code.startswith(m_en):
            part_ko = PART_MAP.get(type_code[len(m_en):], "")
            return m_ko, part_ko
    return None, None

def parse_acquisition_from_name(name):
    """'2T 판금 신발(전리품 제작)' → '전리품 제작'"""
    m = re.search(r'\(([^)]+)\)', name)
    if m:
        raw = m.group(1)
        for key, val in ACQUISITION_NAME_MAP.items():
            if key in raw:
                return val
    return None

def extract_korean(s):
    return re.findall(r'\(([^)]+)\)', str(s) if s else "")

def p_key(*passives):
    parts = [p.strip() for p in passives if p and p not in ("-", None, "")]
    return tuple(sorted(parts))

def copy_fmt(src, dst):
    if src.has_style:
        dst.font      = copy(src.font)
        dst.fill      = copy(src.fill)
        dst.border    = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

# ════════════════════════════════════════════════════════════════
# STEP 1: 방어구_스킬 아이템 리스트 → CID 매핑 테이블 구성
#   키 A: (tier, acquisition, p_key)   ← 메인
#   키 B: (tier, p_key)                ← 보조
# ════════════════════════════════════════════════════════════════
wb_skill = load_workbook("기획서/X7_방어구_스킬.xlsx")
ws_skill = wb_skill["아이템 리스트"]

map_5 = {}   # (tier, acq, mat, part, pk) → (cid, name)  ← 완전 5키 (최우선)
map_4 = {}   # (tier, mat, part, pk) → (cid, name)       ← 타입코드+패시브
map_a = {}   # (tier, acq, pk) → (cid, name)             ← 획득방법+패시브
map_b = {}   # (tier, pk)      → [(cid, name, acq)]       ← 패시브만

for row in ws_skill.iter_rows(min_row=2, values_only=True):
    cid = row[0]
    if not isinstance(cid, int):
        continue
    raw_name = str(row[1])
    p1_str   = str(row[3]) if row[3] else ""
    p2_str   = str(row[4]) if row[4] and row[4] != "-" else ""

    tier_m = re.match(r"(\dT)", raw_name)
    if not tier_m:
        continue
    tier = tier_m.group(1)

    acq = parse_acquisition_from_name(raw_name)
    all_ko = extract_korean(p1_str) + extract_korean(p2_str)
    pk = p_key(*all_ko)

    # 타입코드에서 재질/파츠 추출
    type_code = str(row[2]) if row[2] else ""
    mat_skill, part_skill = parse_type(type_code)

    # 5키: 완전 매칭
    key_5 = (tier, acq, mat_skill, part_skill, pk)
    if key_5 not in map_5:
        map_5[key_5] = (cid, raw_name)

    # 4키: 타입코드 기반
    key_4 = (tier, mat_skill, part_skill, pk)
    if key_4 not in map_4:
        map_4[key_4] = (cid, raw_name)

    # 3키: 획득방법 기반
    key_a = (tier, acq, pk)
    if key_a not in map_a:
        map_a[key_a] = (cid, raw_name)

    map_b.setdefault((tier, pk), []).append((cid, raw_name, acq))

print(f"[CID 맵] 5키: {len(map_5)}  4키: {len(map_4)}  3키: {len(map_a)}  2키: {len(map_b)}")

# ════════════════════════════════════════════════════════════════
# STEP 2: 장비 아이템 리스트 시트에 Cid / 장비 이름 추가
# ════════════════════════════════════════════════════════════════
wb = load_workbook("기획서/X7_장비_제작_및_드랍.xlsx")
ws = wb["📋 장비 아이템 리스트"]

# 헤더 row3
for col, val in [(9, "Cid"), (10, "장비 이름")]:
    cell = ws.cell(row=3, column=col, value=val)
    copy_fmt(ws.cell(row=3, column=1), cell)

matched = unmatched = 0
unmatched_list = []

for ri in range(4, ws.max_row + 1):
    tier = ws.cell(row=ri, column=1).value
    if not tier or not isinstance(tier, str) or not tier.endswith("T"):
        continue

    acq  = ws.cell(row=ri, column=2).value   # 획득 방식
    mat  = ws.cell(row=ri, column=3).value   # 재질
    part = ws.cell(row=ri, column=4).value   # 파츠
    p1   = ws.cell(row=ri, column=5).value
    p2   = ws.cell(row=ri, column=6).value
    p3   = ws.cell(row=ri, column=7).value
    pk   = p_key(p1, p2, p3)

    cid_val = name_val = None

    # ① 5키 완전 매칭 (tier + acq + material + part + passives)  ← 최우선
    key_5 = (tier, acq, mat, part, pk)
    if key_5 in map_5:
        cid_val, name_val = map_5[key_5]

    # ② 4키 매칭 (tier + material + part + passives)  ← 1T 재질/파츠 구분
    elif (tier, mat, part, pk) in map_4:
        cid_val, name_val = map_4[(tier, mat, part, pk)]

    # ③ 3키 매칭 (tier + acquisition + passives)  ← 2T/3T 획득방법 구분
    elif (tier, acq, pk) in map_a:
        cid_val, name_val = map_a[(tier, acq, pk)]

    # ③ 2키 매칭 (tier + passives) — 후보 1개이거나 획득방법 일치 시
    elif (tier, pk) in map_b:
        candidates = map_b[(tier, pk)]
        if len(candidates) == 1:
            cid_val, name_val, _ = candidates[0]
        else:
            matched_acq = [c for c in candidates if c[2] == acq]
            if len(matched_acq) == 1:
                cid_val, name_val, _ = matched_acq[0]
            else:
                cid_val, name_val, _ = candidates[0]
                name_val = f"{name_val} ⚠️ 중복CID"

    if cid_val:
        matched += 1
    else:
        name_val = f"⚠️ CID 미등록 ({p1}/{p2})"
        unmatched += 1
        unmatched_list.append((ri, tier, acq, p1, p2, p3))

    cid_cell  = ws.cell(row=ri, column=9,  value=cid_val)
    name_cell = ws.cell(row=ri, column=10, value=name_val)
    copy_fmt(ws.cell(row=ri, column=1), cid_cell)
    copy_fmt(ws.cell(row=ri, column=1), name_cell)

print(f"\n[매칭] ✅ {matched}개  ❌ {unmatched}개")
if unmatched_list:
    print("  미매칭:")
    for u in unmatched_list:
        print(f"    row{u[0]}: {u[1]} {u[2]} | {u[3]}/{u[4]}/{u[5]}")

# ── 저장 ──────────────────────────────────────────────────────
TMP = "기획서/X7_장비_제작_및_드랍_tmp.xlsx"
wb.save(TMP)
try:
    os.replace(TMP, "기획서/X7_장비_제작_및_드랍.xlsx")
    print("\n✅ 저장: X7_장비_제작_및_드랍.xlsx")
    LOAD_PATH = "기획서/X7_장비_제작_및_드랍.xlsx"
except PermissionError:
    print(f"\n⚠️  Excel이 열려있어 임시파일로 저장: {TMP}")
    print("   Excel을 닫은 후 해당 파일을 원본으로 교체하세요.")
    LOAD_PATH = TMP

# ── 최종 검증 출력 ─────────────────────────────────────────────
wb2 = load_workbook(LOAD_PATH)
ws2 = wb2["📋 장비 아이템 리스트"]

print("\n=== 2T 전리품제작 12개 ===")
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if len(rv) >= 10 and rv[0] == "2T" and rv[1] == "전리품 제작":
        flag = "✅" if (rv[8] and "미등록" not in str(rv[9])) else "❌"
        print(f"  {flag} {rv[2]} {str(rv[3]):4s} | "
              f"{str(rv[4]):20s}/{str(rv[5]):18s} | "
              f"CID={str(rv[8]):>10}  {rv[9]}")

print("\n=== 전체 요약 ===")
dup_cids, no_cid = [], []
cid_seen = {}
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if rv[0] not in ("1T", "2T", "3T"):
        continue
    cid  = rv[8] if len(rv) > 8 else None
    name = rv[9] if len(rv) > 9 else None
    if cid is None:
        no_cid.append(rv)
    elif cid in cid_seen:
        dup_cids.append((cid, rv, cid_seen[cid]))
    else:
        cid_seen[cid] = rv

t1=t2=t3=0
for rv in ws2.iter_rows(min_row=4, values_only=True):
    if rv[0]=="1T": t1+=1
    elif rv[0]=="2T": t2+=1
    elif rv[0]=="3T": t3+=1

print(f"  아이템 수: 1T={t1} / 2T={t2} / 3T={t3}")
print(f"  CID 매칭: {len(cid_seen)}개 / 미등록: {len(no_cid)}개 / 중복CID: {len(dup_cids)}쌍")
if dup_cids:
    print("  ⚠️ 중복 CID 목록:")
    for cid, r1, r2 in dup_cids:
        print(f"     CID={cid}: {r1[0]} {r1[1]} {r1[2]} {r1[3]} vs {r2[0]} {r2[1]} {r2[2]} {r2[3]}")
