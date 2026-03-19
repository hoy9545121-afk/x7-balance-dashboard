"""
X7_무기_스킬.xlsx 3T 스킬 데이터 갱신 스크립트
- constants.py의 최신 SKILLS 데이터를 기준으로 3T 섹션 업데이트
- 수식(J/K/L 컬럼)은 유지하며 D(쿨타임)/E(MP)/G(DMG배율)/H(버프설명)/I(DPS) 업데이트
- 활 시트: 3T 섹션 신규 추가
"""
import sys
sys.path.insert(0, "C:/AI_simulator")

import openpyxl
from openpyxl import load_workbook

EXCEL_PATH = "C:/AI_simulator/기획서/X7_무기_스킬.xlsx"

def sv(ws, row, col, value):
    """셀 값 설정"""
    ws.cell(row=row, column=col).value = value

def dps_formula(row):
    return f"=G{row}/D{row}"

def count_formula(row):
    return f"=60/D{row}"

def total_dmg_formula(row):
    return f"=J{row}*G{row}"

def total_mp_formula(row):
    return f"=J{row}*E{row}"

def write_skill_rows(ws, start_row, cmd, lv_data, skill_type, buff_text=None, is_pure_buff=False):
    """
    스킬 5레벨 행 작성 (lv=1~5)
    lv_data: [(cd, mp, dmg_ratio), ...] 5개
    is_pure_buff: dmg=0인 순수 버프 스킬 (K='0', I='업타임XX%')
    buff_text: H 컬럼에 쓸 버프 설명 (None이면 빈칸)
    """
    for i, (cd, mp, dmg) in enumerate(lv_data):
        row = start_row + i
        if i == 0:
            sv(ws, row, 2, cmd)  # B: CMD
        sv(ws, row, 3, i + 1)    # C: Level
        sv(ws, row, 4, cd)       # D: CD
        sv(ws, row, 5, mp)       # E: MP
        sv(ws, row, 6, skill_type)  # F: type
        sv(ws, row, 7, dmg)      # G: DMG ratio
        if buff_text:
            sv(ws, row, 8, buff_text)  # H: buff
        else:
            sv(ws, row, 8, None)
        if is_pure_buff:
            sv(ws, row, 9, None)       # I: DPS - 순수버프는 직접 입력
            sv(ws, row, 11, 0)         # K: '0'
        else:
            sv(ws, row, 9, dps_formula(row))   # I: =G/D
            sv(ws, row, 11, total_dmg_formula(row))  # K: =J*G
        sv(ws, row, 10, count_formula(row))   # J: =60/D
        sv(ws, row, 12, total_mp_formula(row))  # L: =J*E

def write_summary_tables(ws, data_start_Q, data_start_W, data_start_E, data_start_R, summary_start):
    """3T 요약 테이블 (DPS/딜비중/마나소모)"""
    r = summary_start

    # 📊 지속 딜링 능력
    sv(ws, r, 2, "📊 지속 딜링 능력 (1분 기준) — 3티어")
    r += 1
    for label, col in [("레벨", 2), ("Q", 3), ("W", 4), ("E", 5), ("R", 6), ("1분 합계", 7), ("지속DPS", 8), ("레벨업 증가율", 9)]:
        sv(ws, r, col if col > 2 else 2, label)
    sv(ws, r, 3, "Q"); sv(ws, r, 4, "W"); sv(ws, r, 5, "E"); sv(ws, r, 6, "R")
    sv(ws, r, 7, "1분 합계"); sv(ws, r, 8, "지속DPS"); sv(ws, r, 9, "레벨업 증가율")
    r += 1

    for lv in range(5):
        sv(ws, r, 2, lv + 1)
        sv(ws, r, 3, f"=K{data_start_Q + lv}")
        sv(ws, r, 4, f"=K{data_start_W + lv}")
        sv(ws, r, 5, f"=K{data_start_E + lv}")
        sv(ws, r, 6, f"=K{data_start_R + lv}")
        sv(ws, r, 7, f"=SUM(C{r}:F{r})")
        sv(ws, r, 8, f"=G{r}/60")
        sv(ws, r, 9, "-" if lv == 0 else f"=H{r}/H{r-1}")
        r += 1

    r += 1  # blank
    # 📊 딜 비중표
    sv(ws, r, 2, "📊 딜 비중표 (%) — 3티어")
    r += 1
    sv(ws, r, 2, "레벨"); sv(ws, r, 3, "Q"); sv(ws, r, 4, "W"); sv(ws, r, 5, "E"); sv(ws, r, 6, "R")
    r += 1
    ref_row = summary_start + 2  # first data row in DPS table
    for lv in range(5):
        sv(ws, r, 2, lv + 1)
        sv(ws, r, 3, f"=C{ref_row + lv}/G{ref_row + lv}")
        sv(ws, r, 4, f"=D{ref_row + lv}/G{ref_row + lv}")
        sv(ws, r, 5, f"=E{ref_row + lv}/G{ref_row + lv}")
        sv(ws, r, 6, f"=F{ref_row + lv}/G{ref_row + lv}")
        r += 1

    r += 1  # blank
    # 💧 마나 소모량
    sv(ws, r, 2, "💧 마나 소모량 (1분 기준) — 3티어")
    r += 1
    sv(ws, r, 2, "레벨"); sv(ws, r, 3, "Q"); sv(ws, r, 4, "W"); sv(ws, r, 5, "E"); sv(ws, r, 6, "R")
    sv(ws, r, 7, "합계"); sv(ws, r, 8, "증가율")
    r += 1
    for lv in range(5):
        sv(ws, r, 2, lv + 1)
        sv(ws, r, 3, f"=L{data_start_Q + lv}")
        sv(ws, r, 4, f"=L{data_start_W + lv}")
        sv(ws, r, 5, f"=L{data_start_E + lv}")
        sv(ws, r, 6, f"=L{data_start_R + lv}")
        sv(ws, r, 7, f"=SUM(C{r}:F{r})")
        sv(ws, r, 8, "-" if lv == 0 else f"=G{r}/G{r-1}")
        r += 1


# ============================================================
# 메인
# ============================================================
wb = load_workbook(EXCEL_PATH)

# ─────────────── 1. 양손검 3T ───────────────
ws = wb["양손검"]

sv(ws, 68, 2, "▶ 3티어  |  ✅ 전 레벨 데이터 확보 (시뮬레이터 구현)")

# 스킬명 업데이트
sv(ws, 70, 3, "가벼운 손놀림"); sv(ws, 70, 4, "자기중심 광역 (평타 강화)")
sv(ws, 71, 3, "대지 가르기");  sv(ws, 71, 4, "방향 공격")
sv(ws, 72, 3, "휘몰이");       sv(ws, 72, 4, "범위 공격 (6-hit 합산)")
sv(ws, 73, 3, "적진으로");     sv(ws, 73, 4, "돌진 + 착지 범위 공격 (max100)")

# Q (77-81): 가벼운 손놀림 — 강화평타 부여
q_data = [(5.0,10,2.0),(4.5,11,2.2),(4.0,13,2.4),(3.5,14,2.6),(3.0,15,2.8)]
write_skill_rows(ws, 77, "Q", q_data, "자기중심 광역 (평타 강화)", buff_text="강화 평타 부여 +200% / 5s")
# I는 수식으로 복원 (Q는 dmg가 있으므로)
for r in range(77, 82):
    sv(ws, r, 9, dps_formula(r))

# W (82-86): 대지 가르기
w_data = [(12,30,2.5),(11,34,2.75),(10,38,3.0),(9,41,3.25),(8,45,3.5)]
write_skill_rows(ws, 82, "W", w_data, "방향 공격")

# E (87-91): 휘몰이
e_data = [(15,15,3.0),(14,16,3.3),(13,17,3.6),(12,18,3.9),(11,20,4.2)]
write_skill_rows(ws, 87, "E", e_data, "범위 공격 (6-hit 합산)")

# R (92-96): 적진으로
r_data = [(40,50,4.0),(38,55,4.4),(36,60,4.8),(34,65,5.2),(32,70,5.6)]
write_skill_rows(ws, 92, "R", r_data, "돌진 + 착지 범위 공격 (max100)")

print("✓ 양손검 3T 업데이트 완료")

# ─────────────── 2. 한손검 3T ───────────────
ws = wb["한손검"]

sv(ws, 68, 2, "▶ 3티어  |  ✅ 전 레벨 데이터 확보 (시뮬레이터 구현)")

# 스킬명 업데이트
sv(ws, 70, 3, "성스러운 일격"); sv(ws, 70, 4, "단일 공격")
sv(ws, 71, 3, "가호의 빛");    sv(ws, 71, 4, "공격 + 방어 버프")
sv(ws, 72, 3, "폭주하는 광휘"); sv(ws, 72, 4, "범위 공격")
sv(ws, 73, 3, "광휘의 심판");  sv(ws, 73, 4, "단일 공격")

# Q (77-81): 성스러운 일격 (치명타 적용)
q_data = [(6.0,15,1.5),(5.5,16,1.65),(5.0,17,1.8),(4.5,18,1.95),(4.0,20,2.1)]
write_skill_rows(ws, 77, "Q", q_data, "단일 공격")

# W (82-86): 가호의 빛 — def_pct 버프 + 피해
# dmg=200/220/240/260/280 → 2.0/2.2/2.4/2.6/2.8
# 버프: 방어력 +30%, 지속 5/5.5/6/6.5/7s
w_dur = [5.0, 5.5, 6.0, 6.5, 7.0]
w_data = [(12,35,2.0),(11,39,2.2),(10,43,2.4),(9,47,2.6),(8,50,2.8)]
for i, (cd, mp, dmg) in enumerate(w_data):
    row = 82 + i
    if i == 0:
        sv(ws, row, 2, "W")
    sv(ws, row, 3, i + 1)
    sv(ws, row, 4, cd)
    sv(ws, row, 5, mp)
    sv(ws, row, 6, "공격 + 방어 버프")
    sv(ws, row, 7, dmg)
    sv(ws, row, 8, f"방어력 +30% / {w_dur[i]}s")
    sv(ws, row, 9, dps_formula(row))      # I: DPS 수식
    sv(ws, row, 10, count_formula(row))   # J
    sv(ws, row, 11, total_dmg_formula(row))  # K: =J*G (dmg>0이므로)
    sv(ws, row, 12, total_mp_formula(row))   # L

# E (87-91): 폭주하는 광휘
e_data = [(20,55,2.5),(19,60,2.75),(18,65,3.0),(17,70,3.25),(16,75,3.5)]
write_skill_rows(ws, 87, "E", e_data, "범위 공격")

# R (92-96): 광휘의 심판
r_data = [(45,180,4.5),(43,198,4.95),(41,215,5.4),(39,233,5.85),(37,250,6.3)]
write_skill_rows(ws, 92, "R", r_data, "단일 공격")

print("✓ 한손검 3T 업데이트 완료")

# ─────────────── 3. 단검 3T ───────────────
ws = wb["단검"]

sv(ws, 68, 2, "▶ 3티어  |  ✅ 전 레벨 데이터 확보 (시뮬레이터 구현)")

# 스킬명 업데이트
sv(ws, 70, 3, "추격자의 발톱"); sv(ws, 70, 4, "돌진 공격")
sv(ws, 71, 3, "침묵의 일격");   sv(ws, 71, 4, "버프 (공격속도)")
sv(ws, 72, 3, "사방의 비수");   sv(ws, 72, 4, "자기중심 광역")
sv(ws, 73, 3, "유혈의 장막");   sv(ws, 73, 4, "방향 범위 공격")

# Q (77-81): 추격자의 발톱 (치명타)
q_data = [(6.0,50,2.0),(5.5,55,2.2),(5.0,60,2.4),(4.5,65,2.6),(4.0,70,2.8)]
write_skill_rows(ws, 77, "Q", q_data, "돌진 공격")

# W (82-86): 침묵의 일격 — 순수 버프 (dmg=0)
# 사이클: 버프(5~6s) + RemovalEffect CD 8s = 총 13~14s
# 업타임: 5/13≈38%, 5.25/13≈40%, 5.5/13.5≈41%, 5.75/13.5≈43%, 6/14≈43%
w_cd  = [13.0, 13.0, 13.5, 13.5, 14.0]
w_mp  = [20, 23, 25, 28, 30]
w_dur = [5.0, 5.25, 5.5, 5.75, 6.0]
w_up  = ["38%", "40%", "41%", "43%", "43%"]
for i in range(5):
    row = 82 + i
    if i == 0:
        sv(ws, row, 2, "W")
    sv(ws, row, 3, i + 1)
    sv(ws, row, 4, w_cd[i])
    sv(ws, row, 5, w_mp[i])
    sv(ws, row, 6, "버프 (공격속도)")
    sv(ws, row, 7, 0)
    sv(ws, row, 8, f"공격속도 +35% / {w_dur[i]}s  (사이클 {w_cd[i]}s)")
    sv(ws, row, 9, f"업타임 {w_up[i]}")   # I: 텍스트
    sv(ws, row, 10, count_formula(row))    # J
    sv(ws, row, 11, 0)                     # K: dmg=0
    sv(ws, row, 12, total_mp_formula(row)) # L

# E (87-91): 사방의 비수
e_data = [(15,40,2.5),(14,44,2.75),(13,48,3.0),(12,51,3.25),(11,55,3.5)]
write_skill_rows(ws, 87, "E", e_data, "자기중심 광역")

# R (92-96): 유혈의 장막
r_data = [(45,150,4.5),(43,165,4.95),(41,180,5.4),(39,195,5.85),(37,210,6.3)]
write_skill_rows(ws, 92, "R", r_data, "방향 범위 공격")

print("✓ 단검 3T 업데이트 완료")

# ─────────────── 4. 지팡이 3T ───────────────
ws = wb["지팡이"]

sv(ws, 116, 2, "▶ 3티어  |  ✅ 전 레벨 데이터 확보 (시뮬레이터 구현)")

# 스킬명/설명 업데이트
sv(ws, 118, 3, "얼음 화살");    sv(ws, 118, 4, "방향 투사체")
sv(ws, 119, 3, "다후타의 손짓"); sv(ws, 119, 4, "DoT 장판 (5틱 합산)")
sv(ws, 120, 3, "심판의 낙뢰");  sv(ws, 120, 4, "위치 지정 (2-hit)")
sv(ws, 121, 3, "방울방울");     sv(ws, 121, 4, "CC + 피해 (부유)")
sv(ws, 121, 5, "부유 CC로 적을 띄우고 피해를 입힙니다. (300%~420% / 부유 8~10s)")

# Q (125-129): 얼음 화살
q_data = [(3.0,30,2.0),(2.5,34,2.2),(2.5,38,2.4),(2.0,41,2.6),(2.0,45,2.8)]
write_skill_rows(ws, 125, "Q", q_data, "방향 투사체")

# W (130-134): 다후타의 손짓 (DoT)
w_data = [(15,50,4.0),(14,55,4.4),(13,60,4.8),(12,65,5.2),(11,70,5.6)]
write_skill_rows(ws, 130, "W", w_data, "DoT 장판 (5틱 합산)")

# E (135-139): 심판의 낙뢰 (2-hit)
e_data = [(40,50,6.0),(38,65,6.6),(36,80,7.2),(34,95,7.8),(32,110,8.4)]
write_skill_rows(ws, 135, "E", e_data, "위치 지정 (2-hit)")

# R (140-144): 방울방울 — 부유 CC + 피해
r_dur = [8.0, 8.5, 9.0, 9.5, 10.0]
r_dmg = [3.0, 3.3, 3.6, 3.9, 4.2]
r_cd  = [20, 19, 18, 17, 16]
r_mp  = [50, 55, 60, 65, 70]
for i in range(5):
    row = 140 + i
    if i == 0:
        sv(ws, row, 2, "R")
    sv(ws, row, 3, i + 1)
    sv(ws, row, 4, r_cd[i])
    sv(ws, row, 5, r_mp[i])
    sv(ws, row, 6, "CC + 피해 (부유)")
    sv(ws, row, 7, r_dmg[i])
    sv(ws, row, 8, f"부유 {r_dur[i]}s")
    sv(ws, row, 9, dps_formula(row))      # I: DPS 수식 (dmg>0)
    sv(ws, row, 10, count_formula(row))   # J
    sv(ws, row, 11, total_dmg_formula(row))  # K
    sv(ws, row, 12, total_mp_formula(row))   # L

print("✓ 지팡이 3T 업데이트 완료")

# ─────────────── 5. 활 3T — 신규 추가 ───────────────
ws = wb["활"]

# 기존 3T placeholder 텍스트 업데이트
sv(ws, 72, 2, "▶ 3티어  |  ✅ 전 레벨 데이터 확보 (시뮬레이터 구현)")
sv(ws, 73, 2, "GroupId 300001~300155  ·  Q 충격사격(치명)  /  W 연속쏘기(범위)  /  E 급습(이동+공속버프)  /  R 발묶음(CC+피해)")

# ── row 74~max_row 전체 클리어 (이전 실행 잔여 데이터 제거) ─────────────
# 먼저 해당 범위의 모든 병합 셀 해제
old_max = ws.max_row
merges_to_remove = [str(m) for m in list(ws.merged_cells.ranges)
                    if m.max_row >= 74]
for merge_str in merges_to_remove:
    ws.unmerge_cells(merge_str)

# row 74~old_max 전체 셀 값 클리어
for r in range(74, old_max + 1):
    for c in range(2, 13):
        cell = ws.cell(row=r, column=c)
        try:
            cell.value = None
        except AttributeError:
            pass

# 여기서는 행 삽입 없이 row 74부터 직접 작성
# (삽입 없이 덮어쓰기 방식으로 중복 방지)

# ── 3T 스킬 목록 (row 74부터 시작) ─────────────
# R=74: blank, R=75: header, R=76~79: skill names
sv(ws, 75, 2, "커맨드"); sv(ws, 75, 3, "스킬명"); sv(ws, 75, 4, "유형"); sv(ws, 75, 5, "설명")
sv(ws, 76, 2, "Q"); sv(ws, 76, 3, "충격 사격"); sv(ws, 76, 4, "방향 투사체 (치명타)")
sv(ws, 76, 5, "충격파 화살을 발사하여 피해를 입힙니다. 치명타 적용.")
sv(ws, 77, 2, "W"); sv(ws, 77, 3, "연속 쏘기"); sv(ws, 77, 4, "방향 범위")
sv(ws, 77, 5, "여러 발의 화살을 연속으로 발사하여 피해를 입힙니다.")
sv(ws, 78, 2, "E"); sv(ws, 78, 3, "급습"); sv(ws, 78, 4, "이동 + 공격속도 버프")
sv(ws, 78, 5, "지정 위치로 순간이동 후 공격 속도를 높입니다. (+50~70% / 7~9s)")
sv(ws, 79, 2, "R"); sv(ws, 79, 3, "발묶음"); sv(ws, 79, 4, "CC + 피해 (속박)")
sv(ws, 79, 5, "함정으로 적의 발을 묶어 행동불능 상태로 만들고 피해를 입힙니다.")

# ── 스킬 강화 정보 테이블 (R81부터) ─────────────
sv(ws, 81, 2, "⚔️ 스킬 강화 정보 — 3티어")
header_row = 82
sv(ws, header_row, 2, "커맨드"); sv(ws, header_row, 3, "레벨")
sv(ws, header_row, 4, "쿨타임(s)"); sv(ws, header_row, 5, "마나소모")
sv(ws, header_row, 6, "공격유형"); sv(ws, header_row, 7, "피해량(배율)")
sv(ws, header_row, 8, "버프/디버프"); sv(ws, header_row, 9, "DPS")
sv(ws, header_row, 10, "분당횟수"); sv(ws, header_row, 11, "1분누적피해")
sv(ws, header_row, 12, "1분마나소모")

# Q (83-87): 충격 사격 (치명타)
q_data = [(5.0,10,2.2),(4.5,11,2.42),(4.0,13,2.64),(3.5,14,2.86),(3.0,15,3.08)]
write_skill_rows(ws, 83, "Q", q_data, "방향 투사체 (치명타)")

# W (88-92): 연속 쏘기
w_data = [(12,35,2.8),(11,39,3.08),(10,43,3.36),(9,46,3.64),(8,50,3.92)]
write_skill_rows(ws, 88, "W", w_data, "방향 범위")

# E (93-97): 급습 — 공격속도 버프 + 피해 (dmg=350/385/420/455/490)
# 버프: AtkSpeedVaryper +50~70%, 7~9s
e_dur  = [7.0, 7.5, 8.0, 8.5, 9.0]
e_spd  = [50,  55,  60,  65,  70]
e_cd   = [20,  19,  18,  17,  16]
e_mp   = [55,  60,  65,  70,  75]
e_dmg  = [3.5, 3.85, 4.2, 4.55, 4.9]
for i in range(5):
    row = 93 + i
    if i == 0:
        sv(ws, row, 2, "E")
    sv(ws, row, 3, i + 1)
    sv(ws, row, 4, e_cd[i])
    sv(ws, row, 5, e_mp[i])
    sv(ws, row, 6, "이동 + 공격속도 버프")
    sv(ws, row, 7, e_dmg[i])
    sv(ws, row, 8, f"공격속도 +{e_spd[i]}% / {e_dur[i]}s")
    sv(ws, row, 9, dps_formula(row))
    sv(ws, row, 10, count_formula(row))
    sv(ws, row, 11, total_dmg_formula(row))
    sv(ws, row, 12, total_mp_formula(row))

# R (98-102): 발묶음 — CC(속박) + 피해
r_data = [(45,180,5.0),(43,198,5.5),(41,215,6.0),(39,233,6.5),(37,250,7.0)]
write_skill_rows(ws, 98, "R", r_data, "CC + 피해 (속박)", buff_text="속박 CC (발묶음)")

# ── 요약 테이블 (R104부터) ─────────────
write_summary_tables(ws, 83, 88, 93, 98, summary_start=104)

# 주석 행 (마나 소모표 Lv5=R126 이후 R128)
sv(ws, 128, 2, "※ 피해량 배율 1.0 = ATK×100%  |  E급습은 공격속도 버프 효과 미포함 (순수 스킬 피해만 집계)")

# ── 1T/2T 스킬 계수 요약 테이블 복원 (R130부터) ─────────────
sv(ws, 130, 2, "■ 스킬 계수 요약 (1T/2T Lv.1 기준)")
sv(ws, 131, 2, "티어"); sv(ws, 131, 3, "커맨드"); sv(ws, 131, 4, "스킬명")
sv(ws, 131, 5, "쿨타임(s)"); sv(ws, 131, 6, "마나"); sv(ws, 131, 7, "피해계수(×)")
sv(ws, 131, 8, "유형"); sv(ws, 131, 9, "특이사항")
summary_1t2t = [
    ("1T","Q","저격",      5, 30, 1.65,"방향 투사체","느려짐→기절 연쇄 / 기절대상 +25%"),
    ("1T","W","부채 사격",  12, 30, 2.25,"방향 범위","느려짐+밀어냄 / 광역 CC"),
    ("1T","E","불화살 소나기",15,30,3.0,"위치지정","60%×5발"),
    ("1T","R","폭탄 화살",  50,100,10.0,"방향 공격","500%×2 / 기절"),
    ("2T","Q","섬광 화살",   4, 25, 1.5,"방향 공격","아군 강화평타200%+버프 / 적 150%"),
    ("2T","W","천상의 그물", 10, 30, 2.0,"방향 범위","느려짐"),
    ("2T","E","재정비 도약",  0, 15, 0.0,"이동(서브)","피해없음 / 이속버프"),
    ("2T","R","금빛 성광",   30, 80, 4.0,"방향 공격","캐스팅 1s / 보호막"),
]
for i, (tier, cmd, name, cd, mp, dmg, typ, note) in enumerate(summary_1t2t):
    r = 132 + i
    sv(ws, r, 2, tier); sv(ws, r, 3, cmd); sv(ws, r, 4, name)
    sv(ws, r, 5, cd);   sv(ws, r, 6, mp);  sv(ws, r, 7, dmg)
    sv(ws, r, 8, typ);  sv(ws, r, 9, note)

print("✓ 활 3T 신규 추가 완료")

# ─────────────── 저장 ───────────────
output_path = EXCEL_PATH
wb.save(output_path)
print(f"\n✅ 저장 완료: {output_path}")
