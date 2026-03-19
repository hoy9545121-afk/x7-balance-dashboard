"""
tools/update_dashboard.py
index.html + 성장 밸런스 데이터를 합쳐 X7_대시보드.html 생성
새 기획서가 생길 때마다 재실행하면 됩니다.
"""

import openpyxl, json, re

XLSX     = r"C:\AI_simulator\기획서\성장_경제_밸런스_수치표.xlsx"
HTML_IN  = r"C:\Users\hoy5343\Documents\카카오톡 받은 파일\index.html"
HTML_OUT = r"C:\AI_simulator\기획서\X7_대시보드.html"
HTML_SYNC = r"C:\Users\hoy5343\Documents\카카오톡 받은 파일\index.html"  # 원본도 동기화

# ── 1. 데이터 읽기 ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)

# 캐릭터 레벨 1~100  [lv, h_cum, min_req, xp_lv, xp_cum, area]
level_data = []
for sn in ['📊 캐릭터 레벨 1~60', '📊 캐릭터 레벨 61~100']:
    ws = wb[sn]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], (int, float)) and row[0]:
            level_data.append([
                int(row[0]),
                round(float(row[1]) if row[1] else 0, 2),
                round(float(row[2]) if row[2] else 0, 1),
                int(row[3]) if row[3] else 0,
                int(row[4]) if row[4] else 0,
                str(row[6]) if row[6] else '',
            ])

# 숙련도 — 티어 구조로 재편
ws_p = wb['🏆 숙련도 경험치']
prof_raw = []
for row in ws_p.iter_rows(min_row=2, values_only=True):
    if any(c for c in row[:5]):
        prof_raw.append(row)
prof_raw = prof_raw[1:]  # 헤더 제거

tiers = []
cur = None
for row in prof_raw:
    tier_s, lv_s, target, hours, xp_lv, tier_total, note = row
    if tier_s:
        cur = {'t': str(tier_s), 'target': str(target) if target else '',
               'hours': str(hours) if hours else '',
               'total': int(tier_total) if tier_total else 0, 'lvs': []}
        tiers.append(cur)
    lv_n = int(lv_s[2:]) if lv_s and lv_s.startswith('Lv') else 0
    cur['lvs'].append({'lv': lv_n, 'xp': int(xp_lv) if xp_lv else 0,
                       'note': str(note) if note else ''})

# 지역 기준표
AREAS = [
    ('이니스 섬',    '1~10',   10,   2,   2,  '1T'),
    ('솔즈리드 반도','11~20',  20,   6,   4,  '2T'),
    ('릴리엇 구릉지','21~30',  30,  12,   6,  '3T'),
    ('가랑돌 평원',  '31~35',  35,  20,   8,  '4T'),
    ('하얀 숲',      '36~40',  40,  30,  10,  '4T'),
    ('마리아 노플',  '41~45',  45,  50,  20,  '5T'),
    ('황금 평원',    '46~50',  50,  90,  40,  '5T'),
    ('지옥 늪지대',  '51~55',  55, 170,  80,  '6T'),
    ('긴 모래톱',    '56~60',  60, 330, 160,  '7T'),
]

TIER_COLORS = {
    '1T': '#50c870', '2T': '#5090e8', '3T': '#e8c050',
    '4T': '#e87830', '5T': '#c850c8', '6T': '#d43030', '7T': '#ff5555',
}

# ── 2. 성장 섹션 HTML 생성 ─────────────────────────────────────────────────

def fmt_xp(n):
    if n >= 10000: return f'{n//10000}만 {n%10000:,}' if n%10000 else f'{n//10000}만'
    return f'{n:,}'

def build_growth_section():
    # JS 데이터
    js_levels = json.dumps(level_data, ensure_ascii=False)
    js_tiers  = json.dumps(tiers, ensure_ascii=False)

    # 지역 테이블 행
    area_rows = ''
    for area, lvs, comp, h_total, h_area, ptier in AREAS:
        c = TIER_COLORS.get(ptier, '#aaa')
        area_rows += f'''<tr>
          <td>{area}</td><td>{lvs}</td>
          <td class="td-num">{comp}레벨</td>
          <td class="td-num">{h_total}h</td>
          <td class="td-num">{h_area}h</td>
          <td><span style="color:{c};font-weight:700">{ptier}</span></td>
        </tr>\n'''

    # 숙련도 티어 카드
    prof_cards = ''
    for td in tiers:
        c = TIER_COLORS.get(td['t'], '#aaa')
        lv_rows = ''
        max_lv = len(td['lvs'])
        for lv_info in td['lvs']:
            is_last = lv_info['lv'] == max_lv
            note_html = f' <span style="color:{c};font-size:0.7rem">← {lv_info["note"]}</span>' if lv_info['note'] else ''
            bold = ' style="font-weight:700"' if is_last else ''
            lv_rows += f'<tr{bold}><td>Lv.{lv_info["lv"]}</td><td class="td-num">{lv_info["xp"]:,}</td><td style="color:var(--text2)">{lv_info["xp"]*lv_info["lv"]//lv_info["lv"]:,}</td></tr>\n'
        char_req = td['target'].replace('캐릭터 ', '캐릭터 <strong>').replace('레벨 완료', '레벨</strong> 완료') if td['target'] else '—'
        prof_cards += f'''
<div class="card" style="border-color:{c}33">
  <div class="card-head" style="background:linear-gradient(90deg,{c}18,transparent)">
    <span style="color:{c};font-weight:900;font-size:1.1rem">{td["t"]}</span>
    <div>
      <div class="card-head-title" style="color:{c}">{td["t"]} 숙련도</div>
      <div class="card-head-sub">입장 조건: {char_req} / 총 {td["total"]:,} XP / 달성 {td["hours"]}</div>
    </div>
  </div>
  <div class="card-body">
    <table>
      <tr><th>레벨</th><th>레벨 필요 XP</th><th>누계 XP</th></tr>
      {lv_rows}
    </table>
  </div>
</div>'''

    return f'''
<!-- ══ GROWTH ══ -->
<section class="section" id="sec-growth">
  <div class="sec-header">
    <div class="sec-header-text">
      <div class="sec-title">📈 성장 밸런스</div>
      <div class="sec-desc">캐릭터 레벨 · 숙련도 경험치 설계</div>
    </div>
  </div>

  <div class="tab-bar">
    <button class="tab-btn active" onclick="switchTab('grw','intent',this)">📌 기획 의도</button>
    <button class="tab-btn" onclick="switchTab('grw','charlevel',this)">📊 캐릭터 레벨</button>
    <button class="tab-btn" onclick="switchTab('grw','proficiency',this)">🏆 숙련도</button>
  </div>

  <!-- INTENT ─────────────────── -->
  <div id="grw-intent">
    <div class="stats-row" style="margin-bottom:20px">
      <div class="stat-box"><div class="stat-num">100</div><div class="stat-label">최대 레벨</div></div>
      <div class="stat-box"><div class="stat-num">7</div><div class="stat-label">숙련도 티어</div></div>
      <div class="stat-box"><div class="stat-num">330h</div><div class="stat-label">60레벨 달성</div></div>
      <div class="stat-box"><div class="stat-num">1,200h</div><div class="stat-label">100레벨 달성</div></div>
      <div class="stat-box"><div class="stat-num">300 XP/분</div><div class="stat-label">숙련도 기준</div></div>
    </div>

    <div class="eco-grid" style="margin-bottom:20px">
      <div class="route-card">
        <div class="route-card-head"><span class="route-icon">📊</span>
          <div><div class="route-name">캐릭터 레벨 설계</div>
          <div class="route-tagline">구간별 등차 성장</div></div>
        </div>
        <div class="route-body"><ul class="route-features">
          <li class="route-feature">1레벨 4,300 XP → 60레벨 1,234,300 XP (구간별 등차 증가)</li>
          <li class="route-feature">60레벨 = 330시간 / 100레벨 = 1,200시간 (지역 완료 기준 앵커)</li>
          <li class="route-feature">레벨업 소요 시간이 급증하는 구간: 40레벨 이후 (하얀 숲~)</li>
        </ul></div>
      </div>
      <div class="route-card">
        <div class="route-card-head"><span class="route-icon">🏆</span>
          <div><div class="route-name">숙련도 시스템</div>
          <div class="route-tagline">티어 제한 경험치 획득</div></div>
        </div>
        <div class="route-body"><ul class="route-features">
          <li class="route-feature">분당 300 XP 기준 / 캐릭터 XP의 약 60% 수준</li>
          <li class="route-feature"><strong style="color:var(--gold)">자신의 숙련도 티어 이상 몬스터에서만 경험치 획득</strong></li>
          <li class="route-feature">지역 클리어 시 해당 티어 자동 달성 — 레벨·숙련도 동기화 설계</li>
          <li class="route-feature">상위 티어 입장 조건: 지정 캐릭터 레벨 달성 必</li>
        </ul></div>
      </div>
    </div>

    <div class="card">
      <div class="card-head"><span>🗺</span>
        <div class="card-head-title">지역별 성장 기준표</div>
        <div class="card-head-sub">지역 완료 기준 누계 플레이 타임</div>
      </div>
      <div class="card-body">
        <table>
          <tr><th>지역</th><th>레벨 구간</th><th>완료 레벨</th><th>누계 시간</th><th>구간 소요</th><th>숙련도 티어</th></tr>
          {area_rows}
        </table>
      </div>
    </div>
  </div>

  <!-- CHAR LEVEL ─────────────── -->
  <div id="grw-charlevel" style="display:none">
    <div class="card">
      <div class="card-head"><span>📊</span>
        <div class="card-head-title">캐릭터 레벨 경험치 테이블 (1~100)</div>
        <div class="card-head-sub">필요 XP · 누계 시간 · 지역</div>
      </div>
      <div class="card-body" style="padding:0">
        <div class="tbl-wrap" style="max-height:520px;overflow-y:auto">
          <table id="grw-lv-tbl">
            <thead style="position:sticky;top:0;z-index:2">
              <tr><th>레벨</th><th>지역</th><th>필요 XP</th><th>누계 XP</th><th>누계 시간</th><th>성장폭</th></tr>
            </thead>
            <tbody id="grw-lv-body"></tbody>
          </table>
        </div>
      </div>
    </div>
  </div>

  <!-- PROFICIENCY ─────────────── -->
  <div id="grw-proficiency" style="display:none">
    <div class="notice-card" style="margin-bottom:16px">
      <span class="notice-char">⚠️</span>
      <div class="notice-text">
        <strong>숙련도 티어 규칙</strong> — 자신의 숙련도 티어 이상의 몬스터를 대상으로만 경험치 획득 가능합니다.
        <br>예) 2티어 숙련도 → 2티어 이상 몬스터 사냥 시에만 숙련도 XP 획득
      </div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px">
      {prof_cards}
    </div>
  </div>

</section>

<script>
// ── GROWTH DATA ──────────────────────────────────────────────────────────
const GROWTH_LEVEL = {js_levels};
const GROWTH_PROF  = {js_tiers};
const TIER_COLORS_G = {json.dumps(TIER_COLORS)};

const AREA_ZONE = {{
  '이니스 섬':'1T','솔즈리드 반도':'2T','릴리엇 구릉지':'3T',
  '가랑돌 평원':'4T','하얀 숲':'4T','마리아 노플':'5T',
  '황금 평원':'5T','지옥 늪지대':'6T','긴 모래톱':'7T','(미정)':'—'
}};

function renderGrowthLevel() {{
  const body = document.getElementById('grw-lv-body');
  if (!body || body.children.length > 0) return;
  const maxXp = Math.max(...GROWTH_LEVEL.map(r=>r[3]));
  let prev = null;
  body.innerHTML = GROWTH_LEVEL.map(r => {{
    const [lv, h_cum, min_req, xp_lv, xp_cum, area] = r;
    const tc = TIER_COLORS_G[AREA_ZONE[area]] || '#888';
    const isMile = lv % 10 === 0;
    const growth = prev ? ((xp_lv/prev - 1)*100).toFixed(0)+'%↑' : '—';
    prev = xp_lv;
    const bg = isMile ? 'background:rgba(232,192,80,0.07)' : '';
    return `<tr style="${{bg}}">
      <td style="font-weight:${{isMile?700:400}};color:${{isMile?'var(--gold)':'inherit'}}">${{lv}}</td>
      <td style="color:${{tc}};font-size:0.78rem">${{area}}</td>
      <td class="td-num">${{xp_lv.toLocaleString()}}</td>
      <td class="td-num" style="color:var(--text2)">${{xp_cum.toLocaleString()}}</td>
      <td class="td-num">${{h_cum}}h</td>
      <td style="color:var(--text3);font-size:0.75rem">${{growth}}</td>
    </tr>`;
  }}).join('');
}}
renderGrowthLevel();
</script>
'''


# ── 3. HTML 파일 수정 ──────────────────────────────────────────────────────
with open(HTML_IN, encoding='utf-8') as f:
    html = f.read()

# (a) 성장 섹션 교체
old_growth = re.search(
    r'<!-- ══ GROWTH ══ -->.*?</section>',
    html, re.DOTALL
)
if old_growth:
    html = html[:old_growth.start()] + build_growth_section() + html[old_growth.end():]
    print('✅ 성장 섹션 교체 완료')
else:
    print('⚠ 성장 섹션을 찾지 못했습니다')

# (b) 네비게이션 뱃지: 성장 밸런스 "준비중" → "데이터 있음"
html = html.replace(
    '성장 밸런스 <span class="nav-badge">준비중</span>',
    '성장 밸런스 <span class="nav-badge ok">데이터 있음</span>'
)

# (c) Overview 성장 밸런스 상태 업데이트
html = html.replace(
    '<div class="cat-item-name">캐릭터 레벨</div>\n          <span class="cat-item-status status-soon">⏳ 업로드 대기</span>',
    '<div class="cat-item-name">캐릭터 레벨</div>\n          <span class="cat-item-status status-ok">✓ 데이터 확인</span>'
)
html = html.replace(
    '<div class="cat-item-name">숙련도</div>\n          <span class="cat-item-status status-soon">⏳ 업로드 대기</span>',
    '<div class="cat-item-name">숙련도</div>\n          <span class="cat-item-status status-ok">✓ 데이터 확인</span>'
)

# (d) 스킬 레벨 표기 수정 (Lv.0→5 → Lv.1→5)
html = html.replace('Lv.0→5 성장', 'Lv.1→5 성장')
# (e) 탭 한글화
html = html.replace('🏠 Overview', '🏠 개요')

with open(HTML_OUT, 'w', encoding='utf-8') as f:
    f.write(html)

# 원본(카카오톡 받은 파일)도 동기화
import shutil
shutil.copy2(HTML_OUT, HTML_SYNC)

print(f'✅ 저장: {HTML_OUT}')
print(f'✅ 동기화: {HTML_SYNC}')
print(f'   원본 크기: {len(open(HTML_IN,encoding="utf-8").read()):,}자')
print(f'   출력 크기: {len(html):,}자')
