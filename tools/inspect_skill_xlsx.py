"""X7_무기_스킬.xlsx 내용 덤프 (구조 파악용)"""
import openpyxl

XLSX_PATH = r"C:\AI_simulator\기획서\X7_무기_스킬.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"[{sheet_name}] max_row={ws.max_row}, max_col={ws.max_column}")
    print(f"{'='*60}")
    # 첫 30행, 각 행의 첫 12열 출력
    for r in range(1, min(31, ws.max_row + 1)):
        row_vals = []
        for c in range(1, 13):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:20] if v is not None else "")
        # 비어있는 행은 건너뜀
        if any(row_vals):
            print(f"  R{r:2d}: {' | '.join(row_vals)}")
