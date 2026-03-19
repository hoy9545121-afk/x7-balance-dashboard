"""
X7_장비_제작_및_드랍.xlsx — 📋 장비 아이템 리스트 CID/이름 동기화
  기준: X7_방어구_스킬.xlsx 아이템 리스트

매칭 키: (획득방법, 재질, 파츠, 패시브 sorted tuple)
  ① 완전 4키 매칭  (획득+재질+파츠+패시브)  ← 최우선
  ② 재질 제외 3키  (획득+파츠+패시브)       ← 재질 불일치 케이스
  ③ 미매칭 → ⚠️ 경고 출력

패시브 특이 처리:
  방어구_스킬 "마력 자연 회복 ×2" → ("마력 자연 회복", "마력 자연 회복") 로 전개
"""

import re, os
from openpyxl import load_workbook
from copy import copy

# ── 유틸 ──────────────────────────────────────────────────────
def copy_fmt(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.border      = copy(src.border)
        dst.alignment   = copy(src.alignment)
        dst.number_format = src.number_format

def extract_ko_passives(p1, p2):
    """방어구_스킬 패시브1/패시브2 → 한국어 이름 리스트
    '×N' 표기 처리: '마력 자연 회복 ×2' → ['마력 자연 회복', '마력 자연 회복']
    """
    result = []
    for s in [p1, p2]:
        if not s or s == '-':
            continue
        for m in re.finditer(r'\(([^)]+)\)', s):
            ko = m.group(1).strip()
            xn = re.match(r'^(.*?)\s+×(\d+)$', ko)
            if xn:
                result.extend([xn.group(1).strip()] * int(xn.group(2)))
            else:
                result.append(ko)
    return result

def parse_item_name(name):
    """'2T 판금 투구(전리품 제작)' → (tier, material, part_en, acq)"""
    tier = name[:2]
    mat  = "판금" if "판금" in name else "가죽" if "가죽" in name else "천"
    part = ("아머"  if "갑옷" in name else
            "헬멧"  if "투구" in name else
            "장갑"  if "장갑" in name else "신발")
    acq  = ("기본 제작"     if "기본템"      in name else
            "전리품 제작"   if "전리품 제작" in name else
            "던전 코어 제작" if "던전 코어"   in name else "보스 드랍")
    return tier, mat, part, acq

# ════════════════════════════════════════════════════════════════
# STEP 1: 방어구_스킬 아이템 리스트 → 룩업 테이블 구성
# ════════════════════════════════════════════════════════════════
wb_skill = load_workbook("기획서/X7_방어구_스킬.xlsx")
ws_skill = wb_skill["아이템 리스트"]

# key4: (acq, mat, part, pk) → (cid, name)
# key3: (acq, part, pk)      → (cid, name)
map4, map3 = {}, {}

for row in ws_skill.iter_rows(min_row=2, values_only=True):
    cid, name, _, p1, p2 = row[0], row[1], row[2], row[3], row[4]
    if not isinstance(cid, int):
        continue

    tier, mat, part, acq = parse_item_name(name)
    passives = extract_ko_passives(p1, p2)
    pk = tuple(sorted(passives))

    k4 = (acq, mat, part, pk)
    k3 = (acq, part, pk)

    if k4 not in map4:
        map4[k4] = (cid, name)
    if k3 not in map3:
        map3[k3] = (cid, name)

print(f"[방어구_스킬 룩업] 4키: {len(map4)}  3키: {len(map3)}")

# ════════════════════════════════════════════════════════════════
# STEP 2: 장비_제작_및_드랍 📋 장비 아이템 리스트 갱신
# ════════════════════════════════════════════════════════════════
wb = load_workbook("기획서/X7_장비_제작_및_드랍.xlsx")
ws = wb["📋 장비 아이템 리스트"]

ACQ_MAP = {
    "기본 제작":       "기본 제작",
    "전리품 제작":     "전리품 제작",
    "던전 코어 제작":  "던전 코어 제작",
    "보스 드랍":       "보스 드랍",
}

matched4 = matched3 = unmatched = 0
material_mismatch = []
changes = []

for ri in range(4, ws.max_row + 1):
    tier = ws.cell(row=ri, column=1).value
    if tier not in ("1T", "2T", "3T"):
        continue

    acq  = ws.cell(row=ri, column=2).value
    mat  = ws.cell(row=ri, column=3).value
    part = ws.cell(row=ri, column=4).value
    p1   = ws.cell(row=ri, column=5).value
    p2   = ws.cell(row=ri, column=6).value
    p3   = ws.cell(row=ri, column=7).value

    old_cid  = ws.cell(row=ri, column=9).value
    old_name = ws.cell(row=ri, column=10).value

    # 패시브 키 생성
    passives = [v for v in [p1, p2, p3] if v and v != '-']
    pk = tuple(sorted(passives))

    acq_norm = ACQ_MAP.get(str(acq).strip(), str(acq).strip())

    # ① 4키 완전 매칭
    k4 = (acq_norm, mat, part, pk)
    if k4 in map4:
        new_cid, new_name = map4[k4]
        matched4 += 1
        match_type = "4키"
        mat_ok = True
    else:
        # ② 3키 재질 제외 매칭
        k3 = (acq_norm, part, pk)
        if k3 in map3:
            new_cid, new_name = map3[k3]
            matched3 += 1
            match_type = "3키"
            # 재질 불일치 여부 체크
            _, correct_mat, _, _ = parse_item_name(new_name)
            mat_ok = (mat == correct_mat)
            if not mat_ok:
                material_mismatch.append((ri, mat, correct_mat, new_name, pk))
        else:
            new_cid  = None
            new_name = f"⚠️ 미매칭 ({acq}/{mat}/{part}/{pk})"
            unmatched += 1
            match_type = "❌"
            mat_ok = True

    # 변경사항 기록
    if new_cid != old_cid:
        changes.append((ri, tier, acq, mat, part, old_cid, new_cid, old_name, new_name))

    # 셀 업데이트
    ref = ws.cell(row=ri, column=1)
    for col, val in [(9, new_cid), (10, new_name)]:
        cell = ws.cell(row=ri, column=col, value=val)
        copy_fmt(ref, cell)

print(f"\n[매칭 결과] 4키: {matched4}  3키: {matched3}  미매칭: {unmatched}")

# ════════════════════════════════════════════════════════════════
# STEP 3: 결과 보고
# ════════════════════════════════════════════════════════════════
if changes:
    print(f"\n[CID 변경] {len(changes)}건:")
    for ri, tier, acq, mat, part, old_c, new_c, old_n, new_n in changes:
        print(f"  row{ri:3d} {tier} {acq:7s} {mat:3s} {part:3s}  "
              f"{str(old_c):>10} → {str(new_c):>10}  ({old_n} → {new_n})")

if material_mismatch:
    print(f"\n⚠️  [재질 불일치] {len(material_mismatch)}건 (CID는 갱신됨, 재질 컬럼은 유지):")
    for ri, cur_mat, correct_mat, name, pk in material_mismatch:
        print(f"  row{ri:3d}  현재재질={cur_mat}  올바른재질={correct_mat}  이름={name}")
        print(f"         패시브={pk}")

# ════════════════════════════════════════════════════════════════
# STEP 4: 저장
# ════════════════════════════════════════════════════════════════
TMP = "기획서/X7_장비_제작_및_드랍_tmp.xlsx"
wb.save(TMP)
try:
    os.replace(TMP, "기획서/X7_장비_제작_및_드랍.xlsx")
    print("\n✅ 저장: X7_장비_제작_및_드랍.xlsx")
    LOAD = "기획서/X7_장비_제작_및_드랍.xlsx"
except PermissionError:
    print(f"\n⚠️  Excel 열려있음 → 임시파일: {TMP}")
    LOAD = TMP

# ════════════════════════════════════════════════════════════════
# STEP 5: 최종 검증 — CID 전수 확인
# ════════════════════════════════════════════════════════════════
wb2 = load_workbook(LOAD)
ws2 = wb2["📋 장비 아이템 리스트"]

total = ok = dup = no_cid = 0
cid_seen = {}

for row in ws2.iter_rows(min_row=4, values_only=True):
    if row[0] not in ("1T","2T","3T"):
        continue
    total += 1
    cid  = row[8]
    name = row[9]
    if not cid or "⚠️" in str(name):
        no_cid += 1
    elif cid in cid_seen:
        dup += 1
        print(f"  ⚠️ 중복 CID {cid}: row{cid_seen[cid]} vs 현재 ({name})")
    else:
        cid_seen[cid] = row
        ok += 1

print(f"\n=== 최종 검증 ===")
print(f"  총 아이템: {total}  /  정상: {ok}  /  미매칭: {no_cid}  /  중복CID: {dup}")
if ok == 75:
    print("  ✅ 75개 전부 매칭 완료!")
