"""숙련도 옵션 기획서 Excel 생성 스크립트"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 색상 상수 ──────────────────────────────────────────────────
BG_DARK    = '001A1F3A'
BG_MED     = '001E2040'
BG_DARK2   = '00181830'
BG_HEADER  = '000F2D5A'
BG_SECTION = '001A3A6A'
FC_TITLE   = '00F5D878'
FC_SECTION = '00E8C050'
FC_VALUE   = '00EEE8D8'
FC_DIM     = '007A6A5A'
FC_NOTE    = '009A7A28'

TIER_FC = {
    1: '008A9BB0', 2: '004BE8A0', 3: '004B8CE8',
    4: '00C89040', 5: '00A04BE8', 6: '00E84B4B', 7: '00E8B84B'
}
BG_ROW = [BG_DARK2, BG_MED]


def c(ws, row, col, val, fc=FC_VALUE, bg=BG_DARK, bold=False, sz=10,
      ha='center', va='center', wrap=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name='맑은 고딕', color=fc, bold=bold, size=sz)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    return cell


# ── 숙련도 수치 데이터 ────────────────────────────────────────
# 기준ATK = 해당 티어 최고등급 무기 0강 ATK
# T1:80, T2:120, T3:180, T4:260, T5:360, T6:480, T7:600
# 3% DPS 등가 (티어 총합 기준)  — % 계열 소수점 1자리 허용
# ATK(flat)·HP(flat)·SkillAccel은 정수 / CritChance·AtkSpd는 소수점 1자리
# HP flat 기준: 기준HP = 1500+(참조레벨-1)×100, 티어별 ~3% 환산
#   T1(lv10=2400):70  T2(lv20=3400):100  T3(lv30=4400):45  T4(lv40=5400):55
#   T5(lv50=6400):40  T6(lv60=7400):45   T7(lv70=8400):25
# (tier, level, ref_atk, ATK_flat, HP_flat, CritChance%, AtkSpd%, SkillAccel)
WEAPON_DATA = [
    (1, 1, 80,   2,  70,  1.8, 3.0, 3),
    (2, 1, 120,  4, 100,  1.8, 3.0, 3),
    (3, 1, 180,  2,  45,  0.6, 1.0, 1),
    (3, 2, 180,  2,  45,  0.6, 1.0, 1),
    (3, 3, 180,  2,  45,  0.6, 1.0, 1),
    (4, 1, 260,  2,  55,  0.6, 1.0, 1),
    (4, 2, 260,  2,  55,  0.6, 1.0, 1),
    (4, 3, 260,  2,  55,  0.6, 1.0, 1),
    (5, 1, 360,  2,  40,  0.4, 0.6, 1),
    (5, 2, 360,  2,  40,  0.4, 0.6, 1),
    (5, 3, 360,  2,  40,  0.4, 0.6, 1),
    (5, 4, 360,  2,  40,  0.4, 0.6, 1),
    (5, 5, 360,  2,  40,  0.4, 0.6, 1),
    (6, 1, 480,  3,  45,  0.4, 0.6, 1),
    (6, 2, 480,  3,  45,  0.4, 0.6, 1),
    (6, 3, 480,  3,  45,  0.4, 0.6, 1),
    (6, 4, 480,  3,  45,  0.4, 0.6, 1),
    (6, 5, 480,  3,  45,  0.4, 0.6, 1),
    (7, 1, 600,  2,  25,  0.2, 0.3, 1),
    (7, 2, 600,  2,  25,  0.2, 0.3, 1),
    (7, 3, 600,  2,  25,  0.2, 0.3, 1),
    (7, 4, 600,  2,  25,  0.2, 0.3, 1),
    (7, 5, 600,  2,  25,  0.2, 0.3, 1),
    (7, 6, 600,  2,  25,  0.2, 0.3, 1),
    (7, 7, 600,  2,  25,  0.2, 0.3, 1),
    (7, 8, 600,  2,  25,  0.2, 0.3, 1),
    (7, 9, 600,  2,  25,  0.2, 0.3, 1),
]

# (tier, level, ref_atk, DmgRed%, CritDmg%, DmgUp%)
# 1파츠 기준 — 4파츠 착용 합산 기준 ~3% DPS 등가  — % 소수점 1자리 허용
# 가죽 CritDmg: CC=25% 기준, 0.75%/파츠/티어 → ΔCritDmg = 0.75%×6 = 4.5%/tier
# 판금·천: 0.75%/tier/파츠 → T3+ 0.3%↓ 소수로 표시
ARMOR_DATA = [
    (1, 1, 80,   0.8, 4.5, 0.8),
    (2, 1, 120,  0.8, 4.5, 0.8),
    (3, 1, 180,  0.3, 1.5, 0.3),
    (3, 2, 180,  0.3, 1.5, 0.3),
    (3, 3, 180,  0.3, 1.5, 0.3),
    (4, 1, 260,  0.3, 1.5, 0.3),
    (4, 2, 260,  0.3, 1.5, 0.3),
    (4, 3, 260,  0.3, 1.5, 0.3),
    (5, 1, 360,  0.2, 0.9, 0.2),
    (5, 2, 360,  0.2, 0.9, 0.2),
    (5, 3, 360,  0.2, 0.9, 0.2),
    (5, 4, 360,  0.2, 0.9, 0.2),
    (5, 5, 360,  0.2, 0.9, 0.2),
    (6, 1, 480,  0.2, 0.9, 0.2),
    (6, 2, 480,  0.2, 0.9, 0.2),
    (6, 3, 480,  0.2, 0.9, 0.2),
    (6, 4, 480,  0.2, 0.9, 0.2),
    (6, 5, 480,  0.2, 0.9, 0.2),
    (7, 1, 600,  0.1, 0.5, 0.1),
    (7, 2, 600,  0.1, 0.5, 0.1),
    (7, 3, 600,  0.1, 0.5, 0.1),
    (7, 4, 600,  0.1, 0.5, 0.1),
    (7, 5, 600,  0.1, 0.5, 0.1),
    (7, 6, 600,  0.1, 0.5, 0.1),
    (7, 7, 600,  0.1, 0.5, 0.1),
    (7, 8, 600,  0.1, 0.5, 0.1),
    (7, 9, 600,  0.1, 0.5, 0.1),
]

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# SHEET 1: 설계 의도
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '📋 설계 의도'
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 65
ws1.column_dimensions['C'].width = 26
ws1.row_dimensions[1].height = 30

c(ws1, 1, 1, '⚙  숙련도 옵션 기획서 — 설계 의도', FC_TITLE, BG_DARK, True, 14, 'left')
ws1.merge_cells('A1:C1')
c(ws1, 2, 1, '숙련도를 올리면 해당 숙련도에 대응하는 장비 능력치 옵션을 획득한다.',
  FC_DIM, BG_DARK, sz=9, ha='left')
ws1.merge_cells('A2:C2')

# 설계 원칙
c(ws1, 4, 1, '◆ 핵심 설계 원칙', FC_SECTION, BG_HEADER, True, 11, 'left')
ws1.merge_cells('A4:C4')
PRINC = [
    ('장비 컨셉 강화',
     '장비 종류별 특성에 맞는 능력치를 부여하여 장비의 정체성을 강화한다.'),
    ('3% 등가 설계',
     '각 티어의 숙련도 총합이 해당 티어 유저 스펙(최고등급·0강 기준 ATK)의 약 3% DPS 향상과 등가.\n'
     '  ex) T3 기준ATK = 180,  3% ≈ 5 ATK 상당 / T7 기준ATK = 600,  3% = 18 ATK 상당\n'
     '  ※ T5+ HP·CC·Spd·가속은 정수 하한선(1%/단계) 적용으로 실효 기여가 3%를 소폭 상회할 수 있음'),
    ('동가값 보장',
     '스탯 종류가 달라도 동일한 DPS·생존 기여도를 제공한다.\n'
     '  기준: 공격력 기준 역산  →  ΔDPS 공식으로 각 스탯 등가 수치 산출'),
    ('티어별 차등',
     '티어가 높을수록 레벨 수가 많아 1단계당 수치는 작아지지만, 티어 총합 기여도는 동일하다.\n'
     '  T1~T2: 1단계 (큰 한 방)  /  T3~T4: 3단계  /  T5~T6: 5단계  /  T7: 9단계'),
]
for i, (k, v) in enumerate(PRINC):
    r = 5 + i
    ws1.row_dimensions[r].height = 40
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws1, r, 1, k, '00C8E0FF', bg, True, 10, 'left')
    c(ws1, r, 2, v, FC_VALUE, bg, sz=9, ha='left', wrap=True)
    ws1.merge_cells(f'B{r}:C{r}')

# 숙련도 종류
c(ws1, 10, 1, '◆ 숙련도 종류 및 옵션', FC_SECTION, BG_HEADER, True, 11, 'left')
ws1.merge_cells('A10:C10')

c(ws1, 11, 1, '▶ 무기 숙련도', FC_SECTION, BG_DARK, True, 10, 'left')
ws1.merge_cells('A11:C11')
for ci, hh in enumerate(['무기 종류', '옵션', '스탯 필드']):
    c(ws1, 12, ci+1, hh, FC_SECTION, BG_SECTION, True, 9)
WPN = [('한손검', '최대 체력',       'MaxHpVary (flat)'),
       ('양손검', '공격력',          'AttackVary'),
       ('단검',   '치명타 확률',     'CriVaryper'),
       ('활',     '공격 속도',       'AtkSpeedVaryper'),
       ('지팡이', '스킬 가속',       'SkillCooldownAccVary')]
for i, (wpn, opt, fld) in enumerate(WPN):
    r = 13 + i
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws1, r, 1, wpn, FC_VALUE, bg, True)
    c(ws1, r, 2, opt, FC_SECTION, bg)
    c(ws1, r, 3, fld, FC_NOTE, bg, sz=9)

c(ws1, 19, 1, '▶ 방어구 숙련도', FC_SECTION, BG_DARK, True, 10, 'left')
ws1.merge_cells('A19:C19')
for ci, hh in enumerate(['방어구 재질', '옵션', '스탯 필드']):
    c(ws1, 20, ci+1, hh, FC_SECTION, BG_SECTION, True, 9)
ARM = [('판금', '받는 피해 감소',   'DamageDownVaryper'),
       ('가죽', '치명타 피해 증가', 'CriDamageVaryper'),
       ('천',   '주는 피해 증가',   'DamageUpVaryper')]
for i, (arm, opt, fld) in enumerate(ARM):
    r = 21 + i
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws1, r, 1, arm, FC_VALUE, bg, True)
    c(ws1, r, 2, opt, FC_SECTION, bg)
    c(ws1, r, 3, fld, FC_NOTE, bg, sz=9)

# 단계 구조
c(ws1, 25, 1, '◆ 숙련도 단계 구조 (티어 > 레벨)', FC_SECTION, BG_HEADER, True, 11, 'left')
ws1.merge_cells('A25:C25')
for ci, hh in enumerate(['티어', '레벨 범위', '총 단계 수']):
    c(ws1, 26, ci+1, hh, FC_SECTION, BG_SECTION, True, 9)
STAGES = [(1,'Lv1','1단계'), (2,'Lv1','1단계'), (3,'Lv1~3','3단계'),
          (4,'Lv1~3','3단계'), (5,'Lv1~5','5단계'), (6,'Lv1~5','5단계'), (7,'Lv1~9','9단계')]
for i, (t, lv, st) in enumerate(STAGES):
    r = 27 + i
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws1, r, 1, f'T{t}', TIER_FC[t], bg, True)
    c(ws1, r, 2, lv, FC_VALUE, bg)
    c(ws1, r, 3, st, FC_VALUE, bg)
c(ws1, 34, 1, '합계', '00FFFFFF', BG_HEADER, True)
c(ws1, 34, 3, '27단계', FC_TITLE, BG_HEADER, True)

print('Sheet1 완료')

# ═══════════════════════════════════════════════════════════════
# SHEET 2: 무기 숙련도
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('⚔ 무기 숙련도')
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 15
ws2.column_dimensions['H'].width = 15
ws2.row_dimensions[1].height = 30

c(ws2, 1, 1, '⚔  무기 숙련도 옵션 수치표', FC_TITLE, BG_DARK, True, 14, 'left')
ws2.merge_cells('A1:H1')
c(ws2, 2, 1,
  '※ 각 티어 총합 기준 ~3% DPS 등가  |  기준ATK = 해당 티어 최고등급 무기 0강  |  CC=10% 기준 역산',
  FC_NOTE, BG_DARK, sz=9, ha='left')
ws2.merge_cells('A2:H2')

# 헤더
ws2.row_dimensions[4].height = 30
WHDR = [(1,'티어'), (2,'레벨'), (3,'기준ATK'),
        (4,'양손검\n공격력\n(누적 +ATK)'),
        (5,'한손검\n최대체력\n(누적 +HP)'),
        (6,'단검\n치명타확률\n(누적 +%)'),
        (7,'활\n공격속도\n(누적 +%)'),
        (8,'지팡이\n스킬가속\n(누적)')]
for col, lbl in WHDR:
    c(ws2, 4, col, lbl, FC_SECTION, BG_SECTION, True, 9, wrap=True)

# 색상 정의 (옵션별)
WPN_COL = {'atk':'00FFB347','hp':'00FF7777','crit':'00FFD700',
           'spd':'0098FB98','accel':'00CC88FF'}

prev_tier = None
cum = [0, 0, 0, 0, 0]  # 누적합: atk, hp, crit, spd, accel
for i, (tier, lv, ref, atk, hp, crit, spd, accel) in enumerate(WEAPON_DATA):
    r = 5 + i
    ws2.row_dimensions[r].height = 18
    bg = BG_DARK2 if tier % 2 == 1 else BG_MED
    tier_bold = (tier != prev_tier)
    cum[0] += atk; cum[1] += hp; cum[2] += crit; cum[3] += spd; cum[4] += accel
    c(ws2, r, 1, f'T{tier}', TIER_FC[tier], bg, tier_bold, 10)
    c(ws2, r, 2, f'Lv{lv}', FC_VALUE, bg, sz=10)
    c(ws2, r, 3, ref, FC_DIM, bg, sz=9)
    c(ws2, r, 4, f'+{cum[0]}',        WPN_COL['atk'],   bg, True, 10)
    c(ws2, r, 5, f'+{cum[1]}',        WPN_COL['hp'],    bg, True, 10)
    c(ws2, r, 6, f'+{cum[2]:.1f}%',  WPN_COL['crit'],  bg, True, 10)
    c(ws2, r, 7, f'+{cum[3]:.1f}%',  WPN_COL['spd'],   bg, True, 10)
    c(ws2, r, 8, f'+{cum[4]}',        WPN_COL['accel'], bg, True, 10)
    prev_tier = tier

# 누적 합계 행
r = 5 + len(WEAPON_DATA)
ws2.row_dimensions[r].height = 22
c(ws2, r, 1, '누적 합계', FC_TITLE, BG_HEADER, True, 10)
ws2.merge_cells(f'A{r}:C{r}')
totW = [sum(x[j] for x in WEAPON_DATA) for j in range(3, 8)]
fmts = [f'+{totW[0]} ATK',
        f'+{int(totW[1])} HP', f'+{totW[2]:.1f}%', f'+{totW[3]:.1f}%',
        f'+{int(totW[4])}']
fcs  = [WPN_COL['atk'],WPN_COL['hp'],WPN_COL['crit'],WPN_COL['spd'],WPN_COL['accel']]
for col, (txt, fc) in enumerate(zip(fmts, fcs)):
    c(ws2, r, col+4, txt, fc, BG_HEADER, True, 11)

# 캡 참고행
r2 = r + 1
ws2.row_dimensions[r2].height = 16
caps = ['캡: 300 (ATK)', '캡: MaxHpVary', '캡: 80%', '캡: 300%', '캡: 100']
for col, cap in enumerate(caps):
    c(ws2, r2, col+4, cap, FC_DIM, BG_DARK, sz=8)

print('Sheet2 완료')

# ═══════════════════════════════════════════════════════════════
# SHEET 3: 방어구 숙련도
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('🛡 방어구 숙련도')
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 20
ws3.column_dimensions['F'].width = 18
ws3.row_dimensions[1].height = 30

c(ws3, 1, 1, '🛡  방어구 숙련도 옵션 수치표', FC_TITLE, BG_DARK, True, 14, 'left')
ws3.merge_cells('A1:F1')
c(ws3, 2, 1,
  '※ 1파츠 기준  |  4파츠 동일 재질 착용 합산 ~3% 등가  |  가죽 CritDmg: CC=25%, mult=1.5 기준 역산  |  판금·천: T3+ 1파츠 <1% → 정수 생략',
  FC_NOTE, BG_DARK, sz=9, ha='left')
ws3.merge_cells('A2:F2')

ws3.row_dimensions[4].height = 30
AHDR = [(1,'티어'), (2,'레벨'), (3,'기준ATK'),
        (4,'판금\n받는피해감소\n(누적 +%)'),
        (5,'가죽\n치명타피해증가\n(누적 +%)'),
        (6,'천\n주는피해증가\n(누적 +%)')]
for col, lbl in AHDR:
    c(ws3, 4, col, lbl, FC_SECTION, BG_SECTION, True, 9, wrap=True)

ARM_COL = {'red':'004FBBFF','cdmg':'00FFD700','up':'00FFA07A'}

prev_tier = None
cum_a = [0, 0, 0]  # 누적합: dmgred, critdmg, dmgup
for i, (tier, lv, ref, dmgred, critdmg, dmgup) in enumerate(ARMOR_DATA):
    r = 5 + i
    ws3.row_dimensions[r].height = 18
    bg = BG_DARK2 if tier % 2 == 1 else BG_MED
    tier_bold = (tier != prev_tier)
    cum_a[0] += dmgred; cum_a[1] += critdmg; cum_a[2] += dmgup
    c(ws3, r, 1, f'T{tier}', TIER_FC[tier], bg, tier_bold, 10)
    c(ws3, r, 2, f'Lv{lv}', FC_VALUE, bg, sz=10)
    c(ws3, r, 3, ref, FC_DIM, bg, sz=9)
    c(ws3, r, 4, f'+{cum_a[0]:.1f}%', ARM_COL['red'],  bg, True, 10)
    c(ws3, r, 5, f'+{cum_a[1]:.1f}%', ARM_COL['cdmg'], bg, True, 10)
    c(ws3, r, 6, f'+{cum_a[2]:.1f}%', ARM_COL['up'],   bg, True, 10)
    prev_tier = tier

r = 5 + len(ARMOR_DATA)
ws3.row_dimensions[r].height = 22
c(ws3, r, 1, '1파츠 누적', FC_TITLE, BG_HEADER, True, 10)
ws3.merge_cells(f'A{r}:C{r}')
totA = [sum(x[j] for x in ARMOR_DATA) for j in range(3, 6)]
for col, (val, fc) in enumerate(zip(totA,
    [ARM_COL['red'], ARM_COL['cdmg'], ARM_COL['up']])):
    c(ws3, r, col+4, f'+{val:.1f}%', fc, BG_HEADER, True, 11)

r2 = r + 1
ws3.row_dimensions[r2].height = 18
c(ws3, r2, 1, '4파츠 합산', '00AADDFF', BG_DARK2, True, 10)
ws3.merge_cells(f'A{r2}:C{r2}')
for col, (val, fc) in enumerate(zip([v*4 for v in totA],
    [ARM_COL['red'], ARM_COL['cdmg'], ARM_COL['up']])):
    c(ws3, r2, col+4, f'+{val:.1f}%', fc, BG_DARK2, True, 11)

r3 = r2 + 1
ws3.row_dimensions[r3].height = 16
for col, cap in enumerate(['캡: 50%', '캡: 300%', '캡: 100%']):
    c(ws3, r3, col+4, cap, FC_DIM, BG_DARK, sz=8)

print('Sheet3 완료')

# ═══════════════════════════════════════════════════════════════
# SHEET 4: 캡 검증 & 등가 산출 근거
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('📊 캡 검증 & 등가 산출')
ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 8
ws4.column_dimensions['E'].width = 28
ws4.column_dimensions['F'].width = 20
ws4.row_dimensions[1].height = 30

c(ws4, 1, 1, '📊  캡 검증 & DPS 등가 산출 근거', FC_TITLE, BG_DARK, True, 14, 'left')
ws4.merge_cells('A1:F1')

# 캡 검증
c(ws4, 3, 1, '◆ 캡 검증 (만렙 전 티어 누적 최대)', FC_SECTION, BG_HEADER, True, 11, 'left')
ws4.merge_cells('A3:F3')
for ci, hh in enumerate(['숙련도 옵션', '누적 최대', '게임 캡', '상태', '스탯 필드', '비고']):
    c(ws4, 4, ci+1, hh, FC_SECTION, BG_SECTION, True, 9)

CAP_DATA = [
    ('양손검 — 공격력',      '+61 ATK',           'AttackVary',               '✅', 'AttackVary',           'flat 수치'),
    ('한손검 — 최대 체력',   '+1,120 HP',          'MaxHpVary (flat)',          '✅', 'MaxHpVary',            'flat 수치'),
    ('단검 — 치명타확률%',   '+13.0%',             '80% (CriVaryper)',          '✅', 'CriVaryper',           'CC=10% 기준'),
    ('활 — 공격속도%',       '+20.7%',             '300% (AtkSpeedVaryper)',    '✅', 'AtkSpeedVaryper',      '넉넉'),
    ('지팡이 — 스킬가속',    '+31',                '100 (SkillCooldownAccVary)','✅', 'SkillCooldownAccVary', 'CD×100/(100+n)'),
    ('판금 — 받는피해감소%', '+6.3% (×4=25.2%)',  '50% (DamageDownVaryper)',   '✅', 'DamageDownVaryper',    '1파츠 기준'),
    ('가죽 — 치명타피해%',   '+31.5% (×4=126%)',  '300% (CriDamageVaryper)',   '✅', 'CriDamageVaryper',     'CC=25% 기준'),
    ('천 — 주는피해증가%',   '+6.3% (×4=25.2%)',  '100% (DamageUpVaryper)',    '✅', 'DamageUpVaryper',      '1파츠 기준'),
]
for i, row in enumerate(CAP_DATA):
    r = 5 + i
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    for ci, val in enumerate(row):
        fc = FC_VALUE
        if ci == 0: fc = FC_SECTION
        if ci == 3: fc = '0044FF44'
        if ci == 4: fc = FC_NOTE
        ha = 'left' if ci in (0, 4, 5) else 'center'
        c(ws4, r, ci+1, val, fc, bg, sz=9, ha=ha)

# DPS 등가 산출 근거
c(ws4, 14, 1, '◆ DPS 등가 산출 근거', FC_SECTION, BG_HEADER, True, 11, 'left')
ws4.merge_cells('A14:F14')
for ci, hh in enumerate(['스탯', '기준 조건', '등가 공식', '3% DPS 등가 결과']):
    c(ws4, 15, ci+1, hh, FC_SECTION, BG_SECTION, True, 9)
ws4.merge_cells('C15:E15')

EQUIV = [
    ('공격력 (flat ATK)',   '기준ATK = 티어 최고등급 0강',
     'ΔDPS/DPS = ΔATK / ATK_ref', '3% × ATK_ref'),
    ('공격속도%',           '선형 비례',
     'ΔDPS/DPS = ΔSpdPct / 100', 'ATK%와 1:1 동일'),
    ('치명타확률%',         'CC=10%, CD=200%, mult=1.2',
     'ΔDPS = Δcc × CD / (100 × mult)', '3% DPS → Δcc ≈ 1.8%'),
    ('스킬가속',            '소수 n에서 n% 근사',
     'CD단축 = n/(100+n)',           '+3 accel ≈ 3% 스킬사용↑'),
    ('최대 체력 (flat)',     '기준HP×3% 역산 (티어별 참조레벨)',
     'ref_hp × 3% → flat HP',       '총 +1,120 HP'),
    ('받는피해감소%',       '선형 생존력',
     '+DmgRed% → 유효HP 증가',      '3% 감소 ≈ 3% 생존↑'),
    ('치명타피해%',         'CC=25%, mult=1.5',
     'ΔDPS = CC × ΔCD / (100×mult)', '3% DPS → ΔCD ≈ 18%'),
    ('주는피해증가%',       '선형 비례',
     'ΔDPS/DPS ≈ ΔDmgUp%',          'ATK%와 근사 동일'),
]
for i, (stat, cond, formula, result) in enumerate(EQUIV):
    r = 16 + i
    ws4.row_dimensions[r].height = 18
    bg = BG_DARK2 if i % 2 == 0 else BG_MED
    c(ws4, r, 1, stat,    FC_SECTION, bg, True,  9, 'left')
    c(ws4, r, 2, cond,    FC_DIM,     bg, sz=8,  ha='left')
    c(ws4, r, 3, formula, FC_VALUE,   bg, sz=8,  ha='left')
    ws4.merge_cells(f'C{r}:E{r}')
    c(ws4, r, 6, result,  FC_TITLE,   bg, True,  9)

print('Sheet4 완료')

# ── 저장 ──────────────────────────────────────────────────────
out = 'C:/AI_simulator/기획서/X7_숙련도_옵션_기획서.xlsx'
wb.save(out)
print(f'\n저장 완료: {out}')
print('누적합계 확인:')
print(f'  ATK:         +{sum(x[3] for x in WEAPON_DATA)}')
print(f'  HP (flat):   +{sum(x[4] for x in WEAPON_DATA)} HP')
print(f'  CritChance%: +{sum(x[5] for x in WEAPON_DATA)}%')
print(f'  AtkSpd%:     +{sum(x[6] for x in WEAPON_DATA)}%')
print(f'  SkillAccel:  +{sum(x[7] for x in WEAPON_DATA)}')
print(f'  DmgRed%:     +{sum(x[3] for x in ARMOR_DATA)}%')
print(f'  CritDmg%:    +{sum(x[4] for x in ARMOR_DATA)}%')
print(f'  DmgUp%:      +{sum(x[5] for x in ARMOR_DATA)}%')
