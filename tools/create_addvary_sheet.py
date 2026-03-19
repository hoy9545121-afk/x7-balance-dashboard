"""
추가옵션(AddVary) 시트 생성 스크립트 v2
X7_장비_능력치_기획서.xlsx에 '⚡ 추가옵션' 시트 추가
- 기획 의도 섹션 3개 하위 항목으로 재구성
- 파라미터 행(PARAM_ROW=16, D~I열) 수정 시 모든 FLOOR 수식 자동 갱신
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILEPATH = r'C:\AI_simulator\기획서\X7_장비_능력치_기획서.xlsx'

# ──────────────────────────────────────────────────────────────────────────────
# 색상 팔레트 (기존 기획서 다크 테마 일치)
# ──────────────────────────────────────────────────────────────────────────────
C = {
    'title':   'FF1A1F3A',   # 진한 네이비 — 제목
    'section': 'FF0F2D5A',   # 섹션 헤더
    'param':   'FF1A3A20',   # 파라미터 행 (초록빛 강조)
    'header':  'FF1A1F3A',   # 컬럼 헤더
    'odd':     'FF181830',   # 데이터 홀수행
    'even':    'FF1E2040',   # 데이터 짝수행
    'note':    'FF0D0D1A',   # 설명 텍스트 배경
    'blank':   'FF0D0D1A',   # 구분 공백행
    'white':   'FFFFFFFF',
    'yellow':  'FFFFEE88',   # 범위 텍스트
    'green':   'FF55EE77',   # 파라미터 강조
    'light':   'FFCCCCCC',   # 설명 텍스트
    'dim':     'FF666688',   # N/A 텍스트
    'param_v': 'FFCCFFCC',   # 파라미터 값 색상
}


def f(hex_clr):
    """PatternFill helper"""
    return PatternFill('solid', fgColor=hex_clr)


def ft(name='맑은 고딕', size=10, bold=False, clr='FFFFFFFF', italic=False):
    """Font helper"""
    return Font(name=name, size=size, bold=bold, color=clr, italic=italic)


def al(h='center', v='center', wrap=False, indent=0):
    """Alignment helper"""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def thin():
    s = Side(style='thin', color='FF333355')
    return Border(left=s, right=s, top=s, bottom=s)


def sc(ws, row, col, value, fill_=None, font_=None, align_=None, num_fmt=None):
    """Set cell"""
    cell = ws.cell(row, col)
    cell.value = value
    if fill_:   cell.fill      = fill_
    if font_:   cell.font      = font_
    if align_:  cell.alignment = align_
    if num_fmt: cell.number_format = num_fmt
    return cell


def mr(ws, row, c1, c2, value, fill_=None, font_=None, align_=None):
    """Merge row and set"""
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    return sc(ws, row, c1, value, fill_=fill_, font_=font_, align_=align_)


# ──────────────────────────────────────────────────────────────────────────────
# 베이스 스탯 (0강, 일반 등급)
# ──────────────────────────────────────────────────────────────────────────────
WPN_BASE  = {1:60, 2:80, 3:120, 4:180, 5:260, 6:360, 7:480}
ARM_BASE  = {1:30, 2:40, 3:60,  4:90,  5:130, 6:180, 7:240}
RING_BASE = {3:30, 4:45, 5:65,  6:90,  7:120}
NECK_BASE = {3:60, 4:90, 5:130, 6:180, 7:240}
EAR_BASE  = {3:60, 4:90, 5:130, 6:180, 7:240}

# ──────────────────────────────────────────────────────────────────────────────
# 파일 열기
# ──────────────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILEPATH)

if '⚡ 추가옵션' in wb.sheetnames:
    del wb['⚡ 추가옵션']

ws = wb.create_sheet('⚡ 추가옵션')
ws.sheet_properties.tabColor = 'FFD700'  # 금빛 탭

# 열 너비 (텍스트 잘림 방지)
for col_letter, width in [
    ('A', 2), ('B', 14), ('C', 11),
    ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 8), ('I', 8),
    ('J', 26)
]:
    ws.column_dimensions[col_letter].width = width

# 화면 고정 (행 17 아래에서 스크롤)
ws.freeze_panes = 'B17'

# ──────────────────────────────────────────────────────────────────────────────
# Row 1 — 제목
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 32
mr(ws, 1, 1, 10, '⚡  장비 추가 옵션 (AddVary) — 수치 밸런스 설계',
   fill_=f(C['title']),
   font_=ft(size=13, bold=True),
   align_=al('left', indent=1))

# ──────────────────────────────────────────────────────────────────────────────
# Row 2 — 구분선
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 6
for c in range(1, 11):
    ws.cell(2, c).fill = f(C['blank'])

# ──────────────────────────────────────────────────────────────────────────────
# Row 3 — 기획 의도 섹션 헤더
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 22
mr(ws, 3, 1, 10, '  기획 의도',
   fill_=f('FF1A3A6A'),
   font_=ft(size=11, bold=True),
   align_=al('left', indent=1))

# ──────────────────────────────────────────────────────────────────────────────
# Rows 4~12 — 기획 의도 내용 (3개 소항목)
# ──────────────────────────────────────────────────────────────────────────────
CLR_SUB  = 'FF161E30'   # 소항목 헤더 배경
CLR_NOTE = 'FF0D0D1A'   # 설명 줄 배경

def note_section(ws, r, title, lines):
    """소항목 헤더 + 설명 줄 작성"""
    ws.row_dimensions[r].height = 18
    mr(ws, r, 1, 10, f'  {title}',
       fill_=f(CLR_SUB),
       font_=ft(size=10, bold=True, clr='FFAACCFF'),
       align_=al('left', indent=1))
    r += 1
    for line in lines:
        ws.row_dimensions[r].height = 15
        mr(ws, r, 1, 10, f'      {line}',
           fill_=f(CLR_NOTE),
           font_=ft(size=10, clr=C['light']),
           align_=al('left'))
        r += 1
    return r

r_intent = 4
r_intent = note_section(ws, r_intent,
    '▷ 수치 산출 방식',
    [
        '각 옵션 수치  =  FLOOR( 해당 티어의 기본 능력치 × 비율, 1 )  (소수점 이하 버림)',
        '티어가 높을수록 기본 능력치도 크므로 더 큰 추가 수치 제공  →  티어별 장비 가치 자동 보장',
    ])
r_intent = note_section(ws, r_intent,
    '▷ 슬롯별 적용 능력치',
    [
        '무기 / 반지                       →   AddAttackVary          (추가 공격력)',
        '방어구 / 목걸이 / 귀걸이    →   AddPhysicalDefenseVary  (추가 방어력)',
    ])
r_intent = note_section(ws, r_intent,
    '▷ 확률 설계 철학',
    [
        '옵션 1~3 (낮은 수치)  →  등장 확률 10% / 10% / 10%  |  옵션 4~5  →  20% / 20%  |  옵션 6 (최고)  →  30%',
        '높은 수치가 더 자주 등장하도록 설계  →  풀업 직전의 아쉬운 느낌 유지, 박탈감 최소화',
    ])

# r_intent 는 이제 13 이 됨 (3 헤더 + 6 노트 + 4시작 = 13)

# ──────────────────────────────────────────────────────────────────────────────
# Row 13 — 구분선
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[r_intent].height = 8
for c in range(1, 11):
    ws.cell(r_intent, c).fill = f(C['blank'])
r_intent += 1   # = 14

# ──────────────────────────────────────────────────────────────────────────────
# Row 14 — 옵션 번호 헤더
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[r_intent].height = 18
for c, h in enumerate(['', '옵션 번호', '', '1', '2', '3', '4', '5', '6', ''], 1):
    sc(ws, r_intent, c, h, fill_=f(C['header']), font_=ft(bold=True, size=9), align_=al())
r_intent += 1   # = 15

# ──────────────────────────────────────────────────────────────────────────────
# Row 15 — 등장 확률
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[r_intent].height = 16
for c, v in enumerate(['', '등장 확률', '', 0.10, 0.10, 0.10, 0.20, 0.20, 0.30, ''], 1):
    cell = sc(ws, r_intent, c, v, fill_=f(C['header']), font_=ft(size=10), align_=al())
    if isinstance(v, float):
        cell.number_format = '0%'
r_intent += 1   # = 16

# ──────────────────────────────────────────────────────────────────────────────
# Row 16 — 베이스 대비 % (핵심 파라미터 행 ★)
#   D16~I16 의 %값이 아래 모든 FLOOR 수식에서 참조됨
# ──────────────────────────────────────────────────────────────────────────────
PARAM_ROW = r_intent   # = 16
ws.row_dimensions[PARAM_ROW].height = 20
sc(ws, PARAM_ROW, 1, '', fill_=f(C['param']))
sc(ws, PARAM_ROW, 2, '베이스 대비 %',
   fill_=f(C['param']),
   font_=ft(bold=True, clr=C['green']),
   align_=al())
sc(ws, PARAM_ROW, 3, '',
   fill_=f(C['param']))

pct_vals = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10]
for i, pct in enumerate(pct_vals):
    c = 4 + i   # D=4 ~ I=9
    cell = sc(ws, PARAM_ROW, c, pct,
              fill_=f(C['param']),
              font_=ft(bold=True, size=11, clr=C['param_v']),
              align_=al())
    cell.number_format = '0%'

sc(ws, PARAM_ROW, 10, '★  이 행의 % 값을 수정하면 아래 모든 테이블 수치 자동 갱신',
   fill_=f(C['param']),
   font_=ft(size=9, clr=C['green'], italic=True),
   align_=al('left', indent=1))
r_intent += 1   # = 17

# ──────────────────────────────────────────────────────────────────────────────
# Row 17 — 구분선
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[r_intent].height = 10
for c in range(1, 11):
    ws.cell(r_intent, c).fill = f(C['blank'])

# ──────────────────────────────────────────────────────────────────────────────
# 테이블 빌더 — FLOOR 수식 생성
#   col B = 티어, col C = 베이스 능력치, col D~I = 옵션1~6, col J = 범위
#   FLOOR($C{r}*D$PARAM_ROW, 1)  ← $C 열고정, $PARAM_ROW 행고정
# ──────────────────────────────────────────────────────────────────────────────
def build_table(ws, start_row, title, base_dict, base_col_label):
    r = start_row

    # 섹션 헤더
    ws.row_dimensions[r].height = 22
    mr(ws, r, 1, 10, f'  {title}',
       fill_=f(C['section']),
       font_=ft(size=11, bold=True, clr='FFFFFFAA'),
       align_=al('left', indent=1))
    r += 1

    # 컬럼 헤더
    ws.row_dimensions[r].height = 18
    col_hdrs = ['', '티어', base_col_label, '옵션1', '옵션2', '옵션3', '옵션4', '옵션5', '옵션6', '범위']
    for c, h in enumerate(col_hdrs, 1):
        sc(ws, r, c, h, fill_=f(C['header']), font_=ft(bold=True, size=9), align_=al())
    r += 1

    # 데이터 행
    for idx, tier in enumerate(sorted(base_dict.keys())):
        base = base_dict[tier]
        bg   = C['odd'] if idx % 2 == 0 else C['even']
        ws.row_dimensions[r].height = 16

        sc(ws, r, 1, '',         fill_=f(bg))
        sc(ws, r, 2, f'T{tier}', fill_=f(bg), font_=ft(bold=True), align_=al())
        sc(ws, r, 3, base,       fill_=f(bg), font_=ft(size=10),   align_=al())

        # FLOOR 수식: 옵션 1~6 (D~I = cols 4~9), 파라미터 행 참조
        for opt_i in range(6):
            c       = 4 + opt_i
            pct_col = get_column_letter(c)   # D, E, F, G, H, I
            formula = f'=FLOOR($C{r}*{pct_col}${PARAM_ROW},1)'
            sc(ws, r, c, formula, fill_=f(bg), font_=ft(size=10), align_=al())

        # 범위 표시 (J = col 10)
        range_formula = f'=D{r}&" ~ "&I{r}'
        sc(ws, r, 10, range_formula,
           fill_=f(bg),
           font_=ft(size=10, clr=C['yellow']),
           align_=al())

        r += 1

    return r   # 다음 섹션 시작행 반환


def blank_row(ws, r, height=10):
    ws.row_dimensions[r].height = height
    for c in range(1, 11):
        ws.cell(r, c).fill = f(C['blank'])


# ──────────────────────────────────────────────────────────────────────────────
# 테이블 순서: 무기 → 반지 → 방어구 → 목걸이 → 귀걸이
# ──────────────────────────────────────────────────────────────────────────────
r = r_intent + 1   # Row 18부터 시작

r = build_table(ws, r, '⚔ 무기 — AddAttackVary (T1~T7)',              WPN_BASE,  '베이스 ATK')
blank_row(ws, r); r += 1

r = build_table(ws, r, '💍 반지 — AddAttackVary (T3~T7, ×2 슬롯)',    RING_BASE, '베이스 ATK')
blank_row(ws, r); r += 1

r = build_table(ws, r, '🛡 방어구 — AddPhysicalDefenseVary (T1~T7)',  ARM_BASE,  '베이스 DEF')
blank_row(ws, r); r += 1

r = build_table(ws, r, '📿 목걸이 — AddPhysicalDefenseVary (T3~T7)', NECK_BASE, '베이스 DEF')
blank_row(ws, r); r += 1

r = build_table(ws, r, '👂 귀걸이 — AddPhysicalDefenseVary (T3~T7)', EAR_BASE,  '베이스 DEF')
blank_row(ws, r, 14); r += 1

# ──────────────────────────────────────────────────────────────────────────────
# 요약 테이블 — 슬롯별 옵션 범위 (최솟값 ~ 최댓값)
# ──────────────────────────────────────────────────────────────────────────────
blank_row(ws, r); r += 1

# 요약 헤더
ws.row_dimensions[r].height = 22
mr(ws, r, 1, 10, '  📊 슬롯별 옵션 범위 요약 (최솟값 ~ 최댓값, 5%~10% 기준)',
   fill_=f(C['section']),
   font_=ft(size=11, bold=True, clr='FFFFFFAA'),
   align_=al('left', indent=1))
r += 1

# 요약 컬럼 헤더
ws.row_dimensions[r].height = 16
summ_hdrs = ['', '슬롯 / 능력치', 'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', '']
for c, h in enumerate(summ_hdrs, 1):
    sc(ws, r, c, h, fill_=f(C['header']), font_=ft(bold=True, size=9), align_=al())
r += 1

def rng(base):
    lo = int(base * 0.05)
    hi = int(base * 0.10)
    return f'{lo} ~ {hi}'

summary_slots = [
    ('무기 — AddATK',  WPN_BASE),
    ('반지 — AddATK',  RING_BASE),
    ('방어구 — AddDEF', ARM_BASE),
    ('목걸이 — AddDEF', NECK_BASE),
    ('귀걸이 — AddDEF', EAR_BASE),
]

for idx, (name, bmap) in enumerate(summary_slots):
    bg = C['odd'] if idx % 2 == 0 else C['even']
    ws.row_dimensions[r].height = 16
    sc(ws, r, 1, '', fill_=f(bg))
    sc(ws, r, 2, name, fill_=f(bg), font_=ft(bold=True, size=9),
       align_=al('left', indent=1))
    for tier in range(1, 8):
        c = tier + 2   # T1=col3 … T7=col9
        if tier in bmap:
            sc(ws, r, c, rng(bmap[tier]), fill_=f(bg),
               font_=ft(size=9, clr=C['yellow']), align_=al())
        else:
            sc(ws, r, c, 'N/A', fill_=f(bg),
               font_=ft(size=9, clr=C['dim']), align_=al())
    sc(ws, r, 10, '', fill_=f(bg))
    r += 1

# ──────────────────────────────────────────────────────────────────────────────
# 저장 및 검증
# ──────────────────────────────────────────────────────────────────────────────
wb.save(FILEPATH)
print(f'✅ 저장 완료: {FILEPATH}')
print(f'   시트 목록: {wb.sheetnames}')
print(f'   총 사용 행: {r - 1}')
print()

# 간단 검증 — FLOOR 수식이 들어갔는지 확인
wb2 = openpyxl.load_workbook(FILEPATH, read_only=True)
ws2 = wb2['⚡ 추가옵션']
print(f'=== 파라미터 행 ({PARAM_ROW}행) ===')
for c in range(1, 11):
    cell = ws2.cell(PARAM_ROW, c)
    if cell.value is not None:
        print(f'  {cell.coordinate}: {cell.value}')

# 무기 T1 행 찾기 (Row 19 or nearby)
t1_row = PARAM_ROW + 3   # 예상: header(18) + colhdr(19) + T1(20)
print(f'=== 수식 검증 (무기 T1 ~ Row {t1_row}) ===')
for row in ws2.iter_rows(min_row=t1_row, max_row=t1_row, min_col=1, max_col=10):
    for cell in row:
        if cell.value:
            print(f'  {cell.coordinate}: {cell.value}')
wb2.close()
print()
print(f'⚡ 추가옵션 시트 생성 완료!  (파라미터 행 = Row {PARAM_ROW})')
