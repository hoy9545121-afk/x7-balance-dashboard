"""
X7_방어구_스킬.xlsx 전체 갱신

[STEP 1] 3T 방어구 시트에 누락행 추가
  - 12030038: 3T 가죽 신발(드랍)  — RegenMpVary / AttackVary / PVEDamageUpVaryper
  - 12030039: 3T 천 신발(드랍)    — RegenMpVary / MaxHpVary  / HealAmpVaryper
  - 12030038 → row115에 삽입 (가죽 섹션 끝, 천 섹션 앞)
  - 12030039 → sheet 끝에 추가

[STEP 2] 아이템 리스트 전체 갱신 (방어구 시트 파싱 + CSV 활용)
  패시브1/패시브2 포맷:
    1T (1개) → col4 = "Key (한국어)"              col5 = "-"
    2T (2개) → col4 = "Key1 (한국어1)Key2 (한국어2)" col5 = "-"
    3T (3개) → col4 = "Key1 (한국어1)Key2 (한국어2)" col5 = "Key3 (한국어3)"
  효과타입/Proc/SR → CSV 기준 (1T·3T: CID 직접, 2T: 위치 순서)
"""

import re, os, csv
from openpyxl import load_workbook
from copy import copy

# ── 유틸 ──────────────────────────────────────────────────────
def copy_fmt(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.border      = copy(src.border)
        dst.alignment   = copy(src.alignment)
        dst.number_format = src.number_format

def fmt_passive(key, ko):
    return f"{str(key).strip()} ({str(ko).strip()})"

# ══════════════════════════════════════════════════════════════
# STEP 1 — 3T 방어구 시트에 누락행 삽입
# ══════════════════════════════════════════════════════════════
wb = load_workbook("기획서/X7_방어구_스킬.xlsx")
ws3 = wb["3T 방어구"]

#  새 아이템 데이터 정의
# (part, ko, key, lv1, lv2, lv3, lv4, lv5, cid, name)
NEW_38 = [
    ("👟  신발", "마력 자연 회복",  "RegenMpVary",        8,   9,   10,  12,  14,  12030038, "3T 가죽 신발(드랍)"),
    (None,       "공격력",           "AttackVary",          12,  14,  16,  20,  22,  None,     None),
    (None,       "PVE 피해 증가",    "PVEDamageUpVaryper",  "4%","4.6%","5.3%","6.1%","7%", None, None),
]
NEW_39 = [
    ("👟  신발", "마력 자연 회복",  "RegenMpVary",        8,   9,   10,  12,  14,  12030039, "3T 천 신발(드랍)"),
    (None,       "최대 생명력",      "MaxHpVary",           190, 220, 250, 290, 330, None,     None),
    (None,       "치유력",           "HealAmpVaryper",      "4.3%","5%","5.7%","6.5%","7.5%", None, None),
]

def append_rows(ws, rows_data, style_ref_row):
    """sheet 끝에 rows_data 추가, style은 style_ref_row 복사"""
    next_row = ws.max_row + 1
    for i, data in enumerate(rows_data):
        r = next_row + i
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            src  = ws.cell(row=style_ref_row, column=c_idx)
            copy_fmt(src, cell)

def insert_rows_at(ws, row_idx, rows_data, style_ref_row):
    """row_idx 에 rows_data 삽입 (위에서 아래로)"""
    ws.insert_rows(row_idx, amount=len(rows_data))
    for i, data in enumerate(rows_data):
        r = row_idx + i
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            src  = ws.cell(row=style_ref_row + len(rows_data), column=c_idx)
            copy_fmt(src, cell)

# 12030038: row115에 삽입 (현재 row115 = '🟣 재질 : 천' 섹션 헤더 전)
# 스타일 참조 → row112 (12030035 첫 행)
insert_rows_at(ws3, 115, NEW_38, style_ref_row=112)
print("[STEP 1a] 12030038 가죽 신발(드랍) — row115에 3행 삽입 ✅")

# 12030039: sheet 끝에 추가 (삽입 후 max_row 갱신됨)
# 스타일 참조 → 현재 마지막 신발 CID행 (12030036 → 삽입 후 row128)
append_rows(ws3, NEW_39, style_ref_row=ws3.max_row - 2)
print(f"[STEP 1b] 12030039 천 신발(드랍) — row{ws3.max_row - 2}~{ws3.max_row} 추가 ✅")

# ══════════════════════════════════════════════════════════════
# STEP 2 — 아이템 리스트 전체 갱신
# ══════════════════════════════════════════════════════════════

# ── 2-A: 방어구 시트 전체 파싱 (CID → passives list) ──────────
def parse_passives(ws):
    result = {}
    cur = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        cid, ko, key = row[8], row[1], row[2]
        if isinstance(cid, int):
            cur = cid
            result[cur] = []
        if cur is not None and ko and key:
            result[cur].append((str(key).strip(), str(ko).strip()))
    return result

all_passives = {}
for sn in ["1T 방어구", "2T 방어구", "3T 방어구"]:
    all_passives.update(parse_passives(wb[sn]))
print(f"\n[STEP 2a] 방어구 시트 패시브 파싱: 총 {len(all_passives)}개 CID")

# ── 2-B: CSV 파싱 (CID → etype/proc/sr) ──────────────────────
CSV_PATH = "기획서/[X7] 성장_경제 밸런스 - 방어구 아이템 리스트 .csv"
csv_by_cid  = {}  # {cid: {etype, proc, sr}}
csv_2t_list = []  # 2T items in CSV row order

with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    for row in csv.DictReader(f):
        cid   = int(row["장비 Cid"])
        entry = {
            "etype": row.get("효과 타입",   "").strip(),
            "proc":  row.get("효과 (Proc)", "").strip(),
            "sr":    row.get("SR (Active)", "").strip(),
        }
        csv_by_cid[cid] = entry
        if 12020000 < cid < 12030000:
            csv_2t_list.append(entry)

print(f"[STEP 2b] CSV 파싱: 1T={sum(1 for c in csv_by_cid if 12010000<c<12020000)}, "
      f"2T={len(csv_2t_list)}, 3T={sum(1 for c in csv_by_cid if 12030000<c<12040000)}")

# ── 2-C: 아이템 리스트 셀 업데이트 ───────────────────────────
ws_list = wb["아이템 리스트"]
xlsx_2t_pos = 0  # 2T 위치 카운터

for ri in range(2, ws_list.max_row + 1):
    cid = ws_list.cell(row=ri, column=1).value
    if not isinstance(cid, int):
        continue

    passives = all_passives.get(cid, [])
    n = len(passives)

    # ── 패시브 포맷 결정 ──────────────────────────
    if n == 0:
        p1_val = "⚠️ 없음"
        p2_val = "-"
    elif n == 1:
        p1_val = fmt_passive(*passives[0])
        p2_val = "-"
    elif n == 2:
        p1_val = fmt_passive(*passives[0]) + fmt_passive(*passives[1])
        p2_val = "-"
    else:  # 3개
        p1_val = fmt_passive(*passives[0]) + fmt_passive(*passives[1])
        p2_val = fmt_passive(*passives[2])

    # ── Proc/SR 결정 ──────────────────────────────
    if 12010000 < cid < 12020000:          # 1T: CID 직접
        src = csv_by_cid.get(cid, {})
    elif 12020000 < cid < 12030000:        # 2T: 위치 순서
        src = csv_2t_list[xlsx_2t_pos] if xlsx_2t_pos < len(csv_2t_list) else {}
        xlsx_2t_pos += 1
    else:                                  # 3T: CID 직접
        src = csv_by_cid.get(cid, {})

    etype = src.get("etype", "") or None
    proc  = src.get("proc",  "") or None
    sr    = src.get("sr",    "") or None
    # "-" → None 정리
    proc  = None if proc  == "-" else proc
    sr    = None if sr    == "-" else sr

    # ── 셀 쓰기 ───────────────────────────────────
    ref = ws_list.cell(row=ri, column=1)
    vals = [p1_val, p2_val, etype, proc, sr]
    for ci, val in zip(range(4, 9), vals):
        cell = ws_list.cell(row=ri, column=ci, value=val)
        copy_fmt(ref, cell)

print(f"[STEP 2c] 아이템 리스트 업데이트 완료 (2T 위치 매칭: {xlsx_2t_pos}개)")

# ══════════════════════════════════════════════════════════════
# STEP 3 — 저장
# ══════════════════════════════════════════════════════════════
TMP = "기획서/X7_방어구_스킬_tmp.xlsx"
wb.save(TMP)
try:
    os.replace(TMP, "기획서/X7_방어구_스킬.xlsx")
    print("\n✅ 저장: X7_방어구_스킬.xlsx")
    LOAD = "기획서/X7_방어구_스킬.xlsx"
except PermissionError:
    print(f"\n⚠️  Excel 열려있음 → 임시파일: {TMP}")
    LOAD = TMP

# ══════════════════════════════════════════════════════════════
# STEP 4 — 최종 검증
# ══════════════════════════════════════════════════════════════
wb2 = load_workbook(LOAD)
ws2 = wb2["아이템 리스트"]

print("\n=== 최종 아이템 리스트 검증 ===")
no_passive = no_sr = 0
t1=t2=t3=0

for row in ws2.iter_rows(min_row=2, values_only=True):
    cid = row[0]
    if not isinstance(cid, int):
        continue
    p1, p2, etype, proc, sr = (row[i] if i < len(row) else None for i in range(3, 8))
    if 12010000 < cid < 12020000: t1+=1
    elif 12020000 < cid < 12030000: t2+=1
    else: t3+=1
    if not p1 or "없음" in str(p1): no_passive += 1
    if not sr: no_sr += 1

print(f"  아이템 수:   1T={t1}  2T={t2}  3T={t3}  합계={t1+t2+t3}")
print(f"  패시브 없음: {no_passive}개")
print(f"  SR 없음:     {no_sr}개")

# 12030038/12030039 확인
print("\n  [신규 추가 확인]")
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0] in (12030038, 12030039):
        print(f"    CID={row[0]}  이름={row[1]:25s}  P1={str(row[3])[:50]}  P2={str(row[4])[:30]}")

# 3T 방어구 시트 행 수 확인
ws3_v = wb2["3T 방어구"]
count_3t = sum(1 for r in ws3_v.iter_rows(min_row=3, values_only=True)
               if isinstance(r[8], int))
print(f"\n  [3T 방어구 시트] CID 행 수: {count_3t}개")

print("\n  [전체 샘플 (3T 드랍 신발)]")
for row in ws2.iter_rows(min_row=2, values_only=True):
    cid = row[0]
    if isinstance(cid, int) and 12030030 <= cid <= 12030039:
        print(f"    {cid}  {str(row[1]):28s}  P1={str(row[3])[:45]:45s}  SR={row[7]}")
