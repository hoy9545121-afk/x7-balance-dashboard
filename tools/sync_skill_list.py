"""
스킬 리스트 시트 동기화:
  1. 활 3번 스킬셋: 미구현 → 실제 구현 스킬 (GAME_SPEC.md / 활 시트 기준)
  2. 단검 3번 스킬셋: 스킬명 + 비고 → 단검 시트 / constants.py 기준으로 갱신
"""
import openpyxl

FILEPATH = r'C:\AI_simulator\기획서\X7_무기_스킬.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['스킬 리스트']

# ─── 업데이트 데이터 ─────────────────────────────────────────────────────────
# (장비구분, 커맨드, GroupId, 스킬명, 데이터상태, 비고)

# 활 3번 스킬셋: 미구현 → 구현 완료
# GroupId: 활 시트 row73 노트 "GroupId 300001~300155" 기준
# 스킬 데이터: constants.py / GAME_SPEC.md 기준
bow_set3 = [
    ('3번 스킬셋 활', 'Q', '300001', '충격 사격',   '✅ 전 레벨 확보', 'cd=5~3s/MP=10~15/dmg=220~308% (치명타)'),
    ('3번 스킬셋 활', 'W', '300301', '연속 쏘기',   '✅ 전 레벨 확보', 'cd=12~8s/MP=35~50/dmg=280~392%'),
    ('3번 스킬셋 활', 'E', '300151', '급습',        '✅ 전 레벨 확보', 'cd=20~16s/MP=55~75/공속+50~70%/dmg=350~490%'),
    ('3번 스킬셋 활', 'R', '300051', '발묶음',      '✅ 전 레벨 확보', 'cd=45~37s/MP=180~250/CC(속박)+dmg=500~700%'),
]

# 단검 3번 스킬셋: 스킬명 갱신 (단검 시트 / constants.py 기준)
# GroupId: 단검 시트 내 root_id 주석 기준 (500601~500751)
dagger_set3 = [
    ('3번 스킬셋 단검', 'Q', '500601', '추격자의 발톱', '✅ 전 레벨 확보', 'cd=6~4s/MP=50~70/dmg=200~280%'),
    ('3번 스킬셋 단검', 'W', '500651', '침묵의 일격',   '✅ 전 레벨 확보', 'cd=13~14s/MP=20~30/버프(공속+35%/5~6s)'),
    ('3번 스킬셋 단검', 'E', '500701', '사방의 비수',   '✅ 전 레벨 확보', 'cd=15~11s/MP=40~55/dmg=250~350%'),
    ('3번 스킬셋 단검', 'R', '500751', '유혈의 장막',   '✅ 전 레벨 확보', 'cd=45~37s/MP=150~210/dmg=450~630%'),
]

# ─── 헬퍼: 행 찾기 ──────────────────────────────────────────────────────────
def find_rows(ws, col_b_value, col_c_value):
    """Col B == col_b_value 이고 Col C == col_c_value 인 행 번호 반환."""
    results = []
    for row in ws.iter_rows():
        b = ws.cell(row=row[0].row, column=2).value
        c = ws.cell(row=row[0].row, column=3).value
        if b == col_b_value and c == col_c_value:
            results.append(row[0].row)
    return results


def write_row(ws, row_num, data):
    """(장비구분, 커맨드, GroupId, 스킬명, 데이터상태, 비고) → B~G 열에 쓰기."""
    장비, cmd, gid, name, status, note = data
    ws.cell(row=row_num, column=2).value = 장비
    ws.cell(row=row_num, column=3).value = cmd
    ws.cell(row=row_num, column=4).value = gid
    ws.cell(row=row_num, column=5).value = name
    ws.cell(row=row_num, column=6).value = status
    ws.cell(row=row_num, column=7).value = note


# ─── 활 3번 스킬셋 업데이트 ─────────────────────────────────────────────────
print("=== 활 3번 스킬셋 업데이트 ===")
for entry in bow_set3:
    장비, cmd, *_ = entry
    rows = find_rows(ws, 장비, cmd)
    if rows:
        write_row(ws, rows[0], entry)
        print(f"  Row {rows[0]}: {cmd} → {entry[3]} ✅")
    else:
        print(f"  ⚠ 행 없음: {장비} / {cmd}")

# ─── 단검 3번 스킬셋 업데이트 ───────────────────────────────────────────────
print("\n=== 단검 3번 스킬셋 업데이트 ===")
for entry in dagger_set3:
    장비, cmd, *_ = entry
    rows = find_rows(ws, 장비, cmd)
    if rows:
        write_row(ws, rows[0], entry)
        print(f"  Row {rows[0]}: {cmd} → {entry[3]} ✅")
    else:
        print(f"  ⚠ 행 없음: {장비} / {cmd}")

# ─── 저장 ────────────────────────────────────────────────────────────────────
wb.save(FILEPATH)
print("\n✅ 저장 완료")

# ─── 결과 검증 ───────────────────────────────────────────────────────────────
print("\n=== 검증: 스킬 리스트 단검/활 3번 스킬셋 ===")
wb2 = openpyxl.load_workbook(FILEPATH)
ws2 = wb2['스킬 리스트']
for r in range(47, 63):
    row_data = [(c.column, c.value) for c in ws2[r] if c.value is not None]
    if row_data:
        print(f"  Row {r}: {row_data}")
