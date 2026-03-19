"""
기획서 Excel 파일 감사 스크립트
현재 Excel 내용을 constants.py 기준으로 비교·출력
"""
import openpyxl
import sys

BASE = r'C:\AI_simulator\기획서'

def print_sheet_names(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n{'='*60}")
    print(f"파일: {path.split(chr(92))[-1]}")
    print(f"시트: {wb.sheetnames}")
    wb.close()

def dump_sheet(path, sheet_name, max_rows=60):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ 시트 없음: {sheet_name}")
        wb.close()
        return
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} (max {max_rows}행) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > max_rows:
            print("  ... (이하 생략)")
            break
        if any(v is not None for v in row):
            print(f"  R{i}: {list(row)}")
    wb.close()

files = {
    '무기_스킬': fr'{BASE}\X7_무기_스킬.xlsx',
    '장비_능력치': fr'{BASE}\X7_장비_능력치_기획서.xlsx',
    '성장_경제': fr'{BASE}\X7_성장_경제_밸런스_통합기획서.xlsx',
    '방어구_스킬': fr'{BASE}\X7_방어구_스킬.xlsx',
    '제작_드랍': fr'{BASE}\X7_장비_제작_및_드랍.xlsx',
    '몬스터_카드': fr'{BASE}\X7_몬스터_카드_아이템_목록.xlsx',
}

# 시트 이름 먼저 확인
for k, p in files.items():
    print_sheet_names(p)

# 주요 시트 내용 확인
target = '무기_스킬'
path = files[target]
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"\n\n{'='*60}")
print(f"=== {target} 전체 시트 목록: {wb.sheetnames} ===")
wb.close()

# 스킬 데이터 시트 확인
for sheet in ['양손검', '한손검', '활', '지팡이', '단검']:
    dump_sheet(files['무기_스킬'], sheet, max_rows=80)

# 장비 능력치 시트 확인
dump_sheet(files['장비_능력치'], '장비 능력치', max_rows=80)

# 몬스터 시트 확인
for sheet in ['몬스터 스탯', '몬스터']:
    dump_sheet(files['장비_능력치'], sheet, max_rows=40)
    dump_sheet(files['성장_경제'], sheet, max_rows=40)
